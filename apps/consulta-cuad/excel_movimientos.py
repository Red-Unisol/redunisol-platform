import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PATRON_NDJSON = "resultados_cuad_main2*.ndjson"

ENCABEZADOS = [
    "Empleador",
    "CUIL consultado",
    "Apellido y nombre",
    "DNI",
    "Organizacion",
    "Jurisdiccion",
    "Entidad empleado",
    "Nro afiliado",
    "Organismo",
    "Sector",
    "Entidad movimiento",
    "Cupo",
    "Afectado",
    "% Afectado",
    "PreCancelado",
    "% PreCancelado",
    "Deuda",
]


def parsear_argumentos():
    parser = argparse.ArgumentParser(
        description="Exporta la tabla de movimientos desde un NDJSON de resultados."
    )
    parser.add_argument(
        "archivo_ndjson",
        nargs="?",
        type=Path,
        help="Ruta al archivo .ndjson. Si se omite, toma el mas reciente en corridas/.",
    )
    return parser.parse_args()


def buscar_ndjson_mas_reciente():
    candidatos = [
        path
        for path in Path("corridas").glob(f"*/{PATRON_NDJSON}")
        if path.is_file()
    ]

    if not candidatos:
        return None

    return max(candidatos, key=lambda path: path.stat().st_mtime)


def resolver_archivo_ndjson(path_argumento):
    if path_argumento is not None:
        return path_argumento

    return buscar_ndjson_mas_reciente()


def leer_ndjson(path):
    with path.open("r", encoding="utf-8-sig") as archivo:
        for linea in archivo:
            linea = linea.strip()
            if not linea:
                continue
            yield json.loads(linea)


def cargar_ultimos_registros(path):
    ultimos = {}

    for registro in leer_ndjson(path):
        cuil = texto(registro.get("cuil"))

        if cuil:
            ultimos.pop(cuil, None)
            ultimos[cuil] = registro

    return list(ultimos.values())


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def construir_filas_movimientos(registro):
    payload = registro.get("payload", {})
    parsed = registro.get("parsed", {})
    tabla_empleado = parsed.get("tabla_empleado", {})
    tabla_movimientos = parsed.get("tabla_movimientos", {})
    movimientos = tabla_movimientos.get("registros", [])

    empleador = texto(registro.get("emr_nombre")) or texto(payload.get("Emr_Nombre"))

    filas = []
    for movimiento in movimientos:
        filas.append(
            [
                empleador,
                texto(registro.get("cuil")),
                texto(tabla_empleado.get("apellido_nombre")),
                texto(tabla_empleado.get("documento")),
                texto(tabla_empleado.get("organizacion")),
                texto(tabla_empleado.get("jurisdiccion_codigo")),
                texto(tabla_empleado.get("entidad")),
                texto(tabla_empleado.get("nro_afiliado")),
                texto(movimiento.get("Organismo")),
                texto(movimiento.get("Sector")),
                texto(movimiento.get("Entidad")),
                texto(movimiento.get("Cupo")),
                texto(movimiento.get("Afectado")),
                texto(movimiento.get("%_1")),
                texto(movimiento.get("PreCancelado")),
                texto(movimiento.get("%_2")),
                texto(movimiento.get("Deuda")),
            ]
        )

    return filas


def ajustar_anchos(ws):
    for columna in ws.columns:
        largo_maximo = 0
        indice = columna[0].column

        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            if len(valor) > largo_maximo:
                largo_maximo = len(valor)

        ws.column_dimensions[get_column_letter(indice)].width = max(largo_maximo + 2, 12)


def main():
    args = parsear_argumentos()
    archivo_ndjson = resolver_archivo_ndjson(args.archivo_ndjson)

    if archivo_ndjson is None:
        print("No se encontro ningun archivo NDJSON en corridas/.")
        return

    if not archivo_ndjson.exists():
        print(f"No existe el archivo: {archivo_ndjson}")
        return

    archivo_salida = archivo_ndjson.with_name(
        f"{archivo_ndjson.stem}_movimientos.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    ws.append(ENCABEZADOS)

    registros = cargar_ultimos_registros(archivo_ndjson)

    cantidad_registros_leidos = 0
    cantidad_consultas_exportadas = 0
    cantidad_filas_exportadas = 0

    for registro in registros:
        cantidad_registros_leidos += 1

        if texto(registro.get("status")).lower() != "ok":
            continue

        filas_movimiento = construir_filas_movimientos(registro)

        if not filas_movimiento:
            continue

        for fila in filas_movimiento:
            ws.append(fila)
            cantidad_filas_exportadas += 1

        cantidad_consultas_exportadas += 1

    ws.freeze_panes = "A2"
    ajustar_anchos(ws)

    archivo_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(archivo_salida)

    print(f"Archivo generado: {archivo_salida}")
    print(f"Registros leidos: {cantidad_registros_leidos}")
    print(f"Consultas exportadas: {cantidad_consultas_exportadas}")
    print(f"Filas de movimientos exportadas: {cantidad_filas_exportadas}")


if __name__ == "__main__":
    main()
