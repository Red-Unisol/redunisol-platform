import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

SRC_DIR = Path(__file__).resolve().parents[1] / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from consulta_cuad import exportar  # noqa: E402


REGISTRO = {
    "cuil": "20111111112",
    "status": "ok",
    "emr_nombre": "Santa Fe - ACTIVOS",
    "parsed": {
        "tabla_empleado": {"apellido_nombre": "PEREZ JUAN"},
        "tabla_totales": {"cupo": "40"},
        "tabla_movimientos": {"registros": [{"Organismo": "ORG", "Deuda": "5"}]},
    },
}


class TestExportar(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.archivo = Path(self.temp.name) / "resultados_socios.ndjson"
        self.archivo.write_text(json.dumps(REGISTRO) + "\n", encoding="utf-8")

    def test_exporta_totales_a_xlsx(self):
        salida, leidos, filas = exportar.exportar(self.archivo, "totales")
        hoja = load_workbook(salida).active
        self.assertEqual((leidos, filas), (1, 1))
        self.assertEqual(hoja.title, "Totales")
        self.assertEqual(hoja.cell(2, 11).value, "40")

    def test_exporta_movimientos_a_xlsx(self):
        salida, leidos, filas = exportar.exportar(self.archivo, "movimientos")
        hoja = load_workbook(salida).active
        self.assertEqual((leidos, filas), (1, 1))
        self.assertEqual(hoja.title, "Movimientos")
        self.assertEqual(hoja.cell(2, 9).value, "ORG")
        self.assertEqual(hoja.cell(2, 17).value, "5")
