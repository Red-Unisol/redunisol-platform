"""De donde salen los CUILes cuando no se consultan a Vimarx.

Lee un Excel (.xlsx / .xlsm) o un JSON con una lista, y devuelve siempre la
misma forma: un diccionario con los CUILes validos y el detalle de lo que se
descarto.

POR QUE DEVUELVE TANTO DETALLE
------------------------------
Una corrida de 1000 socios tarda horas. Si el Excel tenia una columna
equivocada, o la mitad de los CUILes venian mal, hay que enterarse ANTES de
arrancar y no a las tres horas. Por eso se cuentan filas leidas, unicos e
invalidos, y se guarda una muestra de los invalidos con su numero de fila para
poder ir a mirarlos.
"""

import json
import logging
import re
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

LARGO_CUIL = 11
EXTENSIONES_EXCEL = {".xlsx", ".xlsm"}

# Encabezados que se reconocen solos, sin que haya que indicar la columna.
ENCABEZADOS_CUIL = {
    "cuil",
    "cuit",
    "cuil/cuit",
    "cuit/cuil",
    "nro cuil",
    "nro cuit",
}


def normalizar_cuil(valor):
    """Deja solo los digitos.

    El `.0` del final aparece cuando openpyxl lee la celda como numero: un
    CUIL de 11 digitos entra en un float y vuelve como 20111111112.0.
    """
    if valor is None:
        return ""

    texto = str(valor).strip()
    if not texto:
        return ""

    if texto.endswith(".0"):
        texto = texto[:-2]

    return re.sub(r"\D", "", texto)


def normalizar_encabezado(valor):
    return re.sub(r"\s+", " ", str(valor or "").strip()).lower()


def indice_columna_excel(columna):
    """Convierte "3" o "C" en un indice 0-based. Devuelve None si no aplica."""
    if columna is None:
        return None

    columna = str(columna).strip()

    if not columna:
        return None

    if columna.isdigit():
        indice = int(columna) - 1
        if indice < 0:
            raise ValueError("La columna numerica debe ser 1 o mayor.")
        return indice

    if re.fullmatch(r"[A-Za-z]+", columna):
        indice = 0
        for caracter in columna.upper():
            indice = indice * 26 + (ord(caracter) - ord("A") + 1)
        return indice - 1

    return None


def resolver_columna_cuil(encabezados, columna_cuiles=None):
    """Decide en que columna estan los CUILes.

    Si se indico una, se busca primero por nombre de encabezado y recien
    despues como letra o numero de columna. Ese orden importa: en una planilla
    con una columna llamada "C" gana el encabezado, que es lo que la persona
    quiso decir.
    """
    if columna_cuiles:
        objetivo = normalizar_encabezado(columna_cuiles)
        for indice, encabezado in enumerate(encabezados):
            if normalizar_encabezado(encabezado) == objetivo:
                return indice

        indice = indice_columna_excel(columna_cuiles)
        if indice is not None:
            if indice >= len(encabezados):
                raise ValueError(f"La columna {columna_cuiles} no existe en la hoja.")
            return indice

        raise ValueError(f"No se encontro la columna '{columna_cuiles}' en la hoja.")

    for indice, encabezado in enumerate(encabezados):
        if normalizar_encabezado(encabezado) in ENCABEZADOS_CUIL:
            return indice

    raise ValueError(
        "No se encontro una columna CUIL/CUIT automatica. "
        "Indicala con --columna-cuiles."
    )


def _resultado(origen, cuiles, invalidos, filas_leidas, hoja=None, columna=None):
    return {
        "origen": str(origen),
        "hoja": hoja,
        "columna": columna,
        "cantidad_filas_leidas": filas_leidas,
        "cantidad_cuiles_unicos": len(cuiles),
        "cantidad_invalidos": len(invalidos),
        "invalidos": invalidos[:20],
        "cuiles": cuiles,
    }


def _recolectar(valores, origen, hoja=None, columna=None, primera_fila=2):
    """Normaliza, valida y deduplica, conservando el orden de aparicion."""
    cuiles = []
    vistos = set()
    invalidos = []
    leidas = 0

    for numero, valor in enumerate(valores, start=primera_fila):
        leidas += 1
        cuil = normalizar_cuil(valor)

        if not cuil:
            continue

        if len(cuil) != LARGO_CUIL:
            invalidos.append({"fila": numero, "valor": str(valor).strip()})
            continue

        if cuil in vistos:
            continue

        vistos.add(cuil)
        cuiles.append(cuil)

    return _resultado(origen, cuiles, invalidos, leidas, hoja, columna)


def cargar_desde_excel(path, hoja_cuiles=None, columna_cuiles=None):
    path = Path(path)

    # read_only deja el archivo abierto hasta que se cierra el libro, y en
    # Windows eso lo BLOQUEA: la persona no puede volver a abrir su propia
    # planilla en Excel ni borrarla. De ahi el try/finally.
    #
    # Como en modo read_only las filas se leen de forma perezosa, hay que
    # materializar los valores ANTES de cerrar.
    libro = load_workbook(path, read_only=True, data_only=True)
    try:
        if hoja_cuiles:
            if hoja_cuiles not in libro.sheetnames:
                raise ValueError(f"La hoja '{hoja_cuiles}' no existe en {path.name}.")
            hoja = libro[hoja_cuiles]
        else:
            hoja = libro[libro.sheetnames[0]]

        nombre_hoja = hoja.title
        filas = hoja.iter_rows(values_only=True)
        encabezados = next(filas, None)

        if encabezados is None:
            raise ValueError(f"La hoja '{nombre_hoja}' esta vacia.")

        encabezados = [
            str(valor).strip() if valor is not None else "" for valor in encabezados
        ]
        indice = resolver_columna_cuil(encabezados, columna_cuiles)

        valores = [fila[indice] if indice < len(fila) else None for fila in filas]
    finally:
        libro.close()

    return _recolectar(
        valores,
        origen=path,
        hoja=nombre_hoja,
        columna=encabezados[indice] or f"columna_{indice + 1}",
    )


def cargar_desde_json(path):
    """Lee una lista de CUILes ya guardada.

    Es el formato en el que la propia app deja los CUILes de una corrida, asi
    que sirve para repetir exactamente la misma lista.
    """
    path = Path(path)
    datos = json.loads(path.read_text(encoding="utf-8"))

    if not isinstance(datos, list):
        raise ValueError(
            f"{path.name} tiene que contener una lista de CUILes, "
            f"y contiene {type(datos).__name__}."
        )

    return _recolectar(datos, origen=path, columna="json", primera_fila=1)


def cargar(path, hoja_cuiles=None, columna_cuiles=None):
    """Elige el lector segun la extension."""
    path = Path(path).expanduser()

    if not path.exists():
        raise ValueError(f"No existe el archivo de CUILes: {path}")

    sufijo = path.suffix.lower()

    if sufijo in EXTENSIONES_EXCEL:
        resultado = cargar_desde_excel(path, hoja_cuiles, columna_cuiles)
    elif sufijo == ".json":
        resultado = cargar_desde_json(path)
    else:
        raise ValueError(
            f"Formato no soportado: '{sufijo}'. "
            f"Usa un Excel ({', '.join(sorted(EXTENSIONES_EXCEL))}) o un .json."
        )

    if not resultado["cuiles"]:
        raise ValueError(f"{path.name} no contiene ningun CUIL valido para consultar.")

    logger.info(
        "%s: %s CUILes unicos de %s filas (%s invalidos).",
        path.name,
        resultado["cantidad_cuiles_unicos"],
        resultado["cantidad_filas_leidas"],
        resultado["cantidad_invalidos"],
    )

    return resultado
