import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PATRON_NDJSON = "resultados_cuad_main2*.ndjson"

ENCABEZADOS = [
    "Empleador",
    "CUIL consultado",
    "Apellido y nombre",
    "DNI",
    "Organizacion",
    "Jurisdiccion",
    "Entidad",
    "Nro afiliado",
    "Bruto",
    "Neto",
    "Cupo",
    "Afectado",
    "% Afectado",
    "PreCancelado",
    "% PreCancelado",
    "Disponible",
    "% Disponible",
    "Deuda",
]


def parsear_argumentos():
    parser = argparse.ArgumentParser(
        description="Exporta la tabla de totales desde un NDJSON de resultados."
    )
    parser.add_argument(
        "archivo_ndjson",
        nargs="?",
        type=Path,
        help="Ruta al archivo .ndjson. Si se omite, toma el mas reciente en corridas/.",
    )
    return parser.parse_args()


def buscar_ndjson_mas_reciente():
    candidatos = [
        path
        for path in Path("corridas").glob(f"*/{PATRON_NDJSON}")
        if path.is_file()
    ]

    if not candidatos:
        return None

    return max(candidatos, key=lambda path: path.stat().st_mtime)


def resolver_archivo_ndjson(path_argumento):
    if path_argumento is not None:
        return path_argumento

    return buscar_ndjson_mas_reciente()


def leer_ndjson(path):
    with path.open("r", encoding="utf-8-sig") as archivo:
        for linea in archivo:
            linea = linea.strip()
            if not linea:
                continue
            yield json.loads(linea)


def cargar_ultimos_registros(path):
    ultimos = {}

    for registro in leer_ndjson(path):
        cuil = texto(registro.get("cuil"))

        if cuil:
            ultimos.pop(cuil, None)
            ultimos[cuil] = registro

    return list(ultimos.values())


def texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def construir_fila(registro):
    payload = registro.get("payload") or {}
    parsed = registro.get("parsed") or {}
    tabla_empleado = parsed.get("tabla_empleado") or {}
    tabla_totales = parsed.get("tabla_totales") or {}

    empleador = texto(registro.get("emr_nombre")) or texto(payload.get("Emr_Nombre"))

    return [
        empleador,
        texto(registro.get("cuil")),
        texto(tabla_empleado.get("apellido_nombre")),
        texto(tabla_empleado.get("documento")),
        texto(tabla_empleado.get("organizacion")),
        texto(tabla_empleado.get("jurisdiccion_codigo")),
        texto(tabla_empleado.get("entidad")),
        texto(tabla_empleado.get("nro_afiliado")),
        texto(tabla_totales.get("bruto")),
        texto(tabla_totales.get("neto")),
        texto(tabla_totales.get("cupo")),
        texto(tabla_totales.get("afectado")),
        texto(tabla_totales.get("afectado_porcentaje")),
        texto(tabla_totales.get("precancelado")),
        texto(tabla_totales.get("precancelado_porcentaje")),
        texto(tabla_totales.get("disponible")),
        texto(tabla_totales.get("disponible_porcentaje")),
        texto(tabla_totales.get("deuda")),
    ]


def ajustar_anchos(ws):
    for columna in ws.columns:
        largo_maximo = 0
        indice = columna[0].column

        for celda in columna:
            valor = "" if celda.value is None else str(celda.value)
            if len(valor) > largo_maximo:
                largo_maximo = len(valor)

        ws.column_dimensions[get_column_letter(indice)].width = max(largo_maximo + 2, 12)


def main():
    args = parsear_argumentos()
    archivo_ndjson = resolver_archivo_ndjson(args.archivo_ndjson)

    if archivo_ndjson is None:
        print("No se encontro ningun archivo NDJSON en corridas/.")
        return

    if not archivo_ndjson.exists():
        print(f"No existe el archivo: {archivo_ndjson}")
        return

    archivo_salida = archivo_ndjson.with_name(f"{archivo_ndjson.stem}_totales.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Totales"

    ws.append(ENCABEZADOS)

    registros = cargar_ultimos_registros(archivo_ndjson)

    cantidad_leidas = 0
    cantidad_exportadas = 0

    for registro in registros:
        cantidad_leidas += 1

        if texto(registro.get("status")).lower() != "ok":
            continue

        ws.append(construir_fila(registro))
        cantidad_exportadas += 1

    ws.freeze_panes = "A2"
    ajustar_anchos(ws)

    archivo_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(archivo_salida)

    print(f"Archivo generado: {archivo_salida}")
    print(f"Registros leidos: {cantidad_leidas}")
    print(f"Registros exportados: {cantidad_exportadas}")


if __name__ == "__main__":
    main()
