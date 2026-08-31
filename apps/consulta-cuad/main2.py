import argparse
import html
import json
import re
import time
import unicodedata
from collections import OrderedDict
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from openpyxl import load_workbook
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

URL_VIMARX = "https://celesol.dyndns.org:5002/api/Empresa/EvaluateList"
URL_CUAD = "https://www.santafe.gov.ar/cuad/movimiento.asp"
URL_CUAD_GRILLA = "https://www.santafe.gov.ar/cuad/grilla.asp"

EMR_NOMBRE_ACTIVOS = "Santa Fe - ACTIVOS"
EMR_ID_ACTIVOS = "10"
EMR_NOMBRE_PASIVOS = "Santa Fe - PASIVOS"
EMR_ID_PASIVOS = "11"

DEFAULT_LINEA_SUPERIOR_VIMARX = "Haberes CUAD SANTA FE"
DEFAULT_ESTADO_VIMARX = "Activa"
BASE_NOMBRE_CUILES_VIMARX = "cuiles_vimarx_main2"
BASE_NOMBRE_CUILES_ARCHIVO = "cuiles_archivo_main2"
BASE_NOMBRE_RESULTADOS = "resultados_cuad_main2"

CAMPOS_VIMARX = (
    "ID;NroCuenta;LineaPrestamo.Descripcion;SocioTitular.Socio.NombreCompleto;"
    "SocioTitular.Socio.CUIT;DeudaVencidaConPunitorios;DeudaTotalVencida"
)
TIPO_VIMARX = "F.Module.Cuentas.Prestamos.Prestamo"

TIMEOUT_VIMARX = 180
TIMEOUT_CUAD = 180
LIMITE_BLOQUE = 500
TAMANO_VENTANA_ID = 3000
DEMORA_ENTRE_CONSULTAS = 12
PAUSA_CADA = 50
PAUSA_LARGA_SEGUNDOS = 180
MAX_INTENTOS_CUAD = 3
PAUSA_REINTENTO_CUAD_SEGUNDOS = 15

PERIODO_CONSULTA = datetime.now().strftime("%Y-%m")
DIRECTORIO_CORRIDAS = Path("corridas")
DIRECTORIO_PERIODO = DIRECTORIO_CORRIDAS / PERIODO_CONSULTA
MODO_CARGA_CUILES = "vimarx"
LINEA_SUPERIOR_VIMARX = DEFAULT_LINEA_SUPERIOR_VIMARX
ESTADO_VIMARX = DEFAULT_ESTADO_VIMARX
BASE_CMD_VIMARX = (
    f"[LineaPrestamo.Superior.Descripcion] = '{LINEA_SUPERIOR_VIMARX}' "
    f"AND [Estado] = '{ESTADO_VIMARX}'"
)
ARCHIVO_CUILES_FUENTE = DIRECTORIO_PERIODO / f"{BASE_NOMBRE_CUILES_VIMARX}.json"
ARCHIVO_RESULTADOS_CUAD = DIRECTORIO_PERIODO / f"{BASE_NOMBRE_RESULTADOS}.ndjson"
ARCHIVO_CUILES_ORIGEN = None
HOJA_CUILES = None
COLUMNA_CUILES = None

INICIAR_NUEVA_CORRIDA = False
LIMITE_CUAD = None

COOKIE_CUAD = ""


NOMBRES_SET_EMPLEADO = [
    "apellido_nombre",
    "tipo_documento",
    "organizacion",
    "jurisdiccion_nombre",
    "jurisdiccion_codigo",
    "entidad",
    "nro_afiliado",
    "nro_cuil",
    "documento",
    "es_socio",
    "suspendido",
]

NOMBRES_SET_TOTALES = [
    "bruto",
    "neto",
    "cupo",
    "afectado",
    "afectado_porcentaje",
    "precancelado",
    "precancelado_porcentaje",
    "disponible",
    "disponible_porcentaje",
    "deuda",
    "disponible_negativo",
    "disponible_negativo_porcentaje",
    "habilita_cupo_especial",
    "cupo_especial",
    "deuda_especial",
    "afectado_especial",
    "afectado_especial_porcentaje",
    "disponible_especial",
    "disponible_especial_porcentaje",
    "disponible_valor_oculto",
    "disponible_negativo_valor_oculto",
]

ESTADOS_REINTENTABLES_CUAD = {"error_http", "error_conexion", "timeout"}
ESTADOS_PENDIENTES_CUAD = ESTADOS_REINTENTABLES_CUAD | {
    "sesion_invalida",
    "respuesta_no_reconocida",
}


class TablaCuadHTMLParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.active_table = None
        self.current_row = None
        self.current_cell = None
        self.headers = []
        self.rows = []

    def handle_starttag(self, tag, attrs):
        attrs_dict = dict(attrs)

        if tag == "table":
            table_id = attrs_dict.get("id")
            if table_id in {"oTableCab", "oTable"}:
                self.active_table = table_id
            return

        if self.active_table is None:
            return

        if tag == "tr":
            self.current_row = []
        elif tag == "td" and self.current_row is not None:
            self.current_cell = []

    def handle_data(self, data):
        if self.current_cell is not None:
            self.current_cell.append(data)

    def handle_endtag(self, tag):
        if tag == "table":
            self.active_table = None
            return

        if self.active_table is None:
            return

        if tag == "td" and self.current_cell is not None:
            texto = html.unescape("".join(self.current_cell)).replace("\xa0", " ")
            texto = re.sub(r"\s+", " ", texto).strip()
            self.current_row.append(texto)
            self.current_cell = None
        elif tag == "tr" and self.current_row is not None:
            fila = self.current_row
            self.current_row = None

            if self.active_table == "oTableCab":
                self.headers = [celda for celda in fila if celda]
            elif self.active_table == "oTable" and fila:
                self.rows.append(fila)


def ahora_iso():
    return datetime.now().isoformat(timespec="seconds")


def sello_archivo():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def slug_archivo(texto):
    texto_normalizado = unicodedata.normalize("NFKD", str(texto))
    texto_ascii = texto_normalizado.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", texto_ascii).strip("_").lower()
    return slug or "lista"


def escapar_literal_vimarx(valor):
    return str(valor).replace("'", "''")


def construir_base_cmd_vimarx(linea_superior, estado):
    return (
        "[LineaPrestamo.Superior.Descripcion] = "
        f"'{escapar_literal_vimarx(linea_superior)}' "
        f"AND [Estado] = '{escapar_literal_vimarx(estado)}'"
    )


def construir_archivo_periodo(nombre_base, extension, sufijo=""):
    if sufijo:
        return DIRECTORIO_PERIODO / f"{nombre_base}_{sufijo}{extension}"
    return DIRECTORIO_PERIODO / f"{nombre_base}{extension}"


def construir_sufijo_salida(
    linea_superior=None, etiqueta_salida=None, archivo_cuiles=None
):
    if etiqueta_salida:
        return slug_archivo(etiqueta_salida)

    if archivo_cuiles is not None:
        return slug_archivo(Path(archivo_cuiles).stem)

    if str(linea_superior).strip() == DEFAULT_LINEA_SUPERIOR_VIMARX:
        return ""

    return slug_archivo(linea_superior)


def configurar_consulta(
    linea_superior,
    estado,
    etiqueta_salida=None,
    archivo_cuiles=None,
    hoja_cuiles=None,
    columna_cuiles=None,
    limite_cuad=None,
):
    global MODO_CARGA_CUILES
    global LINEA_SUPERIOR_VIMARX
    global ESTADO_VIMARX
    global BASE_CMD_VIMARX
    global ARCHIVO_CUILES_FUENTE
    global ARCHIVO_RESULTADOS_CUAD
    global ARCHIVO_CUILES_ORIGEN
    global HOJA_CUILES
    global COLUMNA_CUILES
    global LIMITE_CUAD

    linea_superior = str(linea_superior).strip()
    estado = str(estado).strip()

    if archivo_cuiles is not None:
        archivo_cuiles = Path(archivo_cuiles).expanduser()

        if not archivo_cuiles.exists():
            raise ValueError(f"No existe el archivo de CUILes: {archivo_cuiles}")

        MODO_CARGA_CUILES = "archivo"
        ARCHIVO_CUILES_ORIGEN = archivo_cuiles
        HOJA_CUILES = str(hoja_cuiles).strip() if hoja_cuiles else None
        COLUMNA_CUILES = str(columna_cuiles).strip() if columna_cuiles else None
        sufijo_salida = construir_sufijo_salida(
            etiqueta_salida=etiqueta_salida,
            archivo_cuiles=archivo_cuiles,
        )
        ARCHIVO_CUILES_FUENTE = construir_archivo_periodo(
            BASE_NOMBRE_CUILES_ARCHIVO, ".json", sufijo_salida
        )
    else:
        if not linea_superior:
            raise ValueError("La linea de Vimarx no puede estar vacia.")

        if not estado:
            raise ValueError("El estado de Vimarx no puede estar vacio.")

        MODO_CARGA_CUILES = "vimarx"
        LINEA_SUPERIOR_VIMARX = linea_superior
        ESTADO_VIMARX = estado
        BASE_CMD_VIMARX = construir_base_cmd_vimarx(linea_superior, estado)
        ARCHIVO_CUILES_ORIGEN = None
        HOJA_CUILES = None
        COLUMNA_CUILES = None
        sufijo_salida = construir_sufijo_salida(linea_superior, etiqueta_salida)
        ARCHIVO_CUILES_FUENTE = construir_archivo_periodo(
            BASE_NOMBRE_CUILES_VIMARX, ".json", sufijo_salida
        )

    ARCHIVO_RESULTADOS_CUAD = construir_archivo_periodo(
        BASE_NOMBRE_RESULTADOS, ".ndjson", sufijo_salida
    )

    if limite_cuad is not None:
        if limite_cuad <= 0:
            raise ValueError("El limite de CUAD debe ser mayor a cero.")
        LIMITE_CUAD = limite_cuad


def guardar_json(datos, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as archivo:
        json.dump(datos, archivo, ensure_ascii=False, indent=2)


def leer_json(path):
    with open(path, "r", encoding="utf-8") as archivo:
        return json.load(archivo)


def append_ndjson(path, registro):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "a", encoding="utf-8") as archivo:
        archivo.write(json.dumps(registro, ensure_ascii=False))
        archivo.write("\n")


def cargar_ultimos_resultados_cuad():
    ultimos = OrderedDict()

    if not ARCHIVO_RESULTADOS_CUAD.exists():
        return ultimos

    with open(ARCHIVO_RESULTADOS_CUAD, "r", encoding="utf-8") as archivo:
        for linea in archivo:
            linea = linea.strip()

            if not linea:
                continue

            resultado = json.loads(linea)
            cuil = str(resultado.get("cuil", "")).strip()

            if not cuil:
                continue

            ultimos.pop(cuil, None)
            ultimos[cuil] = resultado

    return ultimos


def resumir_estado_desde_resultados(ultimos_resultados):
    procesados = []
    cantidad_ok = 0
    cantidad_sin_resultado = 0
    cantidad_errores = 0
    cantidad_pendientes_reintento = 0
    ultimo_cuil = None

    for cuil, resultado in ultimos_resultados.items():
        status = resultado.get("status")
        ultimo_cuil = cuil

        if status in ESTADOS_PENDIENTES_CUAD:
            cantidad_pendientes_reintento += 1
            continue

        procesados.append(cuil)

        if status == "ok":
            cantidad_ok += 1
        elif status == "sin_resultado":
            cantidad_sin_resultado += 1
        else:
            cantidad_errores += 1

    return {
        "procesados": procesados,
        "ultimo_cuil": ultimo_cuil,
        "cantidad_ok": cantidad_ok,
        "cantidad_sin_resultado": cantidad_sin_resultado,
        "cantidad_errores": cantidad_errores,
        "cantidad_pendientes_reintento": cantidad_pendientes_reintento,
        "cantidad_total_procesada": len(procesados),
    }


def respaldar_archivo(path, sello):
    if not path.exists():
        return None

    respaldo = path.with_name(f"{path.stem}_{sello}{path.suffix}")
    indice = 1

    while respaldo.exists():
        respaldo = path.with_name(f"{path.stem}_{sello}_{indice}{path.suffix}")
        indice += 1

    path.replace(respaldo)
    return respaldo


def respaldar_directorio(path, sello):
    if not path.exists():
        return None

    respaldo = path.with_name(f"{path.name}_{sello}")
    indice = 1

    while respaldo.exists():
        respaldo = path.with_name(f"{path.name}_{sello}_{indice}")
        indice += 1

    path.replace(respaldo)
    return respaldo


def preparar_corrida_nueva_si_corresponde():
    DIRECTORIO_PERIODO.mkdir(parents=True, exist_ok=True)

    if not INICIAR_NUEVA_CORRIDA:
        return

    sello = sello_archivo()
    respaldos = []

    for path in (
        ARCHIVO_CUILES_FUENTE,
        ARCHIVO_RESULTADOS_CUAD,
    ):
        respaldo = respaldar_archivo(path, sello)
        if respaldo is not None:
            respaldos.append(respaldo)

    print(f"Preparando corrida nueva del periodo {PERIODO_CONSULTA}.")

    if not respaldos:
        print("No habia archivos previos del periodo para respaldar.")
        return

    for respaldo in respaldos:
        print(f"Respaldo creado: {respaldo.name}")


def mapear_argumentos(nombres, valores):
    return {
        nombre: valores[indice] if indice < len(valores) else None
        for indice, nombre in enumerate(nombres)
    }


def extraer_argumentos_script(html, funcion):
    patron = rf"parent\.{funcion}\((.*?)\)\s*;?\s*</script>"
    coincidencia = re.search(patron, html, re.DOTALL | re.IGNORECASE)

    if not coincidencia:
        return None

    contenido = coincidencia.group(1)
    return re.findall(r"'((?:\\'|[^'])*)'", contenido)


def extraer_valor_script(html, patron, flags=0):
    coincidencia = re.search(patron, html, flags)
    if not coincidencia:
        return None
    return coincidencia.group(1)


def normalizar_columnas_movimientos(columnas):
    columnas_normalizadas = []
    porcentaje_indice = 0

    for columna in columnas:
        if columna == "%":
            porcentaje_indice += 1
            columnas_normalizadas.append(f"%_{porcentaje_indice}")
        else:
            columnas_normalizadas.append(columna)

    return columnas_normalizadas


def parsear_grilla_cuad_script(html_text):
    titulos_match = re.search(
        r"contentWindow\.setTitulos\('((?:\\'|[^'])*)'\s*,\s*([0-9]+)\s*,\s*'((?:\\'|[^'])*)'\)",
        html_text,
        re.IGNORECASE,
    )
    datos_match = re.search(
        r"contentWindow\.setDatos\('((?:\\'|[^'])*)'\s*,\s*'((?:\\'|[^'])*)'\)",
        html_text,
        re.IGNORECASE,
    )

    if not titulos_match and not datos_match:
        return None

    columnas_visibles = (
        titulos_match.group(1).split("|")
        if titulos_match and titulos_match.group(1)
        else []
    )
    columnas_normalizadas = normalizar_columnas_movimientos(columnas_visibles)
    filas_raw = (
        datos_match.group(1).split("~") if datos_match and datos_match.group(1) else []
    )
    registros = []

    for fila_raw in filas_raw:
        if not fila_raw:
            continue

        valores = fila_raw.split("|")
        registro = {}

        if columnas_visibles and len(valores) == len(columnas_visibles) + 1:
            registro["id_oculto"] = valores[0]
            valores_visibles = valores[1:]
        else:
            valores_visibles = valores

        for indice, columna in enumerate(columnas_normalizadas):
            registro[columna] = (
                valores_visibles[indice] if indice < len(valores_visibles) else None
            )

        registros.append(registro)

    return {
        "columnas_visibles": columnas_visibles,
        "columnas_normalizadas": columnas_normalizadas,
        "registros": registros,
        "cantidad_registros": len(registros),
    }


def parsear_respuesta_cuad(html):
    empleado_args = extraer_argumentos_script(html, "setEmpleado")
    totales_args = extraer_argumentos_script(html, "setTotales")
    info_empleado_args = extraer_argumentos_script(html, "setInformacion_empleado")
    precancelado_args = extraer_argumentos_script(html, "setPreCancelado")

    grilla = {
        "seleccion": extraer_valor_script(html, r"bSel\s*=\s*'([^']*)'"),
        "paginacion": extraer_valor_script(html, r"bPag\s*=\s*'([^']*)'"),
        "callback": extraer_valor_script(html, r"sCallBack\s*=\s*'([^']*)'"),
        "identificador": extraer_valor_script(html, r"ID\s*=\s*'([^']*)'"),
        "orden": extraer_valor_script(html, r"Orden\s*=\s*'([^']*)'"),
        "multi_seleccion": extraer_valor_script(html, r"MultiSel\s*=\s*'([^']*)'"),
        "estilo": extraer_valor_script(html, r"Estilo\s*=\s*'([^']*)'"),
        "estilo_fila": extraer_valor_script(html, r"EstiloFila\s*=\s*'([^']*)'"),
        "estilo_columna": extraer_valor_script(html, r"EstiloColumna\s*=\s*'([^']*)'"),
        "formato": extraer_valor_script(html, r"Formato\s*=\s*'([^']*)'"),
        "formato_fila": extraer_valor_script(html, r"FormatoFila\s*=\s*'([^']*)'"),
        "formato_columna": extraer_valor_script(
            html, r"FormatoColumna\s*=\s*'([^']*)'"
        ),
        "paginas": extraer_valor_script(html, r"Pags\s*=\s*([0-9]+)"),
        "registros": extraer_valor_script(html, r"Recs\s*=\s*([0-9]+)"),
        "descripcion_objeto": extraer_valor_script(html, r"Obj_Desc\s*=\s*'([^']*)'"),
        "ver_pie": extraer_valor_script(html, r"VerPie\s*=\s*'([^']*)'"),
        "borde": extraer_valor_script(html, r"Borde\s*=\s*'([^']*)'"),
    }

    tabla_empleado = (
        mapear_argumentos(NOMBRES_SET_EMPLEADO, empleado_args)
        if empleado_args
        else None
    )
    tabla_totales = (
        mapear_argumentos(NOMBRES_SET_TOTALES, totales_args) if totales_args else None
    )
    emp_id = extraer_valor_script(html, r"parent\.Emp_Id\s*=\s*([0-9]+)")
    informacion_empleado = info_empleado_args[0] if info_empleado_args else None
    mensaje_precancelado = precancelado_args[0] if precancelado_args else None

    if tabla_empleado is not None:
        tabla_empleado = dict(tabla_empleado)
        tabla_empleado["emp_id"] = emp_id
        tabla_empleado["informacion_empleado"] = informacion_empleado
        tabla_empleado["mensaje_precancelado"] = mensaje_precancelado

    return {
        "empleado": {
            "raw": empleado_args,
            "parsed": tabla_empleado,
        },
        "totales": {
            "raw": totales_args,
            "parsed": tabla_totales,
        },
        "tabla_empleado": tabla_empleado,
        "tabla_totales": tabla_totales,
        "tabla_movimientos": None,
        "emp_id": emp_id,
        "informacion_empleado": informacion_empleado,
        "mensaje_precancelado": mensaje_precancelado,
        "grilla": grilla,
        "contiene_set_empleado": empleado_args is not None,
        "contiene_set_totales": totales_args is not None,
        "tabla_organismos": parsear_grilla_cuad_html(html),
    }


def parsear_grilla_cuad_html(html_text):
    if "oTableCab" not in html_text or "oTable" not in html_text:
        return None

    parser = TablaCuadHTMLParser()
    parser.feed(html_text)

    columnas = parser.headers
    filas = parser.rows

    if not columnas and not filas:
        return None

    if columnas == [
        "Organismo",
        "Sector",
        "Entidad",
        "Cupo",
        "Afectado",
        "%",
        "PreCancelado",
        "%",
        "Deuda",
    ]:
        columnas = [
            "Organismo",
            "Sector",
            "Entidad",
            "Cupo",
            "Afectado",
            "Afectado_%",
            "PreCancelado",
            "PreCancelado_%",
            "Deuda",
        ]
    else:
        columnas_unicas = []
        vistos = {}
        for columna in columnas:
            vistos[columna] = vistos.get(columna, 0) + 1
            if vistos[columna] == 1:
                columnas_unicas.append(columna)
            else:
                columnas_unicas.append(f"{columna}_{vistos[columna]}")
        columnas = columnas_unicas

    filas_dict = []
    for fila in filas:
        registro = {}
        for indice, columna in enumerate(columnas):
            registro[columna] = fila[indice] if indice < len(fila) else None
        filas_dict.append(registro)

    return {
        "columnas": columnas,
        "filas": filas_dict,
        "cantidad_filas": len(filas_dict),
    }


def enviar_post_json(url, payload):
    body = json.dumps(payload).encode("utf-8")

    request = Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
        },
    )

    with urlopen(request, timeout=TIMEOUT_VIMARX) as response:
        respuesta_texto = response.read().decode("utf-8", errors="replace")

    return json.loads(respuesta_texto)


def consultar_vimarx_raw(cmd, campos, max_registros):
    payload = {
        "cmd": cmd,
        "tipo": TIPO_VIMARX,
        "campos": campos,
        "max": max_registros,
    }
    return enviar_post_json(URL_VIMARX, payload)


def construir_cmd_rango(id_desde, id_hasta):
    return f"{BASE_CMD_VIMARX} AND [ID] >= {id_desde} AND [ID] < {id_hasta}"


def obtener_primer_id():
    datos = consultar_vimarx_raw(BASE_CMD_VIMARX, "ID", 1)

    if not datos:
        return None

    return int(datos[0][0])


def encontrar_limite_superior_id(id_inicial):
    limite_inferior = id_inicial
    limite_superior = id_inicial + TAMANO_VENTANA_ID

    while True:
        cmd = f"{BASE_CMD_VIMARX} AND [ID] >= {limite_superior}"
        datos = consultar_vimarx_raw(cmd, "ID", 1)

        if not datos:
            break

        limite_inferior = limite_superior
        limite_superior *= 2

    while limite_inferior + 1 < limite_superior:
        medio = (limite_inferior + limite_superior) // 2
        cmd = f"{BASE_CMD_VIMARX} AND [ID] >= {medio}"
        datos = consultar_vimarx_raw(cmd, "ID", 1)

        if datos:
            limite_inferior = medio
        else:
            limite_superior = medio

    return limite_superior


def obtener_prestamos_en_rango(id_desde, id_hasta):
    cmd = construir_cmd_rango(id_desde, id_hasta)
    datos = consultar_vimarx_raw(cmd, CAMPOS_VIMARX, LIMITE_BLOQUE)

    if len(datos) < LIMITE_BLOQUE:
        print(f"Rango {id_desde} - {id_hasta}: {len(datos)} registros")
        return datos

    if id_hasta - id_desde <= 1:
        print(f"Rango {id_desde} - {id_hasta}: {len(datos)} registros")
        return datos

    medio = (id_desde + id_hasta) // 2
    print(
        f"Rango {id_desde} - {id_hasta}: al menos {LIMITE_BLOQUE}, dividiendo en "
        f"{id_desde} - {medio} y {medio} - {id_hasta}"
    )

    izquierda = obtener_prestamos_en_rango(id_desde, medio)
    derecha = obtener_prestamos_en_rango(medio, id_hasta)
    return izquierda + derecha


def obtener_todos_los_prestamos(id_desde, id_hasta):
    todos = []
    inicio = id_desde

    while inicio < id_hasta:
        fin = min(inicio + TAMANO_VENTANA_ID, id_hasta)
        todos.extend(obtener_prestamos_en_rango(inicio, fin))
        inicio = fin

    return todos


def deduplicar_prestamos_por_id(datos):
    prestamos_unicos = []
    vistos = set()

    for fila in datos:
        prestamo_id = fila[0]

        if prestamo_id in vistos:
            continue

        vistos.add(prestamo_id)
        prestamos_unicos.append(fila)

    prestamos_unicos.sort(key=lambda fila: fila[0])
    return prestamos_unicos


def filtrar_prestamos_con_deuda(datos):
    prestamos_con_deuda = []

    for fila in datos:
        deuda_vencida_con_punitorios = fila[5]
        deuda_total_vencida = fila[6]

        if deuda_vencida_con_punitorios > 0 or deuda_total_vencida > 0:
            prestamos_con_deuda.append(fila)

    return prestamos_con_deuda


def extraer_cuiles_unicos(datos):
    cuiles = []
    vistos = set()

    for fila in datos:
        cuil = str(fila[4]).strip()

        if not cuil or cuil in vistos:
            continue

        vistos.add(cuil)
        cuiles.append(cuil)

    return cuiles


def consultar_vimarx():
    print("Consultando API de Vimarx...")

    try:
        primer_id = obtener_primer_id()

        if primer_id is None:
            return {
                "ok": True,
                "cantidad_total": 0,
                "cantidad_con_deuda": 0,
                "cantidad_cuiles_unicos": 0,
                "cuiles": [],
            }

        limite_superior = encontrar_limite_superior_id(primer_id)
        print(f"Primer ID detectado: {primer_id}")
        print(f"Limite superior estimado: {limite_superior}")

        datos = obtener_todos_los_prestamos(primer_id, limite_superior)
        datos_unicos = deduplicar_prestamos_por_id(datos)
        datos_filtrados = filtrar_prestamos_con_deuda(datos_unicos)
        cuiles_unicos = extraer_cuiles_unicos(datos_filtrados)

        return {
            "ok": True,
            "cantidad_total": len(datos_unicos),
            "cantidad_con_deuda": len(datos_filtrados),
            "cantidad_cuiles_unicos": len(cuiles_unicos),
            "cuiles": cuiles_unicos,
        }

    except HTTPError as error:
        return {"ok": False, "error": f"Error HTTP: {error.code} - {error.reason}"}
    except URLError as error:
        return {"ok": False, "error": f"Error de conexion: {error.reason}"}
    except TimeoutError:
        return {"ok": False, "error": "La consulta a Vimarx supero el tiempo de espera"}
    except json.JSONDecodeError:
        return {"ok": False, "error": "La respuesta no vino en JSON valido"}


def enviar_post_form(url, payload, cookie):
    body = urlencode(payload).encode("utf-8")

    request = Request(
        url,
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "User-Agent": "Mozilla/5.0",
            "Cookie": cookie,
            "Origin": "https://www.santafe.gov.ar",
            "Referer": "https://www.santafe.gov.ar/cuad/movimiento.asp",
        },
    )

    with urlopen(request, timeout=TIMEOUT_CUAD) as response:
        return response.read().decode("latin-1", errors="replace")


def enviar_get_text(url, params, cookie):
    query = urlencode(params)
    request = Request(
        f"{url}?{query}",
        method="GET",
        headers={
            "User-Agent": "Mozilla/5.0",
            "Cookie": cookie,
            "Referer": "https://www.santafe.gov.ar/cuad/movimiento.asp",
        },
    )

    with urlopen(request, timeout=TIMEOUT_CUAD) as response:
        return response.read().decode("latin-1", errors="replace")


def construir_payload_cuad(cuil, emr_nombre, emr_id):
    return {
        "Modo": "BS",
        "Emr_Nombre": emr_nombre,
        "Emr_Id": emr_id,
        "Emt_Nome": "",
        "Emt_Id": "",
        "Emp_Cod": cuil,
        "Per_NroDoc": "",
        "none1": "",
    }


def consultar_cuad_raw(cuil, cookie, emr_nombre, emr_id):
    payload = construir_payload_cuad(cuil, emr_nombre, emr_id)

    try:
        html = enviar_post_form(URL_CUAD, payload, cookie)
        parsed = parsear_respuesta_cuad(html)
        return {
            "cuil": cuil,
            "emr_nombre": emr_nombre,
            "emr_id": emr_id,
            "consultado_en": ahora_iso(),
            "payload": payload,
            "parsed": parsed,
            "html": html,
        }
    except HTTPError as error:
        return {
            "ok": False,
            "status": "error_http",
            "cuil": cuil,
            "emr_nombre": emr_nombre,
            "emr_id": emr_id,
            "consultado_en": ahora_iso(),
            "payload": payload,
            "error": f"Error HTTP en CUAD: {error.code} - {error.reason}",
        }
    except URLError as error:
        return {
            "ok": False,
            "status": "error_conexion",
            "cuil": cuil,
            "emr_nombre": emr_nombre,
            "emr_id": emr_id,
            "consultado_en": ahora_iso(),
            "payload": payload,
            "error": f"Error de conexion en CUAD: {error.reason}",
        }
    except TimeoutError:
        return {
            "ok": False,
            "status": "timeout",
            "cuil": cuil,
            "emr_nombre": emr_nombre,
            "emr_id": emr_id,
            "consultado_en": ahora_iso(),
            "payload": payload,
            "error": "La consulta a CUAD supero el tiempo de espera",
        }


def consultar_grilla_movimientos(cookie, paginas):
    columnas_visibles = None
    columnas_normalizadas = None
    registros = []

    total_paginas = max(1, paginas or 1)

    for pagina in range(1, total_paginas + 1):
        html = enviar_get_text(
            URL_CUAD_GRILLA,
            {"Modo": "SERVICIO", "Pag": pagina, "ID": "MOVIMIENTOS"},
            cookie,
        )
        parsed = parsear_grilla_cuad_script(html)

        if parsed is None:
            continue

        if columnas_visibles is None:
            columnas_visibles = parsed["columnas_visibles"]
            columnas_normalizadas = parsed["columnas_normalizadas"]

        registros.extend(parsed["registros"])

    return {
        "columnas_visibles": columnas_visibles or [],
        "columnas_normalizadas": columnas_normalizadas or [],
        "registros": registros,
        "cantidad_registros": len(registros),
        "paginas_consultadas": total_paginas,
    }


def normalizar_cuil_entrada(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()
    if not texto:
        return ""

    if texto.endswith(".0"):
        texto = texto[:-2]

    return re.sub(r"\D", "", texto)


def normalizar_encabezado(valor):
    return re.sub(r"\s+", " ", str(valor or "").strip()).lower()


def indice_columna_excel(columna):
    if columna is None:
        return None

    columna = str(columna).strip()

    if not columna:
        return None

    if columna.isdigit():
        indice = int(columna) - 1
        if indice < 0:
            raise ValueError("La columna numerica debe ser 1 o mayor.")
        return indice

    if re.fullmatch(r"[A-Za-z]+", columna):
        indice = 0
        for caracter in columna.upper():
            indice = indice * 26 + (ord(caracter) - ord("A") + 1)
        return indice - 1

    return None


def resolver_columna_cuil_excel(encabezados, columna_cuiles=None):
    if columna_cuiles:
        objetivo = normalizar_encabezado(columna_cuiles)
        for indice, encabezado in enumerate(encabezados):
            if normalizar_encabezado(encabezado) == objetivo:
                return indice

        indice = indice_columna_excel(columna_cuiles)
        if indice is not None:
            if indice >= len(encabezados):
                raise ValueError(f"La columna {columna_cuiles} no existe en la hoja.")
            return indice

        raise ValueError(f"No se encontro la columna '{columna_cuiles}' en la hoja.")

    candidatos = {
        "cuil",
        "cuit",
        "cuil/cuit",
        "cuit/cuil",
        "nro cuil",
        "nro cuit",
    }

    for indice, encabezado in enumerate(encabezados):
        if normalizar_encabezado(encabezado) in candidatos:
            return indice

    raise ValueError(
        "No se encontro una columna CUIL/CUIT automatica. "
        "Indicala con --columna-cuiles."
    )


def cargar_cuiles_desde_excel(path, hoja_cuiles=None, columna_cuiles=None):
    wb = load_workbook(path, read_only=True, data_only=True)

    if hoja_cuiles:
        if hoja_cuiles not in wb.sheetnames:
            raise ValueError(f"La hoja '{hoja_cuiles}' no existe en {path.name}.")
        ws = wb[hoja_cuiles]
    else:
        ws = wb[wb.sheetnames[0]]

    filas = ws.iter_rows(values_only=True)
    encabezados = next(filas, None)

    if encabezados is None:
        raise ValueError(f"La hoja '{ws.title}' esta vacia.")

    encabezados = [
        str(valor).strip() if valor is not None else "" for valor in encabezados
    ]
    indice_cuil = resolver_columna_cuil_excel(encabezados, columna_cuiles)

    cuiles = []
    vistos = set()
    invalidos = []
    filas_leidas = 0

    for numero_fila, fila in enumerate(filas, start=2):
        filas_leidas += 1
        valor = fila[indice_cuil] if indice_cuil < len(fila) else None
        cuil = normalizar_cuil_entrada(valor)

        if not cuil:
            continue

        if len(cuil) != 11:
            invalidos.append({"fila": numero_fila, "valor": str(valor).strip()})
            continue

        if cuil in vistos:
            continue

        vistos.add(cuil)
        cuiles.append(cuil)

    return {
        "origen": str(path),
        "hoja": ws.title,
        "columna": encabezados[indice_cuil] or f"columna_{indice_cuil + 1}",
        "cantidad_filas_leidas": filas_leidas,
        "cantidad_cuiles_unicos": len(cuiles),
        "cantidad_invalidos": len(invalidos),
        "invalidos": invalidos[:20],
        "cuiles": cuiles,
    }


def cargar_cuiles_desde_archivo():
    if ARCHIVO_CUILES_ORIGEN is None:
        raise RuntimeError("No hay archivo de CUILes configurado.")

    if ARCHIVO_CUILES_ORIGEN.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise RuntimeError(
            "Formato de archivo no soportado. Usa un Excel .xlsx o .xlsm."
        )

    return cargar_cuiles_desde_excel(
        ARCHIVO_CUILES_ORIGEN,
        hoja_cuiles=HOJA_CUILES,
        columna_cuiles=COLUMNA_CUILES,
    )


def cargar_o_generar_cuiles_vimarx():
    if ARCHIVO_CUILES_FUENTE.exists():
        cuiles = leer_json(ARCHIVO_CUILES_FUENTE)
        print(f"Usando CUILes ya guardados en {ARCHIVO_CUILES_FUENTE}")
        print("CUILes disponibles:", len(cuiles))
        return cuiles

    resultado_vimarx = consultar_vimarx()
    print("\nResultados de Vimarx:")

    if not resultado_vimarx["ok"]:
        raise RuntimeError(resultado_vimarx["error"])

    print("Prestamos recibidos:", resultado_vimarx["cantidad_total"])
    print("Prestamos con deuda:", resultado_vimarx["cantidad_con_deuda"])
    print("CUILes unicos:", resultado_vimarx["cantidad_cuiles_unicos"])
    print("Primeros 10 CUILes:", resultado_vimarx["cuiles"][:10])
    guardar_json(resultado_vimarx["cuiles"], ARCHIVO_CUILES_FUENTE)
    print(f"Archivo generado: {ARCHIVO_CUILES_FUENTE}")
    return resultado_vimarx["cuiles"]


def cargar_cuiles_fuente():
    if MODO_CARGA_CUILES == "vimarx":
        return cargar_o_generar_cuiles_vimarx()

    resultado_archivo = cargar_cuiles_desde_archivo()

    if not resultado_archivo["cuiles"]:
        raise RuntimeError("El archivo no contiene CUILes validos para consultar.")

    print("\nResultados del archivo de CUILes:")
    print(f"Archivo origen: {resultado_archivo['origen']}")
    print(f"Hoja usada: {resultado_archivo['hoja']}")
    print(f"Columna usada: {resultado_archivo['columna']}")
    print("Filas leidas:", resultado_archivo["cantidad_filas_leidas"])
    print("CUILes unicos:", resultado_archivo["cantidad_cuiles_unicos"])
    print("Registros invalidos:", resultado_archivo["cantidad_invalidos"])
    print("Primeros 10 CUILes:", resultado_archivo["cuiles"][:10])

    if resultado_archivo["invalidos"]:
        print("Primeros invalidos detectados:", resultado_archivo["invalidos"][:5])

    guardar_json(resultado_archivo["cuiles"], ARCHIVO_CUILES_FUENTE)
    print(f"Archivo generado: {ARCHIVO_CUILES_FUENTE}")
    return resultado_archivo["cuiles"]


def es_respuesta_sin_resultado(html_text):
    return "parent.Emp_Id = -1" in html_text and "parent.Display('N')" in html_text


def es_sesion_invalida(html_text):
    html_lower = html_text.lower()
    return (
        "login.asp?modo=e" in html_lower
        or "login.asp?modo=m" in html_lower
        or "identificaciÃ³n - cuad" in html_lower
        or "identificacion - cuad" in html_lower
    )


def serializar_resultado_cuad(resultado):
    return {clave: valor for clave, valor in resultado.items() if clave != "html"}


def consultar_cuad_por_regimen(cuil, cookie, emr_nombre, emr_id):
    print(f"Consultando CUAD para {cuil} en {emr_nombre}...")

    resultado = consultar_cuad_raw(cuil, cookie, emr_nombre, emr_id)

    if "html" not in resultado:
        return resultado

    html_text = resultado["html"]
    parsed = resultado["parsed"]

    if es_sesion_invalida(html_text):
        resultado.update(
            {
                "ok": False,
                "status": "sesion_invalida",
                "error": "La sesion de CUAD no es valida o vencio",
            }
        )
        return resultado

    if es_respuesta_sin_resultado(html_text):
        resultado.update(
            {
                "ok": False,
                "status": "sin_resultado",
                "error": f"Sin resultado en CUAD ({emr_nombre})",
            }
        )
        return resultado

    if (
        parsed.get("contiene_set_empleado")
        or parsed.get("contiene_set_totales")
        or parsed.get("emp_id")
        or parsed.get("tabla_organismos")
    ):
        paginas_grilla = 1

        try:
            paginas_grilla = int(parsed.get("grilla", {}).get("paginas") or 1)
        except ValueError:
            paginas_grilla = 1

        try:
            parsed["tabla_movimientos"] = consultar_grilla_movimientos(
                cookie, paginas_grilla
            )
        except HTTPError as error:
            resultado.update(
                {
                    "ok": False,
                    "status": "error_http",
                    "error": (
                        "Error HTTP en grilla de movimientos: "
                        f"{error.code} - {error.reason}"
                    ),
                }
            )
            return resultado
        except URLError as error:
            resultado.update(
                {
                    "ok": False,
                    "status": "error_conexion",
                    "error": f"Error de conexion en grilla de movimientos: {error.reason}",
                }
            )
            return resultado
        except TimeoutError:
            resultado.update(
                {
                    "ok": False,
                    "status": "timeout",
                    "error": "La consulta de grilla de movimientos supero el tiempo de espera",
                }
            )
            return resultado

        resultado.update(
            {
                "ok": True,
                "status": "ok",
            }
        )
        return resultado

    resultado.update(
        {
            "ok": False,
            "status": "respuesta_no_reconocida",
            "error": "La respuesta de CUAD no contiene datos reconocibles",
        }
    )
    return resultado


def consultar_cuad(cuil, cookie):
    resultado = consultar_cuad_por_regimen(
        cuil, cookie, EMR_NOMBRE_ACTIVOS, EMR_ID_ACTIVOS
    )
    resultado["reconsultado_en_pasivos"] = False

    if resultado["status"] != "sin_resultado":
        return resultado

    print(
        f"{cuil} sin resultado en {EMR_NOMBRE_ACTIVOS}. "
        f"Reconsultando en {EMR_NOMBRE_PASIVOS}..."
    )

    resultado_pasivos = consultar_cuad_por_regimen(
        cuil, cookie, EMR_NOMBRE_PASIVOS, EMR_ID_PASIVOS
    )
    resultado_pasivos["reconsultado_en_pasivos"] = True
    resultado_pasivos["consulta_activos_previa"] = serializar_resultado_cuad(resultado)

    if resultado_pasivos["status"] == "sin_resultado":
        resultado_pasivos["error"] = (
            f"Sin resultado en CUAD ({EMR_NOMBRE_ACTIVOS} y {EMR_NOMBRE_PASIVOS})"
        )

    return resultado_pasivos


def consultar_cuad_con_reintentos(cuil, cookie):
    for intento in range(1, MAX_INTENTOS_CUAD + 1):
        resultado = consultar_cuad(cuil, cookie)

        if resultado["status"] not in ESTADOS_REINTENTABLES_CUAD:
            return resultado

        if intento == MAX_INTENTOS_CUAD:
            return resultado

        print(
            f"{cuil} {resultado['status']}: {resultado['error']}. "
            f"Reintentando en {PAUSA_REINTENTO_CUAD_SEGUNDOS} segundos "
            f"(intento {intento + 1} de {MAX_INTENTOS_CUAD})."
        )
        time.sleep(PAUSA_REINTENTO_CUAD_SEGUNDOS)

    return resultado


def resumir_resultado_cuad(resultado):
    emr_nombre = resultado.get("emr_nombre", "CUAD")

    if resultado["ok"]:
        empleado = (resultado.get("parsed") or {}).get("empleado", {}).get(
            "parsed"
        ) or {}
        nombre = empleado.get("apellido_nombre") or "sin_nombre"
        emp_id = (resultado.get("parsed") or {}).get("emp_id")
        return (
            f"{resultado['cuil']} [{emr_nombre}] ok "
            f"nombre: {nombre} emp_id: {emp_id}"
        )

    return (
        f"{resultado['cuil']} [{emr_nombre}] "
        f"{resultado['status']}: {resultado['error']}"
    )


def procesar_cuad_reanudable(cuiles, cookie):
    ultimos_resultados = cargar_ultimos_resultados_cuad()
    estado = resumir_estado_desde_resultados(ultimos_resultados)
    procesados = set(estado["procesados"])

    pendientes = [cuil for cuil in cuiles if cuil not in procesados]
    print("CUILes ya procesados:", len(procesados))
    print("CUILes pendientes:", len(pendientes))
    print(
        "CUILes pendientes por reintento:",
        estado["cantidad_pendientes_reintento"],
    )

    if LIMITE_CUAD is not None:
        pendientes = pendientes[:LIMITE_CUAD]
        print("Limite de esta corrida:", len(pendientes))

    procesados_en_esta_corrida = 0

    for cuil in pendientes:
        resultado = consultar_cuad_con_reintentos(cuil, cookie)

        append_ndjson(ARCHIVO_RESULTADOS_CUAD, serializar_resultado_cuad(resultado))
        ultimos_resultados.pop(cuil, None)
        ultimos_resultados[cuil] = serializar_resultado_cuad(resultado)
        estado = resumir_estado_desde_resultados(ultimos_resultados)
        print(resumir_resultado_cuad(resultado))

        if resultado["status"] == "sesion_invalida":
            print("Proceso detenido: la sesion de CUAD vencio o no es valida.")
            return

        if resultado["status"] in ESTADOS_REINTENTABLES_CUAD:
            print(
                f"{cuil} quedo pendiente para una proxima reanudacion "
                f"por {resultado['status']}."
            )

        procesados_en_esta_corrida += 1

        if procesados_en_esta_corrida % PAUSA_CADA == 0:
            print(
                f"Pausa larga de {PAUSA_LARGA_SEGUNDOS} segundos "
                f"despues de {procesados_en_esta_corrida} consultas."
            )
            time.sleep(PAUSA_LARGA_SEGUNDOS)
        else:
            time.sleep(DEMORA_ENTRE_CONSULTAS)

    if estado["cantidad_pendientes_reintento"] > 0:
        print(
            "Proceso de CUAD finalizado con pendientes para reintento:",
            estado["cantidad_pendientes_reintento"],
        )
        return

    print("Proceso de CUAD completado.")


def parsear_argumentos():
    parser = argparse.ArgumentParser(
        description="Consulta CUAD usando CUILes de Vimarx o de un archivo Excel."
    )
    parser.add_argument(
        "--archivo-cuiles",
        type=Path,
        help=(
            "Archivo Excel .xlsx/.xlsm con una columna de CUIL/CUIT. "
            "Si se informa, no consulta Vimarx."
        ),
    )
    parser.add_argument(
        "--hoja-cuiles",
        help=(
            "Nombre de la hoja de Excel a usar. " "Si se omite, toma la primera hoja."
        ),
    )
    parser.add_argument(
        "--columna-cuiles",
        help=(
            "Columna con CUIL/CUIT en el Excel. Acepta encabezado, letra "
            "de Excel o numero de columna."
        ),
    )
    parser.add_argument(
        "--linea-vimarx",
        default=DEFAULT_LINEA_SUPERIOR_VIMARX,
        help=(
            "Valor de [LineaPrestamo.Superior.Descripcion] en Vimarx. "
            f"Default: {DEFAULT_LINEA_SUPERIOR_VIMARX}"
        ),
    )
    parser.add_argument(
        "--estado-vimarx",
        default=DEFAULT_ESTADO_VIMARX,
        help=f"Valor de [Estado] en Vimarx. Default: {DEFAULT_ESTADO_VIMARX}",
    )
    parser.add_argument(
        "--etiqueta-salida",
        help=(
            "Texto opcional para el sufijo de los archivos de salida. "
            "Si no se informa, se usa la linea de Vimarx o el nombre del archivo."
        ),
    )
    parser.add_argument(
        "--limite-cuad",
        type=int,
        help="Limita la cantidad de CUILes a consultar en esta corrida.",
    )
    return parser.parse_args()


def main():
    args = parsear_argumentos()

    try:
        configurar_consulta(
            linea_superior=args.linea_vimarx,
            estado=args.estado_vimarx,
            etiqueta_salida=args.etiqueta_salida,
            archivo_cuiles=args.archivo_cuiles,
            hoja_cuiles=args.hoja_cuiles,
            columna_cuiles=args.columna_cuiles,
            limite_cuad=args.limite_cuad,
        )
    except ValueError as error:
        print(f"Configuracion invalida: {error}")
        return

    print("Iniciando consulta a APIs...")
    print(f"Periodo de consulta: {PERIODO_CONSULTA}")
    print(f"Directorio de la corrida: {DIRECTORIO_PERIODO}")

    if MODO_CARGA_CUILES == "vimarx":
        print("Origen de CUILes: Vimarx")
        print(f"Linea superior de Vimarx: {LINEA_SUPERIOR_VIMARX}")
        print(f"Estado en Vimarx: {ESTADO_VIMARX}")
    else:
        print("Origen de CUILes: archivo Excel")
        print(f"Archivo origen: {ARCHIVO_CUILES_ORIGEN}")
        if HOJA_CUILES:
            print(f"Hoja solicitada: {HOJA_CUILES}")
        if COLUMNA_CUILES:
            print(f"Columna solicitada: {COLUMNA_CUILES}")

    if INICIAR_NUEVA_CORRIDA:
        print("Modo: corrida nueva")
    else:
        print("Modo: reanudar corrida existente")

    preparar_corrida_nueva_si_corresponde()

    try:
        cuiles = cargar_cuiles_fuente()
    except RuntimeError as error:
        print("Error al obtener CUILes:", error)
        return

    print("\nIniciando proceso reanudable de CUAD...")
    procesar_cuad_reanudable(cuiles, COOKIE_CUAD)
    print(f"Archivo de CUILes: {ARCHIVO_CUILES_FUENTE}")
    print(f"Archivo de resultados: {ARCHIVO_RESULTADOS_CUAD}")


if __name__ == "__main__":
    main()
