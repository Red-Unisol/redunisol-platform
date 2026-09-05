"""Lectura del Excel de entrada y armado del Excel de resultados.

Formato de intercambio: como llegan los datos y como se entregan.
No sabe que Caja existe, asi que se puede probar sin red ni credenciales.

La lectura es tolerante a proposito: la planilla la arma otra persona y no
conviene depender de un encabezado exacto.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import registro

# Encabezados que se consideran la columna del CUIL, en orden de preferencia.
PISTAS_CUIL = ("cuil", "cuit", "documento", "dni")

FORMATO_IMPORTE = "#,##0.00"
FORMATO_PORCENTAJE = '0"%"'

COLOR_ENCABEZADO = "17365D"
COLOR_TEXTO_ENCABEZADO = "FFFFFF"
COLOR_NO_ENCONTRADO = "FFEB9C"
COLOR_ERROR = "FFC7CE"


class PlanillaError(RuntimeError):
    """La planilla de entrada no se pudo interpretar."""


@dataclass
class FilaEntrada:
    """Una fila de la planilla, con su CUIL ya normalizado."""

    numero: int
    cuil_original: str
    cuil: str = ""
    invalido: str = ""
    extra: Dict[str, Any] = field(default_factory=dict)


def leer_entrada(ruta: Path) -> tuple[List[FilaEntrada], List[str]]:
    """Devuelve las filas de la planilla y los nombres de las columnas extra.

    Busca la columna del CUIL por su encabezado. Si la planilla tiene una sola
    columna, usa esa sin importar como se llame.
    """
    ruta = Path(ruta)
    if not ruta.is_file():
        raise PlanillaError(f"No existe la planilla {ruta}")

    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro.active

    filas = list(hoja.iter_rows(values_only=True))
    libro.close()

    if not filas:
        raise PlanillaError("La planilla esta vacia.")

    encabezado = [_texto(c) for c in filas[0]]
    indice = _ubicar_columna_cuil(encabezado)

    if indice is None:
        raise PlanillaError(
            "No se encontro la columna del CUIL. Se buscaron encabezados que "
            f"contengan {', '.join(PISTAS_CUIL)}. Encabezados leidos: "
            f"{', '.join(e or '(vacio)' for e in encabezado)}."
        )

    extras = [e for i, e in enumerate(encabezado) if i != indice and e]
    salida: List[FilaEntrada] = []

    for numero, cruda in enumerate(filas[1:], start=2):
        if cruda is None or all(c is None or _texto(c) == "" for c in cruda):
            continue

        crudo_cuil = _texto(cruda[indice]) if indice < len(cruda) else ""
        extra = {
            encabezado[i]: cruda[i] if i < len(cruda) else None
            for i in range(len(encabezado))
            if i != indice and encabezado[i]
        }

        fila = FilaEntrada(numero=numero, cuil_original=crudo_cuil, extra=extra)
        try:
            fila.cuil = normalizar(crudo_cuil)
        except ValueError as exc:
            fila.invalido = str(exc)
        salida.append(fila)

    if not salida:
        raise PlanillaError("La planilla no tiene filas de datos.")

    return salida, extras


def _ubicar_columna_cuil(encabezado: List[str]) -> int | None:
    if len([e for e in encabezado if e]) == 1:
        return next(i for i, e in enumerate(encabezado) if e)

    for pista in PISTAS_CUIL:
        for i, texto in enumerate(encabezado):
            if pista in texto.lower():
                return i

    # Sin encabezados utiles: si hay una sola columna con contenido, es esa.
    if len(encabezado) == 1:
        return 0
    return None


def normalizar(valor: Any) -> str:
    """Deja solo los digitos y valida que sean 11."""
    import re

    digitos = re.sub(r"\D+", "", _texto(valor))
    if len(digitos) != 11:
        raise ValueError(f"se esperaban 11 digitos y hay {len(digitos)}")
    return digitos


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def escribir_resultados(
    ruta: Path,
    entradas: List[FilaEntrada],
    resultados: Dict[str, Dict[str, str]],
    columnas_extra: List[str],
) -> None:
    """Arma el Excel final: una fila por cada fila de la planilla original.

    Se respeta el orden de entrada y no se pierde ninguna fila, incluidas las
    que no se pudieron consultar. Asi la planilla se puede cruzar contra la que
    se mando.
    """
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Resultados"

    encabezado = [
        *columnas_extra,
        "cuil",
        "nombre",
        "apellido",
        "disponible",
        "tope_descuento",
        "estado",
        "consultado",
        "error",
    ]
    hoja.append(encabezado)

    for celda in hoja[1]:
        celda.font = Font(bold=True, color=COLOR_TEXTO_ENCABEZADO)
        celda.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for entrada in entradas:
        fila = resultados.get(entrada.cuil, {}) if entrada.cuil else {}
        estado = fila.get("estado") or (
            registro.ESTADO_CUIL_INVALIDO if entrada.invalido else "sin_consultar"
        )
        error = fila.get("error") or entrada.invalido

        hoja.append(
            [
                *[entrada.extra.get(c) for c in columnas_extra],
                entrada.cuil or entrada.cuil_original,
                fila.get("nombre", ""),
                fila.get("apellido", ""),
                registro.leer_importe(fila.get("disponible", "")),
                registro.leer_importe(fila.get("tope_descuento", "")),
                estado,
                fila.get("consultado_at", ""),
                error,
            ]
        )

    _dar_formato(hoja, len(columnas_extra), len(entradas))
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    libro.save(ruta)


def _dar_formato(hoja, cantidad_extra: int, cantidad_filas: int) -> None:
    col_disponible = cantidad_extra + 4
    col_tope = cantidad_extra + 5
    col_estado = cantidad_extra + 6

    for fila in range(2, cantidad_filas + 2):
        hoja.cell(row=fila, column=col_disponible).number_format = FORMATO_IMPORTE
        hoja.cell(row=fila, column=col_tope).number_format = FORMATO_PORCENTAJE

        estado = hoja.cell(row=fila, column=col_estado).value
        color = None
        if estado == registro.ESTADO_NO_ENCONTRADO:
            color = COLOR_NO_ENCONTRADO
        elif estado in (registro.ESTADO_ERROR, registro.ESTADO_CUIL_INVALIDO, "sin_consultar"):
            color = COLOR_ERROR
        if color:
            hoja.cell(row=fila, column=col_estado).fill = PatternFill("solid", fgColor=color)

    for i in range(1, hoja.max_column + 1):
        ancho = max(
            (len(str(hoja.cell(row=f, column=i).value or "")) for f in range(1, min(cantidad_filas + 2, 200))),
            default=10,
        )
        hoja.column_dimensions[get_column_letter(i)].width = min(max(ancho + 2, 10), 40)

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(hoja.max_column)}{cantidad_filas + 1}"
