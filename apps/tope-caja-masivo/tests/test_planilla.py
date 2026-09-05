"""Pruebas de lectura y escritura de planillas. Sin red ni credenciales."""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tope_caja_masivo import planilla, registro


def armar_excel(ruta: Path, filas: list[list]) -> None:
    libro = Workbook()
    hoja = libro.active
    for fila in filas:
        hoja.append(fila)
    libro.save(ruta)


class NormalizarTest(unittest.TestCase):
    def test_acepta_con_y_sin_guiones(self) -> None:
        self.assertEqual(planilla.normalizar("20-11111111-2"), "20111111112")
        self.assertEqual(planilla.normalizar("20111111112"), "20111111112")
        self.assertEqual(planilla.normalizar(" 20 11111111 2 "), "20111111112")

    def test_rechaza_lo_que_no_tenga_once_digitos(self) -> None:
        for valor in ("359661305", "", None, "1234", "201111111125"):
            with self.assertRaises(ValueError):
                planilla.normalizar(valor)

    def test_excel_guarda_los_cuils_como_numero(self) -> None:
        """Excel convierte 20111111112 a float y sin cuidado aparece un .0 al final."""
        self.assertEqual(planilla.normalizar(20111111112.0), "20111111112")


class LeerEntradaTest(unittest.TestCase):
    def setUp(self) -> None:
        self.dir = tempfile.TemporaryDirectory()
        self.ruta = Path(self.dir.name) / "entrada.xlsx"

    def tearDown(self) -> None:
        self.dir.cleanup()

    def test_encuentra_la_columna_por_el_encabezado(self) -> None:
        armar_excel(self.ruta, [["legajo", "CUIL", "sucursal"], [1, "20111111112", "centro"]])
        filas, extras = planilla.leer_entrada(self.ruta)

        self.assertEqual(filas[0].cuil, "20111111112")
        self.assertEqual(extras, ["legajo", "sucursal"])
        self.assertEqual(filas[0].extra, {"legajo": 1, "sucursal": "centro"})

    def test_reconoce_variantes_del_encabezado(self) -> None:
        for encabezado in ("cuil", "CUIT", "Cuit/Cuil", "NRO DOCUMENTO"):
            armar_excel(self.ruta, [[encabezado], ["20111111112"]])
            filas, _ = planilla.leer_entrada(self.ruta)
            self.assertEqual(filas[0].cuil, "20111111112", f"fallo con {encabezado!r}")

    def test_una_sola_columna_se_usa_aunque_no_se_llame_cuil(self) -> None:
        armar_excel(self.ruta, [["numeros"], ["20111111112"]])
        filas, extras = planilla.leer_entrada(self.ruta)

        self.assertEqual(filas[0].cuil, "20111111112")
        self.assertEqual(extras, [])

    def test_marca_los_invalidos_sin_cortar_la_lectura(self) -> None:
        armar_excel(self.ruta, [["cuil"], ["20111111112"], ["123"], ["20123456783"]])
        filas, _ = planilla.leer_entrada(self.ruta)

        self.assertEqual(len(filas), 3)
        self.assertEqual(filas[0].invalido, "")
        self.assertTrue(filas[1].invalido)
        self.assertEqual(filas[1].cuil, "")
        self.assertEqual(filas[2].cuil, "20123456783")

    def test_saltea_filas_vacias(self) -> None:
        armar_excel(self.ruta, [["cuil"], ["20111111112"], [None], ["20123456783"]])
        filas, _ = planilla.leer_entrada(self.ruta)
        self.assertEqual(len(filas), 2)

    def test_conserva_el_numero_de_fila_del_excel(self) -> None:
        armar_excel(self.ruta, [["cuil"], ["20111111112"], ["123"]])
        filas, _ = planilla.leer_entrada(self.ruta)
        self.assertEqual(filas[0].numero, 2)
        self.assertEqual(filas[1].numero, 3)

    def test_avisa_si_no_encuentra_la_columna(self) -> None:
        armar_excel(self.ruta, [["nombre", "sucursal"], ["ana", "centro"]])
        with self.assertRaises(planilla.PlanillaError) as ctx:
            planilla.leer_entrada(self.ruta)
        self.assertIn("nombre", str(ctx.exception))

    def test_avisa_si_la_planilla_no_existe(self) -> None:
        with self.assertRaises(planilla.PlanillaError):
            planilla.leer_entrada(Path(self.dir.name) / "no-existe.xlsx")


class EscribirResultadosTest(unittest.TestCase):
    def setUp(self) -> None:
        self.dir = tempfile.TemporaryDirectory()
        self.entrada = Path(self.dir.name) / "entrada.xlsx"
        self.salida = Path(self.dir.name) / "salida.xlsx"

    def tearDown(self) -> None:
        self.dir.cleanup()

    def test_una_fila_por_fila_de_la_planilla_original(self) -> None:
        armar_excel(
            self.entrada,
            [["legajo", "cuil"], [1, "20111111112"], [2, "123"], [3, "20123456783"]],
        )
        filas, extras = planilla.leer_entrada(self.entrada)
        resultados = {
            "20111111112": {
                "estado": registro.ESTADO_OK,
                "nombre": "ANA",
                "apellido": "PEREZ",
                "disponible": "1500.50",
                "tope_descuento": "20.00",
                "consultado_at": "2026-09-01T10:00:00",
                "error": "",
            }
        }
        planilla.escribir_resultados(self.salida, filas, resultados, extras)

        hoja = load_workbook(self.salida).active
        self.assertEqual(hoja.max_row, 4)  # encabezado + 3

        # columnas: legajo | cuil | nombre | apellido | disponible | tope | estado | ...
        datos = list(hoja.iter_rows(min_row=2, values_only=True))
        self.assertEqual(datos[0][2], "ANA")
        self.assertEqual(datos[0][4], 1500.50)
        self.assertEqual(datos[1][6], registro.ESTADO_CUIL_INVALIDO)
        self.assertEqual(datos[2][6], "sin_consultar")

    def test_los_importes_quedan_como_numero_no_como_texto(self) -> None:
        armar_excel(self.entrada, [["cuil"], ["20111111112"]])
        filas, extras = planilla.leer_entrada(self.entrada)
        resultados = {
            "20111111112": {
                "estado": registro.ESTADO_OK,
                "disponible": "-201503.26",
                "tope_descuento": "50.00",
            }
        }
        planilla.escribir_resultados(self.salida, filas, resultados, extras)

        hoja = load_workbook(self.salida).active
        disponible = hoja.cell(row=2, column=4).value
        self.assertIsInstance(disponible, float)
        self.assertEqual(disponible, -201503.26)


if __name__ == "__main__":
    unittest.main()
