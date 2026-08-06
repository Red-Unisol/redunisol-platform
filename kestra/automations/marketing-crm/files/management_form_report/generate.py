#!/usr/bin/env python3
"""Generate the private daily management report for form submissions to Bitrix."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY, WHITE = "17365D", "FFFFFF"
GREEN, YELLOW, ORANGE, RED = "C6EFCE", "FFEB9C", "FCE4D6", "FFC7CE"


def api_get(session: requests.Session, url: str, **params: Any) -> Any:
    error: Exception | None = None
    for attempt in range(5):
        try:
            response = session.get(url, params=params, timeout=(10, 90))
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            error = exc
            if attempt < 4:
                time.sleep(2**attempt)
    raise RuntimeError(f"No se pudo consultar Kestra: {error}")


def executions(session: requests.Session, base: str, tenant: str, namespace: str, flow: str) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    page = 1
    while True:
        payload = api_get(
            session,
            f"{base}/api/v1/{tenant}/executions/search",
            namespace=namespace,
            flowId=flow,
            page=page,
            size=100,
        )
        batch = payload.get("results", [])
        result.extend(batch)
        if not batch or len(result) >= int(payload.get("total", len(result))):
            return result
        page += 1


def state(row: dict[str, Any]) -> str:
    return str((row.get("state") or {}).get("current") or "UNKNOWN")


def iso(value: Any) -> datetime | None:
    if not isinstance(value, str):
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def body(row: dict[str, Any]) -> dict[str, Any]:
    variables = (row.get("trigger") or {}).get("variables") or {}
    value = variables.get("body") or {}
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except ValueError:
            value = {}
    return value if isinstance(value, dict) else {}


def child_id_from_parent(row: dict[str, Any]) -> str:
    for task in row.get("taskRunList") or []:
        if task.get("taskId") != "persistir_bitrix":
            continue
        outputs = task.get("outputs") or {}
        for key in ("executionId", "execution_id", "id"):
            if outputs.get(key):
                return str(outputs[key])
    return ""


def cross_children(parents: list[dict[str, Any]], children: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    children_by_id = {str(row.get("id")): row for row in children}
    result: dict[str, dict[str, Any]] = {}
    for child in children:
        parent_id = child.get("parentId") or child.get("parentExecutionId")
        if parent_id:
            result[str(parent_id)] = child
    for parent in parents:
        child_id = child_id_from_parent(parent)
        if child_id and child_id in children_by_id:
            result[str(parent.get("id"))] = children_by_id[child_id]
    return result


def normalized(row: dict[str, Any], child: dict[str, Any] | None) -> dict[str, Any]:
    request = body(row)
    parent_outputs = row.get("outputs") or {}
    child_outputs = (child or {}).get("outputs") or {}
    outputs = child_outputs if child_outputs.get("lead_id") else parent_outputs
    action = str(outputs.get("action") or parent_outputs.get("action") or "").lower()
    lead_id = str(outputs.get("lead_id") or "")
    technical_error = state(row) in {"FAILED", "KILLED"} or (child is not None and state(child) in {"FAILED", "KILLED"})
    reason = str(outputs.get("message") or outputs.get("reason") or parent_outputs.get("message") or parent_outputs.get("reason") or "")
    if lead_id:
        category = "Lead aprobado" if action != "rejected" else "Lead rechazado en Bitrix"
    elif technical_error:
        category = "Error técnico"
    elif action == "rejected":
        category = "Rechazo antes de Bitrix"
    elif action == "error":
        category = "Error de datos"
    else:
        category = "Pendiente de verificación"
    return {
        "date": iso((row.get("state") or {}).get("startDate")),
        "execution_id": str(row.get("id") or ""),
        "child_execution_id": str((child or {}).get("id") or ""),
        "revision": row.get("flowRevision"),
        "technical_state": state(row),
        "action": action,
        "reason": reason,
        "category": category,
        "lead_id": lead_id,
        "contact_id": str(outputs.get("contact_id") or ""),
        "name": request.get("full_name"),
        "cuil": request.get("cuil"),
        "email": request.get("email"),
        "whatsapp": request.get("whatsapp"),
        "province": request.get("province"),
        "employment": request.get("employment_status"),
        "bank": request.get("payment_bank"),
        "source": request.get("lead_source"),
        "utm_source": request.get("utm_source"),
        "utm_medium": request.get("utm_medium"),
        "utm_campaign": request.get("utm_campaign"),
        "landing": request.get("landing_slug") or request.get("landing_url"),
    }


def compact(ws, widths: dict[int, int] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for number in range(2, ws.max_row + 1):
        ws.row_dimensions[number].height = 18
        for cell in ws[number]:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    widths = widths or {}
    for index, column in enumerate(ws.columns, 1):
        longest = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(longest, widths.get(index, 36))


def build(rows: list[dict[str, Any]]) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen Ejecutivo"
    counts = Counter(row["category"] for row in rows)
    leads = [row for row in rows if row["lead_id"]]
    summary.append(["INFORME FORMULARIO → BITRIX", "Valor"])
    dates = [row["date"] for row in rows if row["date"]]
    metrics = [
        ("Período", f"{min(dates):%d/%m/%Y} al {max(dates):%d/%m/%Y}" if dates else ""),
        ("Formularios recibidos", len(rows)),
        ("Leads confirmados en Bitrix", len({row["lead_id"] for row in leads})),
        ("Leads aprobados/ingresados", counts["Lead aprobado"]),
        ("Leads rechazados en Bitrix", counts["Lead rechazado en Bitrix"]),
        ("Conversión formulario → lead", len(leads) / len(rows) if rows else 0),
        ("Rechazados antes de Bitrix", counts["Rechazo antes de Bitrix"]),
        ("Errores de datos", counts["Error de datos"]),
        ("Errores técnicos", counts["Error técnico"]),
        ("Pendientes de verificación", counts["Pendiente de verificación"]),
    ]
    for metric in metrics:
        summary.append(metric)
    compact(summary, {1: 38, 2: 28})
    summary.auto_filter.ref = None
    for number in range(2, summary.max_row + 1):
        summary.cell(number, 1).font = Font(bold=True, color=NAVY)
        if summary.cell(number, 1).value == "Conversión formulario → lead":
            summary.cell(number, 2).number_format = "0.00%"
    summary["D1"], summary["E1"] = "Resultado", "Cantidad"
    category_names = ("Lead aprobado", "Lead rechazado en Bitrix", "Rechazo antes de Bitrix", "Error de datos", "Error técnico", "Pendiente de verificación")
    categories = [(key, counts[key]) for key in category_names]
    for index, item in enumerate(categories, 2):
        summary.cell(index, 4, item[0])
        summary.cell(index, 5, item[1])
    pie = PieChart()
    pie.title = "Resultado de formularios"
    pie.add_data(Reference(summary, min_col=5, min_row=1, max_row=7), titles_from_data=True)
    pie.set_categories(Reference(summary, min_col=4, min_row=2, max_row=7))
    summary.add_chart(pie, "G2")

    daily_ws = wb.create_sheet("Evolución diaria")
    daily_ws.append(["Fecha", "Formularios", "Leads confirmados", "Conversión", *category_names])
    rows_by_day: dict[Any, list[dict[str, Any]]] = {}
    for row in rows:
        if row["date"]:
            rows_by_day.setdefault(row["date"].date(), []).append(row)
    for day, day_rows in sorted(rows_by_day.items()):
        day_counts = Counter(row["category"] for row in day_rows)
        day_leads = {row["lead_id"] for row in day_rows if row["lead_id"]}
        daily_ws.append([
            day,
            len(day_rows),
            len(day_leads),
            len(day_leads) / len(day_rows),
            *(day_counts[category] for category in category_names),
        ])
    compact(daily_ws, {1: 14, 2: 16, 3: 20, 4: 14})
    for number in range(2, daily_ws.max_row + 1):
        daily_ws.cell(number, 1).number_format = "dd/mm/yyyy"
        daily_ws.cell(number, 4).number_format = "0.00%"
    if daily_ws.max_row > 1:
        daily_chart = BarChart()
        daily_chart.type = "col"
        daily_chart.grouping = "stacked"
        daily_chart.overlap = 100
        daily_chart.title = "Resultados de formularios por día"
        daily_chart.y_axis.title = "Cantidad"
        daily_chart.x_axis.title = "Fecha"
        daily_chart.add_data(
            Reference(daily_ws, min_col=5, max_col=10, min_row=1, max_row=daily_ws.max_row),
            titles_from_data=True,
        )
        daily_chart.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=daily_ws.max_row))
        daily_chart.height = 9
        daily_chart.width = 20
        daily_ws.add_chart(daily_chart, "L2")

    lead_ws = wb.create_sheet("Leads en Bitrix")
    lead_ws.append(["fecha", "lead_id", "contact_id", "nombre", "cuil", "email", "whatsapp", "provincia", "situación laboral", "banco de cobro", "origen", "utm_source", "utm_medium", "utm_campaign", "landing", "resultado", "motivo", "ejecución Kestra", "subejecución", "revisión"])
    for row in sorted(leads, key=lambda item: item["date"] or datetime.min):
        lead_ws.append([row[key] for key in ("date", "lead_id", "contact_id", "name", "cuil", "email", "whatsapp", "province", "employment", "bank", "source", "utm_source", "utm_medium", "utm_campaign", "landing", "category", "reason", "execution_id", "child_execution_id", "revision")])
    compact(lead_ws, {4: 28, 6: 32, 10: 38, 14: 42, 17: 62, 18: 28, 19: 28})

    rejected_ws = wb.create_sheet("No enviados")
    rejected_ws.append(["fecha", "nombre", "cuil", "email", "whatsapp", "provincia", "origen", "resultado", "motivo", "acción recomendada", "estado Kestra", "ejecución Kestra", "revisión"])
    recommendations = {"Rechazo antes de Bitrix": "Sin acción técnica", "Error de datos": "Corregir validación o catálogo", "Error técnico": "Reintentar y revisar logs", "Pendiente de verificación": "Revisar manualmente"}
    for row in sorted((item for item in rows if not item["lead_id"]), key=lambda item: item["date"] or datetime.min):
        rejected_ws.append([row["date"], row["name"], row["cuil"], row["email"], row["whatsapp"], row["province"], row["source"], row["category"], row["reason"], recommendations[row["category"]], row["technical_state"], row["execution_id"], row["revision"]])
    compact(rejected_ws, {2: 28, 4: 32, 8: 28, 9: 62, 10: 44, 12: 28})

    quality = wb.create_sheet("Motivos y Calidad")
    quality.append(["Resultado", "Motivo", "Cantidad", "% total", "Acción recomendada"])
    reason_counts = Counter((row["category"], row["reason"]) for row in rows if not row["lead_id"])
    for (category, reason), count in reason_counts.most_common():
        quality.append([category, reason, count, count / len(rows) if rows else 0, recommendations[category]])
    compact(quality, {1: 28, 2: 72, 5: 48})
    for number in range(2, quality.max_row + 1):
        quality.cell(number, 4).number_format = "0.00%"
    if quality.max_row > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Principales motivos"
        last = min(quality.max_row, 11)
        chart.add_data(Reference(quality, min_col=3, min_row=1, max_row=last), titles_from_data=True)
        chart.set_categories(Reference(quality, min_col=2, min_row=2, max_row=last))
        quality.add_chart(chart, "G2")
    return wb


def publish(workbook: Workbook, root: Path, generated_at: datetime) -> tuple[Path, Path]:
    report_dir = root / "marketing" / "formulario-bitrix"
    history_dir = report_dir / "historico"
    history_dir.mkdir(parents=True, exist_ok=True)
    dated = history_dir / f"{generated_at:%Y-%m-%d}.xlsx"
    latest = report_dir / "ultimo.xlsx"
    with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, dated)
        dated.chmod(0o644)
        with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
            latest_temporary = Path(handle.name)
        shutil.copy2(dated, latest_temporary)
        os.replace(latest_temporary, latest)
        latest.chmod(0o644)
    finally:
        temporary.unlink(missing_ok=True)
    return latest, dated


def main() -> None:
    base = os.environ["REPORTS_KESTRA_URL"].rstrip("/")
    tenant = os.getenv("REPORTS_KESTRA_TENANT", "main")
    namespace = os.getenv("REPORTS_NAMESPACE", "redunisol.prod.marketing-crm")
    session = requests.Session()
    session.auth = (os.environ["REPORTS_KESTRA_USERNAME"], os.environ["REPORTS_KESTRA_PASSWORD"])
    parents = executions(session, base, tenant, namespace, "bitrix24_form_webhook")
    children = executions(session, base, tenant, namespace, "bitrix24_form_persistence")
    crossed = cross_children(parents, children)
    rows = [normalized(row, crossed.get(str(row.get("id")))) for row in parents]
    latest, dated = publish(build(rows), Path(os.getenv("REPORTS_ROOT", "/reports")), datetime.now())
    print(json.dumps({"ok": True, "forms": len(rows), "leads": sum(bool(row["lead_id"]) for row in rows), "latest": str(latest), "history": str(dated)}))


if __name__ == "__main__":
    main()
