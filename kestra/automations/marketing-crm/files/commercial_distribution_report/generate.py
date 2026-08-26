#!/usr/bin/env python3
"""Generate the private commercial classification and distribution audit report."""

from __future__ import annotations

import json
import os
import shutil
import tempfile
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY, WHITE = "17365D", "FFFFFF"
FLOW_IDS = (
    "bitrix24_catamarca_deal_qualification",
    "bitrix24_deal_assignment_queue",
)
ARGENTINA_TIMEZONE = ZoneInfo("America/Argentina/Buenos_Aires")
DEFAULT_AUDIT_FROM = "2026-08-11T13:00:00-03:00"
ASSIGNMENT_STRATEGY_LABELS = {
    "contact_history": "Continuidad con vendedor anterior del contacto",
    "legacy_contact_history": "Continuidad con vendedor histórico del contacto",
    "round_robin": "Siguiente vendedor del round-robin",
    "legacy_round_robin": "Round-robin calculado con negociaciones históricas",
    "round_robin_initial": "Primer vendedor online por falta de antecedentes",
    "single_seller": "Único vendedor configurado para el bucket",
    "outside_hours_manual": "Fuera de horario laboral; gestión manual con Maru",
    "no_online_sellers_manual": "Sin vendedores online; gestión manual con Maru",
    "assignment_queue": "En cola temporal hasta que haya un vendedor disponible",
    "assignment_queue_waiting": "El bucket continúa sin vendedores disponibles",
    "assignment_queue_closed_manual": "Cierre semanal; gestión manual con Maru",
    "assignment_queue_error": "El reintento de la cola no pudo completarse",
    "commercial_rejection_manual": "Rechazo comercial; gestión manual con Maru",
    "rejection_without_distribution": "Rechazo directo; no requiere vendedor",
    "no_matching_bucket": "No existe un bucket aplicable",
    "not_applicable": "La negociación ya no estaba pendiente",
    "commercial_data_pending": "Pendiente de información BCRA; todavía no corresponde distribuir",
    "technical_error": "La ejecución terminó con un error técnico",
}
REASON_LABELS = {
    "amejuca_premium": (
        "Cumple las condiciones BCRA de AMEJUCA Premium: hasta cinco entidades "
        "en situación 2, ninguna superior a 2 y banco de cobro hasta situación 2."
    ),
    "amejuca_special": (
        "El banco de cobro está en situación 1 y el perfil BCRA corresponde a la "
        "línea AMEJUCA Especial."
    ),
    "amejuca_line_ambiguous_for_payment_bank_two": (
        "El banco de cobro está en situación 2 y los datos no permiten elegir con "
        "certeza entre AMEJUCA Premium y Especial."
    ),
    "bcra_more_than_four_high_risk_situations": (
        "Tiene más de cuatro entidades con situación BCRA de riesgo."
    ),
    "payment_bank_situation_above_two": "El banco de cobro está en situación BCRA mayor a 2.",
    "cordoba_bank_situation_above_one": "Bancor está en situación BCRA 2 o superior.",
    "cde_more_than_two_high_risk_situations": (
        "Tiene más de dos entidades con situación BCRA entre 2 y 3."
    ),
    "cde_premium": (
        "No tiene un préstamo Cruz del Eje activo y todas las entidades están en situación 1."
    ),
    "cde_special": (
        "No tiene un préstamo Cruz del Eje activo y presenta situaciones BCRA admitidas para Especial."
    ),
    "cde_ren_premium": (
        "Tiene un préstamo Cruz del Eje activo sin atraso y todas las entidades están en situación 1."
    ),
    "cde_ren_special": (
        "Tiene un préstamo Cruz del Eje activo sin atraso y situaciones BCRA admitidas para Especial."
    ),
    "cde_parallel_requires_manual_review": (
        "Tiene más de un préstamo activo de la familia Cruz del Eje."
    ),
    "cde_active_loan_in_arrears": "Tiene un préstamo Cruz del Eje activo con días de atraso.",
    "cbu_more_than_five_entities": "Tiene más de cinco entidades informadas en BCRA.",
    "cbu_situation_above_one": "Tiene al menos una entidad en situación BCRA mayor a 1.",
    "cbu_passive_age_80_or_more": "La persona jubilada o pensionada tiene 80 años o más.",
    "cbu_gender_required_for_age_limit": (
        "La edad requiere conocer el género para aplicar correctamente el límite de la línea CBU."
    ),
    "cbu_approved": (
        "Tiene hasta cinco entidades, todas en situación BCRA 1, y cumple el límite de edad aplicable."
    ),
    "caja_age_80_or_more": "La persona tiene 80 años o más.",
    "caja_new_payment_bank_above_one": (
        "Es cliente nuevo y el banco de cobro está en situación BCRA mayor a 1."
    ),
    "caja_morosos_payment_bank_above_one": (
        "El banco de cobro está en situación BCRA mayor a 1 para la línea Caja Morosos."
    ),
    "caja_morosos_excluded_entity": (
        "Tiene una entidad excluida en situación BCRA 4 o 5; requiere revisión comercial."
    ),
    "caja_morosos_parallel_minimum_not_met": (
        "Tiene un préstamo Caja activo con menos de cuatro cuotas pagadas."
    ),
    "caja_irregular_parallel_minimum_not_met": (
        "Tiene un préstamo Caja activo con menos de cuatro cuotas pagadas."
    ),
    "caja_general_parallel_minimum_not_met": (
        "Tiene un préstamo Caja activo sin la primera cuota pagada."
    ),
    "caja_morosos": "Presenta situaciones BCRA 4 o 5 y cumple las condiciones de Caja Morosos.",
    "caja_irregulares": "Presenta situaciones BCRA 2 o 3 y cumple las condiciones de Caja Irregulares.",
    "caja_general": "Es cliente recurrente, no presenta situaciones BCRA 2 a 5 y cumple el mínimo de cuotas.",
    "caja_nuevo": "Es cliente nuevo, el banco de cobro está en situación 1 y no presenta situaciones BCRA 2 a 5.",
    "club_mutual_cbu": "Es socio activo de Club Mutual y cumple las condiciones de edad y BCRA de la línea.",
    "unc_activity_not_verifiable": "No se pudo confirmar que sea socio activo de Club Mutual.",
    "unc_gender_required_for_age_limit": "La edad requiere conocer el género para aplicar el límite de la línea UNC.",
    "unc_more_than_three_high_risk_situations": "Tiene más de tres entidades con situación BCRA entre 2 y 3.",
    "unc_banco_nacion_irregular": "Banco Nación está en situación BCRA mayor a 1.",
    "daspu_form_691_or_limit_not_available": "Falta validar el formulario 691 o el límite disponible de DASPU.",
    "missing_birthdate": "No se pudo determinar la edad porque falta la fecha de nacimiento.",
    "missing_vimarx_credit_data": "Faltan datos de préstamos de Vimarx para decidir automáticamente.",
    "missing_bcra_snapshot": "No hay información BCRA suficiente para decidir automáticamente.",
    "bcra_snapshot_not_conclusive": "La consulta BCRA no produjo información concluyente.",
    "bcra_refresh_missing_cuil": "No se pudo actualizar BCRA porque falta el CUIL.",
    "bcra_refresh_failed": "No fue posible actualizar BCRA; el dato anterior no se utilizó.",
    "bcra_retry_scheduled": "BCRA no respondió; Kestra programó un nuevo intento sin tomar una decisión comercial.",
    "bcra_retry_exhausted": "BCRA no respondió durante la ventana de reintentos; el caso requiere revisión manual.",
    "bcra_not_found": "BCRA no devolvió información para el CUIL consultado.",
    "bcra_invalid_identification": "BCRA informó que la identificación consultada no es válida.",
    "payment_bank_not_identifiable": "No se pudo identificar el banco de cobro dentro de la información BCRA.",
    "missing_recurrent_membership_data": "Es socio, pero falta información para evaluar la renovación automáticamente.",
    "missing_membership_data": "No se pudo confirmar si es socio nuevo o recurrente.",
    "missing_prequalification_data": "Faltan datos de la precalificación necesarios para evaluar la negociación.",
    "unsupported_cordoba_employment_status": "La situación laboral no tiene una regla comercial automática en Córdoba.",
    "province_not_supported_for_deal_classification": "La provincia no tiene clasificación comercial automática.",
    "missing_routing_data": "Faltan provincia o situación laboral para determinar el grupo de distribución.",
    "no_matching_bucket": "No existe un grupo de distribución configurado para esos datos.",
    "outside_business_hours": "La negociación ingresó fuera del horario de distribución automática.",
    "no_online_sellers": "No había vendedores del grupo conectados en Bitrix.",
    "assignment_queued": "No había vendedores disponibles; quedó en cola temporal.",
    "assignment_queue_waiting": "El bucket continúa sin vendedores disponibles.",
    "assignment_queue_distributed": "Apareció un vendedor y se distribuyó desde la cola.",
    "assignment_queue_closed": "La cola cerró y el caso quedó con Maru.",
    "deal_not_pending": "La negociación ya había salido de la etapa pendiente cuando Kestra la revisó.",
    "internal_error": "La ejecución no pudo completar el procesamiento.",
}
STAGE_LABELS = {
    "C1:KESTRA_PENDING": "Preclasificación Kestra",
    "C1:KESTRA_REVIEW": "Revisión manual Kestra",
    "C1:KESTRA_ROUTE_REVIEW": "Revisión de enrutamiento Kestra",
    "C1:KESTRA_QUEUE": "Cola de distribución Kestra",
    "C1:NEW": "Presentación",
    "C1:5": "Situación negativa en BCRA",
}
MUTED_FILL = "F2F2F2"
MUTED_FONT = "7F7F7F"
CHAT_TRANSFER_STATUS_LABELS = {
    "transferred": "Todos los chats encontrados fueron transferidos",
    "partially_transferred": "Transferencia parcial",
    "no_chats_found": "No se encontraron chats asociados",
    "no_transferable_session": "Chats encontrados sin sesión transferible",
    "not_evaluated": "No aplica: no hubo transferencia de chat",
}
CHAT_SKIP_REASON_LABELS = {
    "inspection_error": "no se pudo inspeccionar el chat",
    "invalid_dialog": "Bitrix devolvió un diálogo inválido",
    "no_current_transferable_session": "sin sesión actual transferible",
}


def api_get(session: requests.Session, url: str, **params: Any) -> Any:
    error: Exception | None = None
    for attempt in range(5):
        try:
            response = session.get(url, params=params, timeout=(10, 90))
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            error = exc
            if attempt < 4:
                time.sleep(2**attempt)
    raise RuntimeError(f"No se pudo consultar Kestra: {error}")


def bitrix_user_names(api_url: str, user_ids: set[str]) -> dict[str, str]:
    if not api_url or not user_ids:
        return {}
    response = requests.post(
        f"{api_url.rstrip('/')}/user.get.json",
        json={"FILTER": {"ID": sorted(user_ids, key=int)}},
        timeout=(10, 60),
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("error"):
        raise RuntimeError(payload.get("error_description") or payload["error"])
    result = payload.get("result") or []
    return {
        str(user.get("ID") or user.get("id")): " ".join(
            part
            for part in (
                str(user.get("NAME") or user.get("name") or "").strip(),
                str(user.get("LAST_NAME") or user.get("last_name") or "").strip(),
            )
            if part
        )
        for user in result
        if user.get("ID") or user.get("id")
    }


def executions(
    session: requests.Session,
    base: str,
    tenant: str,
    namespace: str,
    flow: str,
) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    page = 1
    while True:
        payload = api_get(
            session,
            f"{base}/api/v1/{tenant}/executions/search",
            namespace=namespace,
            flowId=flow,
            page=page,
            size=100,
        )
        batch = payload.get("results", [])
        result.extend(batch)
        if not batch or len(result) >= int(payload.get("total", len(result))):
            return result
        page += 1


def execution_state(row: dict[str, Any]) -> str:
    return str((row.get("state") or {}).get("current") or "UNKNOWN")


def parse_datetime(value: Any) -> datetime | None:
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(ARGENTINA_TIMEZONE)
    return parsed.replace(tzinfo=None)


def parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in {"true", "1", "yes", "si"}:
        return True
    if normalized in {"false", "0", "no"}:
        return False
    return None


def parse_int(value: Any) -> int:
    try:
        return int(str(value or "0"))
    except ValueError:
        return 0


def distribution_status(
    action: str,
    strategy: str,
    technical_state: str,
    message: str = "",
    distribution_action: str = "",
) -> str:
    if technical_state in {"FAILED", "KILLED"} or action == "error":
        if (
            technical_state == "SUCCESS"
            and "no hay vendedores online disponibles" in message.lower()
        ):
            return "Sin vendedor disponible"
        return "Error técnico"
    if action == "bcra_pending":
        return "Pendiente BCRA"
    if distribution_action == "assigned":
        return "Distribuido"
    if distribution_action == "queued":
        return "En cola de distribución"
    if distribution_action == "manual_owner":
        return "Gestión manual con Maru"
    if distribution_action == "routing_review":
        return "Sin bucket"
    if distribution_action == "error":
        return "Error técnico"
    if distribution_action == "not_applicable":
        return "Rechazado sin distribución"
    if strategy in {
        "outside_hours_manual",
        "commercial_rejection_manual",
        "no_online_sellers_manual",
        "assignment_queue_closed_manual",
    }:
        return "Gestión manual con Maru"
    if strategy == "rejection_without_distribution":
        return "Rechazado sin distribución"
    if strategy == "assignment_queue":
        return "En cola de distribución"
    if strategy == "no_matching_bucket" or action == "routing_review":
        return "Sin bucket"
    if action == "skipped":
        return "Omitido"
    if strategy in {
        "contact_history",
        "legacy_contact_history",
        "round_robin",
        "legacy_round_robin",
        "round_robin_initial",
        "single_seller",
    }:
        return "Distribuido"
    if action == "manual_review":
        return "Revisión manual"
    return "Sin distribución"


def normalized(row: dict[str, Any]) -> dict[str, Any] | None:
    outputs = row.get("outputs") or {}
    action = str(outputs.get("action") or "").strip().lower()
    technical_state = execution_state(row)
    deal_id = str(outputs.get("deal_id") or "").strip()
    if action == "no_pending" or (not deal_id and technical_state not in {"FAILED", "KILLED"}):
        return None

    strategy = str(outputs.get("assignment_strategy") or "").strip()
    commercial_action = str(outputs.get("commercial_action") or "").strip().lower()
    if not commercial_action and action in {
        "approved", "rejected", "commercial_rejected", "manual_review"
    }:
        commercial_action = action
    commercial_reason = str(outputs.get("commercial_reason") or "").strip()
    if not commercial_reason and commercial_action:
        commercial_reason = str(outputs.get("reason") or "").strip()
    distribution_action = str(outputs.get("distribution_action") or "").strip()
    distribution_reason = str(outputs.get("distribution_reason") or "").strip()
    rule_version = str(outputs.get("rule_version") or "").strip()
    assigned_by_id = str(outputs.get("assigned_by_id") or "").strip()
    assigned_by_name = str(outputs.get("assigned_by_name") or "").strip()
    if assigned_by_id and not assigned_by_name:
        assigned_by_name = f"Usuario {assigned_by_id}"
    processed_at = parse_datetime(outputs.get("processed_at")) or parse_datetime(
        (row.get("state") or {}).get("startDate")
    )
    message = str(outputs.get("message") or "")
    status = distribution_status(
        action,
        strategy,
        technical_state,
        message,
        distribution_action,
    )
    if not strategy and not rule_version and technical_state == "SUCCESS":
        status = "Histórico incompleto"
    return {
        "processed_at": processed_at,
        "execution_id": str(row.get("id") or ""),
        "flow_id": str(row.get("flowId") or ""),
        "revision": row.get("flowRevision"),
        "technical_state": technical_state,
        "action": action,
        "commercial_action": commercial_action,
        "commercial_reason": commercial_reason,
        "commercial_stage_id": str(outputs.get("commercial_stage_id") or ""),
        "distribution_action": distribution_action,
        "distribution_reason": distribution_reason,
        "distribution_status": status,
        "reason": str(outputs.get("reason") or ""),
        "message": message,
        "rule_version": rule_version,
        "deal_id": deal_id,
        "deal_title": str(outputs.get("deal_title") or ""),
        "lead_id": str(outputs.get("lead_id") or ""),
        "contact_id": str(outputs.get("contact_id") or ""),
        "province": str(outputs.get("province") or ""),
        "employment_status": str(outputs.get("employment_status") or ""),
        "payment_bank": str(outputs.get("payment_bank") or ""),
        "stage_before": str(outputs.get("stage_before") or ""),
        "stage_after": str(outputs.get("stage_id") or ""),
        "commercial_line": str(outputs.get("commercial_line") or ""),
        "routing_bucket": str(outputs.get("routing_bucket") or ""),
        "previous_assigned_by_id": str(outputs.get("previous_assigned_by_id") or ""),
        "assigned_by_id": assigned_by_id,
        "assigned_by_name": assigned_by_name,
        "assignment_strategy": strategy,
        "assignment_reason": ASSIGNMENT_STRATEGY_LABELS.get(strategy, strategy),
        "configured_pool": str(outputs.get("configured_pool") or ""),
        "online_pool": str(outputs.get("online_pool") or ""),
        "within_business_hours": parse_bool(outputs.get("within_business_hours")),
        "linked_activity_count": parse_int(outputs.get("linked_activity_count")),
        "transferred_chat_count": parse_int(outputs.get("transferred_chat_count")),
        "chat_transfer_status": str(outputs.get("chat_transfer_status") or ""),
        "found_chat_ids": str(outputs.get("found_chat_ids") or ""),
        "transferred_chat_ids": str(outputs.get("transferred_chat_ids") or ""),
        "skipped_chat_ids": str(outputs.get("skipped_chat_ids") or ""),
        "skipped_chat_reasons": str(outputs.get("skipped_chat_reasons") or ""),
        "bcra_snapshot_checked_at": parse_datetime(
            outputs.get("bcra_snapshot_checked_at")
        ),
        "bcra_snapshot_age_days": outputs.get("bcra_snapshot_age_days", ""),
        "bcra_snapshot_refreshed": parse_bool(
            outputs.get("bcra_snapshot_refreshed")
        ),
        "bcra_refresh_outcome": str(outputs.get("bcra_refresh_outcome") or ""),
        "bcra_retry_attempts": parse_int(outputs.get("bcra_retry_attempts")),
        "bcra_next_retry_at": parse_datetime(outputs.get("bcra_next_retry_at")),
    }


def normalized_events(row: dict[str, Any]) -> list[dict[str, Any]]:
    outputs = row.get("outputs") or {}
    raw_events = outputs.get("events_json")
    if not raw_events:
        event = normalized(row)
        return [event] if event is not None else []
    try:
        queue_events = json.loads(raw_events)
    except (TypeError, ValueError):
        return []
    result: list[dict[str, Any]] = []
    for index, event_outputs in enumerate(queue_events):
        if not isinstance(event_outputs, dict):
            continue
        # El estado "waiting" se reintenta cada minuto; conservarlo inflaría el
        # informe sin agregar una nueva decisión comercial.
        if event_outputs.get("action") == "queue_waiting":
            continue
        child = {
            **row,
            "id": f"{row.get('id')}:queue:{index}",
            "outputs": event_outputs,
        }
        event = normalized(child)
        if event is not None:
            result.append(event)
    return result


def events_from(
    rows: list[dict[str, Any]],
    audit_from: datetime,
) -> list[dict[str, Any]]:
    """Keep only auditable events at or after the Argentina-local cutoff."""
    return [
        row
        for row in rows
        if row["processed_at"] is not None and row["processed_at"] >= audit_from
    ]


def user_ids_in(rows: list[dict[str, Any]]) -> set[str]:
    result: set[str] = set()
    for row in rows:
        for key in ("previous_assigned_by_id", "assigned_by_id"):
            value = str(row.get(key) or "").strip()
            if value.isdigit():
                result.add(value)
        for key in ("configured_pool", "online_pool"):
            result.update(
                value
                for value in (part.strip() for part in str(row.get(key) or "").split(","))
                if value.isdigit()
            )
    return result


def user_display(user_id: str, names: dict[str, str], fallback_name: str = "") -> str:
    if not user_id:
        return ""
    name = fallback_name.strip() or names.get(user_id, "").strip() or f"Usuario {user_id}"
    return f"{name} ({user_id})"


def pool_display(value: str, names: dict[str, str]) -> str:
    return ", ".join(
        user_display(user_id, names)
        for user_id in (part.strip() for part in value.split(","))
        if user_id
    )


def add_user_displays(
    rows: list[dict[str, Any]],
    names: dict[str, str],
) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item["previous_assigned_by"] = user_display(
            item["previous_assigned_by_id"], names
        )
        item["assigned_by"] = user_display(
            item["assigned_by_id"], names, item["assigned_by_name"]
        )
        item["configured_pool_display"] = pool_display(item["configured_pool"], names)
        item["online_pool_display"] = pool_display(item["online_pool"], names)
        enriched.append(item)
    return enriched


def business_decision(row: dict[str, Any]) -> str:
    action = row["commercial_action"]
    if action == "pending_data":
        return "Pendiente de información BCRA"
    if action == "approved":
        line = row["commercial_line"] or "comercial definida"
        return f"Asignado a la línea {line}"
    if action in {"rejected", "commercial_rejected"}:
        return "Rechazado"
    if action == "manual_review":
        if row["assigned_by_id"] == "57":
            return "Enviado a revisión manual con Maru"
        return "Enviado a revisión manual"
    if action == "routing_review":
        return "Enviado a revisión de enrutamiento"
    if row["distribution_status"] == "Sin vendedor disponible":
        return "No asignado automáticamente"
    if row["distribution_status"] == "Error técnico":
        return "Procesamiento incompleto"
    if action == "skipped":
        return "Sin cambios"
    return "Sin decisión comercial"


def business_reason(row: dict[str, Any]) -> str:
    if row["distribution_status"] == "Sin vendedor disponible":
        return "No había vendedores del grupo conectados en Bitrix."
    reason = REASON_LABELS.get(row["commercial_reason"])
    if reason:
        return reason
    if row["message"]:
        return row["message"]
    if row["commercial_reason"]:
        return f"Motivo pendiente de descripción comercial: {row['commercial_reason']}."
    return "La ejecución no dejó información suficiente para explicar la decisión."


def add_business_fields(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        item["business_decision"] = business_decision(item)
        item["business_reason"] = business_reason(item)
        item["stage_before_display"] = STAGE_LABELS.get(
            item["stage_before"], item["stage_before"]
        )
        item["stage_after_display"] = STAGE_LABELS.get(
            item["stage_after"], item["stage_after"]
        )
        result.append(item)
    return result


def latest_cases(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in sorted(
        rows, key=lambda item: item["processed_at"] or datetime.min, reverse=True
    ):
        key = row["deal_id"] or f"execution:{row['execution_id']}"
        result.setdefault(key, row)
    return list(result.values())


def current_versions(rows: list[dict[str, Any]]) -> dict[str, tuple[str, Any]]:
    result: dict[str, tuple[str, Any]] = {}
    for row in sorted(
        rows, key=lambda item: item["processed_at"] or datetime.min, reverse=True
    ):
        if row["rule_version"] and row["revision"] is not None:
            result.setdefault(
                row["flow_id"] or "Flow sin identificar",
                (row["rule_version"], row["revision"]),
            )
    return result


def latest_rule_version(rows: list[dict[str, Any]]) -> str:
    for row in sorted(
        rows, key=lambda item: item["processed_at"] or datetime.min, reverse=True
    ):
        if row["rule_version"]:
            return row["rule_version"]
    return ""


def mute_previous_version(
    ws,
    row_number: int,
    item: dict[str, Any],
    current: dict[str, tuple[str, Any]],
) -> None:
    flow_version = current.get(item["flow_id"] or "Flow sin identificar")
    if flow_version is None or (item["rule_version"], item["revision"]) == flow_version:
        return
    for cell in ws[row_number]:
        cell.fill = PatternFill("solid", fgColor=MUTED_FILL)
        cell.font = Font(color=MUTED_FONT)


def compact(ws, widths: dict[int, int] | None = None) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = widths or {}
    for index, column in enumerate(ws.columns, 1):
        longest = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(
            longest, widths.get(index, 38)
        )


def portal_base(api_base_url: str) -> str:
    base = api_base_url.strip().rstrip("/")
    return base[:-5] if base.lower().endswith("/rest") else base


def add_link(cell, url: str) -> None:
    if not cell.value or not url:
        return
    cell.hyperlink = url
    cell.style = "Hyperlink"


def chat_transfer_display(item: dict[str, Any]) -> str:
    status = item["chat_transfer_status"]
    if status:
        return CHAT_TRANSFER_STATUS_LABELS.get(status, status)
    if item["transferred_chat_count"] > 0:
        return "Transferido; detalle no disponible en la traza histórica"
    if item["distribution_status"] == "Distribuido":
        return "Sin detalle histórico"
    return "No aplica: no hubo distribución"


def skipped_chat_reason_display(value: str) -> str:
    result: list[str] = []
    for entry in value.split(";"):
        entry = entry.strip()
        if not entry:
            continue
        chat_id, separator, reason = entry.partition(":")
        label = CHAT_SKIP_REASON_LABELS.get(reason, reason)
        result.append(f"Chat {chat_id}: {label}" if separator else entry)
    return "; ".join(result)


def build(
    rows: list[dict[str, Any]],
    bitrix_base_url: str = "",
    audit_from: datetime | None = None,
) -> Workbook:
    if rows and "assigned_by" not in rows[0]:
        rows = add_user_displays(rows, {})
    rows = add_business_fields(rows)
    cases = latest_cases(rows)
    versions = current_versions(rows)
    rule_version = latest_rule_version(rows)
    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen"
    status_counts = Counter(row["distribution_status"] for row in cases)
    summary.append(["AUDITORÍA COMERCIAL Y DISTRIBUCIÓN", "Valor"])
    summary_rows = (
        ("Eventos incluidos desde", audit_from or "Sin corte configurado"),
        ("Negociaciones incluidas", len(cases)),
        ("Ejecuciones registradas", len(rows)),
        ("Distribuidos", status_counts["Distribuido"]),
        ("Gestión manual con Maru", status_counts["Gestión manual con Maru"]),
        ("Sin vendedor disponible", status_counts["Sin vendedor disponible"]),
        ("Sin bucket", status_counts["Sin bucket"]),
        ("Errores técnicos", status_counts["Error técnico"]),
        ("Históricos sin trazabilidad completa", status_counts["Histórico incompleto"]),
        ("Chats transferidos", sum(row["transferred_chat_count"] for row in cases)),
        ("Versión de reglas vigente", rule_version or "Sin datos"),
        (
            "Flows y revisiones vigentes",
            "; ".join(
                f"{flow_id}: rev. {version[1]}"
                for flow_id, version in sorted(versions.items())
            ) or "Sin datos",
        ),
        (
            "Lectura de versiones",
            "Las filas se atenúan solo si usaron una revisión anterior del mismo flow.",
        ),
    )
    for item in summary_rows:
        summary.append(item)
    compact(summary, {1: 38, 2: 72})
    summary.auto_filter.ref = None
    for row_number in range(2, summary.max_row + 1):
        summary.cell(row_number, 1).font = Font(bold=True, color=NAVY)
    if audit_from is not None:
        summary.cell(2, 2).number_format = "dd/mm/yyyy hh:mm:ss"

    by_seller = wb.create_sheet("Por vendedor")
    by_seller.append(["Responsable", "Negociaciones", "Distribuidas", "Chats transferidos"])
    seller_rows: dict[str, list[dict[str, Any]]] = {}
    for row in cases:
        key = row["assigned_by"]
        if key:
            seller_rows.setdefault(key, []).append(row)
    for seller, items in sorted(
        seller_rows.items(), key=lambda item: (-len(item[1]), item[0])
    ):
        by_seller.append([
            seller,
            len(items),
            sum(item["distribution_status"] == "Distribuido" for item in items),
            sum(item["transferred_chat_count"] for item in items),
        ])
    compact(by_seller, {1: 36})
    if by_seller.max_row > 1:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Distribuciones por responsable"
        chart.add_data(
            Reference(by_seller, min_col=3, min_row=1, max_row=by_seller.max_row),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(by_seller, min_col=1, min_row=2, max_row=by_seller.max_row)
        )
        by_seller.add_chart(chart, "G2")

    business_headers = [
        "Fecha y hora", "Decisión tomada", "Razón de la decisión", "Negociación",
        "Cliente", "Provincia", "Situación laboral", "Banco de cobro",
        "Estado de consulta BCRA", "Fecha de consulta BCRA", "Antigüedad BCRA (días)",
        "Intentos BCRA", "Próximo intento BCRA",
        "Línea comercial", "Grupo de distribución", "Responsable", "Responsable anterior",
        "Resultado de distribución", "Motivo de asignación", "Etapa anterior",
        "Etapa resultante", "Chat transferido", "Resultado de chats",
        "Chats encontrados", "Chats transferidos", "Chats no transferidos",
        "Motivo sin transferencia", "Versión de reglas", "Flow Kestra",
        "Revisión del flow Kestra",
    ]
    business_keys = [
        "processed_at", "business_decision", "business_reason", "deal_id", "deal_title",
        "province", "employment_status", "payment_bank", "bcra_refresh_display",
        "bcra_snapshot_checked_at", "bcra_snapshot_age_days", "bcra_retry_attempts",
        "bcra_next_retry_at", "commercial_line",
        "routing_bucket", "assigned_by", "previous_assigned_by", "distribution_status",
        "assignment_reason", "stage_before_display", "stage_after_display",
        "chat_transferred_display", "chat_transfer_display", "found_chat_ids_display",
        "transferred_chat_ids_display", "skipped_chat_ids_display",
        "skipped_chat_reason_display", "rule_version", "flow_id", "revision",
    ]
    technical_headers = [
        "Fecha y hora", "Estado de distribución", "Acción técnica", "Código de motivo",
        "Decisión comercial", "Motivo comercial", "Etapa comercial objetivo",
        "Acción de distribución", "Motivo de distribución",
        "Versión de reglas", "Flow Kestra", "Revisión del flow Kestra",
        "Negociación", "Título",
        "Lead", "Contacto", "Provincia", "Situación laboral", "Banco de cobro",
        "Estado de consulta BCRA", "Fecha de consulta BCRA", "Antigüedad BCRA (días)",
        "Intentos BCRA", "Próximo intento BCRA",
        "Etapa anterior", "Etapa resultante", "Línea comercial", "Grupo de distribución",
        "Responsable anterior", "Responsable", "Motivo de asignación", "Estrategia técnica",
        "Pool configurado", "Pool online", "Dentro del horario laboral",
        "Resultado de chats", "Chats encontrados", "Chats transferidos",
        "Chats no transferidos", "Motivo sin transferencia",
        "Actividades vinculadas", "Ejecución Kestra",
        "Estado técnico", "Mensaje técnico",
    ]
    technical_keys = [
        "processed_at", "distribution_status", "action", "reason",
        "commercial_action", "commercial_reason", "commercial_stage_id",
        "distribution_action", "distribution_reason", "rule_version", "flow_id",
        "revision", "deal_id", "deal_title", "lead_id", "contact_id", "province",
        "employment_status",
        "payment_bank", "bcra_refresh_display", "bcra_snapshot_checked_at",
        "bcra_snapshot_age_days", "bcra_retry_attempts", "bcra_next_retry_at",
        "stage_before", "stage_after", "commercial_line",
        "routing_bucket", "previous_assigned_by", "assigned_by", "assignment_reason",
        "assignment_strategy", "configured_pool_display", "online_pool_display",
        "within_business_hours", "chat_transfer_display", "found_chat_ids_display",
        "transferred_chat_ids_display", "skipped_chat_ids_display",
        "skipped_chat_reason_display", "linked_activity_count", "execution_id",
        "technical_state", "message",
    ]

    for item in rows:
        has_current_chat_trace = bool(item["chat_transfer_status"])
        item["chat_transferred_display"] = "Sí" if (
            item["chat_transfer_status"] in {"transferred", "partially_transferred"}
            if has_current_chat_trace
            else item["transferred_chat_count"] > 0
        ) else "No"
        item["chat_transfer_display"] = chat_transfer_display(item)
        historical_transfer_without_ids = (
            not has_current_chat_trace and item["transferred_chat_count"] > 0
        )
        item["found_chat_ids_display"] = (
            item["found_chat_ids"]
            or ("No registrado" if historical_transfer_without_ids else "")
        )
        item["transferred_chat_ids_display"] = (
            item["transferred_chat_ids"]
            or ("No registrado" if historical_transfer_without_ids else "")
        )
        item["skipped_chat_ids_display"] = item["skipped_chat_ids"]
        item["skipped_chat_reason_display"] = skipped_chat_reason_display(
            item["skipped_chat_reasons"]
        )
        item["bcra_refresh_display"] = {
            "reused_fresh": "Consulta vigente reutilizada",
            "ok": "Consulta actualizada",
            "not_found": "Consulta actualizada sin datos",
            "invalid_identification": "Identificación inválida",
            "missing_cuil": "No se pudo consultar: falta CUIL",
            "temporary_error": "No se pudo actualizar",
            "rate_limited": "No se pudo actualizar: límite del servicio",
            "retry_scheduled": "Reintento programado",
            "retry_exhausted": "Reintentos agotados",
            "error": "No se pudo actualizar",
            "not_evaluated": "No evaluada",
        }.get(item["bcra_refresh_outcome"], item["bcra_refresh_outcome"] or "No evaluada")

    cases_sheet = wb.create_sheet("Casos")
    cases_sheet.append(business_headers)
    exceptions = wb.create_sheet("Excepciones")
    exceptions.append(business_headers)
    technical = wb.create_sheet("Trazabilidad técnica")
    technical.append(technical_headers)
    legacy = wb.create_sheet("Histórico incompleto")
    legacy.append(technical_headers)
    base = portal_base(bitrix_base_url)

    for item in cases:
        values = [item[key] for key in business_keys]
        target_sheets = [cases_sheet]
        if item["distribution_status"] not in {"Distribuido", "Histórico incompleto"}:
            target_sheets.append(exceptions)
        for sheet in target_sheets:
            sheet.append(values)
            row_number = sheet.max_row
            sheet.cell(row_number, 1).number_format = "dd/mm/yyyy hh:mm:ss"
            sheet.cell(row_number, 13).number_format = "dd/mm/yyyy hh:mm:ss"
            for cell in sheet[row_number]:
                cell.alignment = Alignment(vertical="center")
            sheet.cell(row_number, 3).alignment = Alignment(
                wrap_text=False, vertical="center"
            )
            sheet.row_dimensions[row_number].height = 20
            mute_previous_version(sheet, row_number, item, versions)
            add_link(
                sheet.cell(row_number, 4),
                f"{base}/crm/deal/details/{item['deal_id']}/" if base else "",
            )

    for item in sorted(
        rows, key=lambda row: row["processed_at"] or datetime.min, reverse=True
    ):
        values = [item[key] for key in technical_keys]
        target_sheets = [technical]
        if item["distribution_status"] == "Histórico incompleto":
            target_sheets.append(legacy)
        for sheet in target_sheets:
            sheet.append(values)
            row_number = sheet.max_row
            sheet.cell(row_number, 1).number_format = "dd/mm/yyyy hh:mm:ss"
            mute_previous_version(sheet, row_number, item, versions)
            add_link(
                sheet.cell(row_number, 13),
                f"{base}/crm/deal/details/{item['deal_id']}/" if base else "",
            )
            add_link(
                sheet.cell(row_number, 15),
                f"{base}/crm/lead/details/{item['lead_id']}/" if base else "",
            )

    business_widths = {2: 36, 3: 72, 5: 34, 7: 30, 8: 38, 11: 38, 12: 38, 14: 48}
    technical_widths = {
        4: 36, 6: 48, 9: 48, 13: 34, 17: 30, 18: 38,
        25: 48, 27: 60, 28: 60, 34: 60, 37: 60,
    }
    compact(cases_sheet, business_widths)
    compact(exceptions, business_widths)
    compact(technical, technical_widths)
    compact(legacy, technical_widths)
    return wb


def publish(
    workbook: Workbook,
    root: Path,
    generated_at: datetime,
) -> tuple[Path, Path]:
    report_dir = root / "marketing" / "distribucion-negociaciones"
    history_dir = report_dir / "historico"
    history_dir.mkdir(parents=True, exist_ok=True)
    dated = history_dir / f"{generated_at:%Y-%m-%d}.xlsx"
    latest = report_dir / "ultimo.xlsx"
    with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, dated)
        dated.chmod(0o644)
        with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
            latest_temporary = Path(handle.name)
        shutil.copy2(dated, latest_temporary)
        os.replace(latest_temporary, latest)
        latest.chmod(0o644)
    finally:
        temporary.unlink(missing_ok=True)
    return latest, dated


def main() -> None:
    base = os.environ["REPORTS_KESTRA_URL"].rstrip("/")
    tenant = os.getenv("REPORTS_KESTRA_TENANT", "main")
    namespace = os.getenv("REPORTS_NAMESPACE", "redunisol.prod.marketing-crm")
    session = requests.Session()
    session.auth = (
        os.environ["REPORTS_KESTRA_USERNAME"],
        os.environ["REPORTS_KESTRA_PASSWORD"],
    )
    audit_from_value = os.getenv("REPORTS_AUDIT_FROM", DEFAULT_AUDIT_FROM)
    audit_from = parse_datetime(audit_from_value)
    if audit_from is None:
        raise RuntimeError(
            "REPORTS_AUDIT_FROM debe ser una fecha ISO válida; "
            f"valor recibido: {audit_from_value!r}"
        )
    raw = [
        {**row, "flowId": row.get("flowId") or flow_id}
        for flow_id in FLOW_IDS
        for row in executions(session, base, tenant, namespace, flow_id)
    ]
    normalized_rows = [
        event
        for row in raw
        for event in normalized_events(row)
    ]
    rows = events_from(normalized_rows, audit_from)
    bitrix_api_url = "/".join(
        part.strip("/")
        for part in (
            os.getenv("REPORTS_BITRIX_BASE_URL", ""),
            os.getenv("REPORTS_BITRIX_WEBHOOK_PATH", ""),
        )
        if part.strip("/")
    )
    try:
        names = bitrix_user_names(bitrix_api_url, user_ids_in(rows))
    except (requests.RequestException, RuntimeError, ValueError):
        names = {}
    rows = add_user_displays(rows, names)
    latest, dated = publish(
        build(rows, os.getenv("REPORTS_BITRIX_BASE_URL", ""), audit_from),
        Path(os.getenv("REPORTS_ROOT", "/reports")),
        datetime.now(ARGENTINA_TIMEZONE).replace(tzinfo=None),
    )
    print(json.dumps({
        "ok": True,
        "audit_from": audit_from.isoformat(),
        "events": len(rows),
        "latest": str(latest),
        "history": str(dated),
    }))


if __name__ == "__main__":
    main()
