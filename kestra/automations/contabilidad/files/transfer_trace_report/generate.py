from __future__ import annotations

import json
import os
import shutil
import tempfile
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time as datetime_time, timedelta
from pathlib import Path
from statistics import mean
from typing import Any, Iterable
from zoneinfo import ZoneInfo

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from kestra import Kestra
except ImportError:  # pragma: no cover - optional outside Kestra
    Kestra = None


ARGENTINA_TIMEZONE = ZoneInfo("America/Argentina/Buenos_Aires")
DEFAULT_REPORT_ROOT = "/reports"
DEFAULT_COVERAGE_FROM = "2026-09-03"
NAVY, BLUE, GREEN, ORANGE, WHITE = "17365D", "4472C4", "70AD47", "ED7D31", "FFFFFF"
TRANSFER_EVENT_TYPES = {
    "transfer_started", "pre_transfer_validation", "transfer_payload_built",
    "transfer_response_received", "initial_response_classified",
    "confirmation_poll_skipped_already_confirmed", "confirmation_poll_skipped_terminal_rejection",
    "confirmation_poll_result", "confirmation_poll_error", "confirmation_finished",
    "receipt_written", "receipt_write_failed", "mark_paid_request_started",
    "mark_paid_request_succeeded", "mark_paid_request_failed",
    "mark_paid_request_skipped_missing_receipt", "cancellation_leg_payload_built",
    "cancellation_leg_already_transferred",
}


@dataclass
class Operation:
    started_at: datetime
    finished_at: datetime | None
    request_oid: str
    mode: str
    operation_type: str
    result: str
    flow_duration_seconds: float | None
    paid_at: datetime | None
    credit_line: str
    credit_line_id: str
    amount: float | None
    operator: str
    session_id: str
    client_instance_id: str
    application_version: str
    external_ids: tuple[str, ...]
    cancellation_legs: int
    event_count: int


@dataclass
class Candidate:
    request_oid: str
    first_observed_at: datetime
    last_observed_at: datetime
    observations: int
    request_status: str
    operation_type: str
    credit_line: str
    credit_line_id: str
    amount: float | None
    completed_via_app: bool
    paid_at: datetime | None
    paid_mode: str
    paid_operation_type: str
    time_to_paid_seconds: float | None


def env(name: str, default: str = "") -> str:
    return os.getenv(name, default).strip()


def require_env(name: str) -> str:
    value = env(name)
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def parse_run_date(raw: str | None = None, *, today: date | None = None) -> date:
    value = (raw if raw is not None else env("TRANSFER_TRACE_REPORT_DATE")).strip()
    if not value:
        return (today or datetime.now(ARGENTINA_TIMEZONE).date()) - timedelta(days=1)
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise RuntimeError("TRANSFER_TRACE_REPORT_DATE must use YYYY-MM-DD format") from exc


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str) and value.strip():
        try:
            parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        except ValueError:
            return None
    else:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=ARGENTINA_TIMEZONE)
    return parsed.astimezone(ARGENTINA_TIMEZONE)


def api_get(session: requests.Session, url: str, **params: Any) -> Any:
    last_error: Exception | None = None
    for attempt in range(5):
        try:
            response = session.get(url, params=params, timeout=(10, 90))
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt < 4:
                time.sleep(2**attempt)
    raise RuntimeError(f"No se pudo consultar la trazabilidad de transferencias: {last_error}")


def fetch_trace_events(session: requests.Session, base_url: str, occurred_from: datetime, occurred_to: datetime, *, event_type: str | None = None, page_size: int = 500) -> list[dict[str, Any]]:
    offset = 0
    result: list[dict[str, Any]] = []
    while True:
        params = {"limit": page_size, "offset": offset, "occurred_from": occurred_from.isoformat(), "occurred_to": occurred_to.isoformat()}
        if event_type:
            params["event_type"] = event_type
        payload = api_get(session, f"{base_url.rstrip('/')}/api/v1/transfer-trace-events", **params)
        batch = payload.get("items") or []
        result.extend(batch)
        total = int((payload.get("pagination") or {}).get("total", len(result)))
        if not batch or len(result) >= total:
            break
        offset += page_size
    return result


def normalize_mode(value: Any) -> str:
    normalized = str(value or "").strip().casefold()
    if normalized == "manual":
        return "Manual"
    if normalized in {"automatico", "automatica", "automático", "automática", "automatic"}:
        return "Automática"
    return "Sin modalidad"


def normalize_identifier(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def supports_candidate_observation(version: Any) -> bool:
    try:
        parts = tuple(int(part) for part in str(version or "").split(".")[:3])
    except ValueError:
        return False
    return parts + (0,) * (3 - len(parts)) >= (2, 0, 1)


def has_candidate_coverage(events: Iterable[dict[str, Any]]) -> bool:
    return any(
        item.get("event_type") == "transfer_candidate_observed"
        or supports_candidate_observation(item.get("application_version"))
        for item in events
    )


def parse_amount(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if not isinstance(value, str) or not value.strip():
        return None
    normalized = value.strip().replace("$", "").replace(" ", "")
    if "," in normalized and "." in normalized:
        normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def event_status_classification(event: dict[str, Any]) -> str:
    status = (event.get("data") or {}).get("status")
    if isinstance(status, dict):
        return str(status.get("classification") or "").strip().casefold()
    return str(status or "").strip().casefold()


def operation_result(events: list[dict[str, Any]]) -> str:
    event_types = {str(item.get("event_type") or "") for item in events}
    if "mark_paid_request_succeeded" in event_types:
        return "Completada"
    if "mark_paid_request_failed" in event_types or "mark_paid_request_skipped_missing_receipt" in event_types:
        return "Transferida; registro pendiente"
    classifications = {event_status_classification(item) for item in events}
    if "confirmed" in classifications or "mark_paid_request_started" in event_types:
        return "Transferida; registro pendiente"
    if "rejected" in classifications:
        return "Rechazada"
    if "pending" in classifications:
        return "Pendiente"
    validations = [(item.get("data") or {}).get("can_transfer") for item in events if item.get("event_type") == "pre_transfer_validation"]
    if any(value is False for value in validations):
        return "Bloqueada"
    if "transfer_response_received" in event_types:
        return "En curso"
    return "Iniciada sin cierre"


def terminal_time(events: list[dict[str, Any]]) -> datetime | None:
    terminal_types = {"mark_paid_request_succeeded", "mark_paid_request_failed", "mark_paid_request_skipped_missing_receipt", "confirmation_finished", "receipt_write_failed"}
    values = [parse_datetime(item.get("occurred_at")) for item in events if item.get("event_type") in terminal_types]
    return max((value for value in values if value is not None), default=None)


def event_time(events: Iterable[dict[str, Any]], event_type: str) -> datetime | None:
    values = [parse_datetime(item.get("occurred_at")) for item in events if item.get("event_type") == event_type]
    return max((value for value in values if value is not None), default=None)


def external_ids(events: Iterable[dict[str, Any]]) -> tuple[str, ...]:
    values: set[str] = set()
    for event in events:
        data = event.get("data") or {}
        for key in ("id_coelsa", "idCoelsa"):
            value = normalize_identifier(data.get(key))
            if value:
                values.add(value)
    return tuple(sorted(values))


def build_operations(events: list[dict[str, Any]]) -> list[Operation]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        request_oid = normalize_identifier(event.get("request_oid"))
        if request_oid and event.get("event_type") in TRANSFER_EVENT_TYPES:
            grouped[(str(event.get("session_id") or ""), request_oid)].append(event)
    operations: list[Operation] = []
    minimum = datetime.min.replace(tzinfo=ARGENTINA_TIMEZONE)
    for (session_id, request_oid), request_events in grouped.items():
        ordered = sorted(request_events, key=lambda item: parse_datetime(item.get("occurred_at")) or minimum)
        starts = [index for index, item in enumerate(ordered) if item.get("event_type") == "transfer_started"]
        for position, start_index in enumerate(starts):
            next_index = starts[position + 1] if position + 1 < len(starts) else len(ordered)
            window = ordered[start_index:next_index]
            start_event = window[0]
            started_at = parse_datetime(start_event.get("occurred_at"))
            if started_at is None:
                continue
            finished_at = terminal_time(window)
            paid_at = event_time(window, "mark_paid_request_succeeded")
            data = start_event.get("data") or {}
            is_cancellation = any(str(item.get("event_type") or "").startswith("cancellation_") for item in window)
            legs = {normalize_identifier((item.get("data") or {}).get("leg")) for item in window if str(item.get("event_type") or "").startswith("cancellation_leg_")}
            legs.discard("")
            duration = (finished_at - started_at).total_seconds() if finished_at else None
            operations.append(Operation(
                started_at=started_at, finished_at=finished_at, request_oid=request_oid,
                mode=normalize_mode(start_event.get("mode")), operation_type="Cancelación" if is_cancellation else "Transferencia",
                result=operation_result(window), flow_duration_seconds=max(duration, 0.0) if duration is not None else None,
                paid_at=paid_at,
                credit_line=str(data.get("credit_line") or ""), credit_line_id=normalize_identifier(data.get("credit_line_id")),
                amount=parse_amount(data.get("transfer_amount")), operator=str(start_event.get("operator") or data.get("operator") or ""),
                session_id=session_id, client_instance_id=str(start_event.get("client_instance_id") or ""),
                application_version=str(start_event.get("application_version") or ""), external_ids=external_ids(window),
                cancellation_legs=len(legs), event_count=len(window),
            ))
    return sorted(operations, key=lambda item: item.started_at)


def build_candidates(events: list[dict[str, Any]], operations: list[Operation]) -> list[Candidate]:
    observations: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        if event.get("event_type") == "transfer_candidate_observed":
            request_oid = normalize_identifier(event.get("request_oid"))
            if request_oid:
                observations[request_oid].append(event)
    completed_by_oid: dict[str, Operation] = {}
    paid_by_oid: dict[str, list[Operation]] = defaultdict(list)
    for operation in operations:
        if operation.result in {"Completada", "Transferida; registro pendiente"}:
            current = completed_by_oid.get(operation.request_oid)
            if current is None or (operation.finished_at or operation.started_at) < (current.finished_at or current.started_at):
                completed_by_oid[operation.request_oid] = operation
        if operation.paid_at is not None:
            paid_by_oid[operation.request_oid].append(operation)
    candidates: list[Candidate] = []
    maximum = datetime.max.replace(tzinfo=ARGENTINA_TIMEZONE)
    for request_oid, rows in observations.items():
        ordered = sorted(rows, key=lambda item: parse_datetime(item.get("occurred_at")) or maximum)
        first, last = ordered[0], ordered[-1]
        first_at, last_at = parse_datetime(first.get("occurred_at")), parse_datetime(last.get("occurred_at"))
        if first_at is None or last_at is None:
            continue
        data = last.get("data") or {}
        completed = completed_by_oid.get(request_oid)
        paid = min(
            (operation for operation in paid_by_oid.get(request_oid, []) if operation.paid_at >= first_at),
            key=lambda operation: operation.paid_at,
            default=None,
        )
        paid_at = paid.paid_at if paid else None
        time_to_paid = (paid_at - first_at).total_seconds() if paid_at else None
        candidates.append(Candidate(
            request_oid=request_oid, first_observed_at=first_at, last_observed_at=last_at, observations=len(ordered),
            request_status=str(data.get("request_status") or "A Transferir"),
            operation_type="Cancelación" if data.get("is_cancellation") is True else "Transferencia",
            credit_line=str(data.get("credit_line") or ""), credit_line_id=normalize_identifier(data.get("credit_line_id")),
            amount=parse_amount(data.get("transfer_amount") or data.get("request_amount")), completed_via_app=completed is not None,
            paid_at=paid_at, paid_mode=paid.mode if paid else "", paid_operation_type=paid.operation_type if paid else "",
            time_to_paid_seconds=max(time_to_paid, 0.0) if time_to_paid is not None else None,
        ))
    return sorted(candidates, key=lambda item: (item.first_observed_at, item.request_oid))


def average_time_to_paid(candidates: Iterable[Candidate], *, mode: str | None = None, operation_type: str | None = None) -> float | None:
    values = [item.time_to_paid_seconds for item in candidates if item.time_to_paid_seconds is not None and (mode is None or item.paid_mode == mode) and (operation_type is None or item.paid_operation_type == operation_type)]
    return mean(values) if values else None


def distinct_completed(operations: Iterable[Operation], *, mode: str | None = None, operation_type: str | None = None) -> int:
    return len({item.request_oid for item in operations if item.result in {"Completada", "Transferida; registro pendiente"} and (mode is None or item.mode == mode) and (operation_type is None or item.operation_type == operation_type)})


def compact_sheet(sheet, *, widths: dict[int, int] | None = None, freeze: str = "A2") -> None:
    sheet.freeze_panes = freeze
    sheet.sheet_view.showGridLines = False
    if sheet.max_row and sheet.max_column:
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    limits = widths or {}
    for index, column in enumerate(sheet.columns, 1):
        longest = max(len(str(cell.value or "")) for cell in column) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(longest, limits.get(index, 34))


def configure_print(sheet, *, print_area: str | None = None) -> None:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4
    if print_area:
        sheet.print_area = print_area


def build_workbook(run_date: date, events: list[dict[str, Any]], operations: list[Operation], candidates: list[Candidate], *, coverage_from: date) -> Workbook:
    day_operations = [item for item in operations if item.started_at.date() == run_date]
    day_paid_candidates = [item for item in candidates if item.paid_at is not None and item.paid_at.date() == run_date]
    new_candidates = [item for item in candidates if item.first_observed_at.date() == run_date]
    pending_candidates = [item for item in candidates if item.first_observed_at.date() <= run_date and not item.completed_via_app]
    minimum = datetime.min.replace(tzinfo=ARGENTINA_TIMEZONE)
    day_events = [item for item in events if (parse_datetime(item.get("occurred_at")) or minimum).date() == run_date]
    candidate_coverage = has_candidate_coverage(events)

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:H1")
    summary["A1"] = "REPORTE DIARIO DE TRANSFERENCIAS"
    summary["A1"].fill = PatternFill("solid", fgColor=NAVY)
    summary["A1"].font = Font(color=WHITE, bold=True, size=16)
    summary["A1"].alignment = Alignment(horizontal="center")
    summary.row_dimensions[1].height = 34
    summary.merge_cells("A2:H2")
    summary["A2"] = f"Fecha operativa: {run_date:%d/%m/%Y} · Hora Argentina · Cobertura desde {coverage_from:%d/%m/%Y}"
    summary["A2"].alignment = Alignment(horizontal="center")
    metrics = [
        ("Solicitudes observadas nuevas", len(new_candidates) if candidate_coverage else "Sin cobertura", "Primera observación en A Transferir durante el día"),
        ("Pendientes no realizadas vía app", len(pending_candidates) if candidate_coverage else "Sin cobertura", "Observadas hasta el cierre sin finalización vía app"),
        ("Realizadas vía app", distinct_completed(day_operations), "OID únicos confirmados durante el día"),
        ("Realizadas manualmente", distinct_completed(day_operations, mode="Manual"), "OID únicos"),
        ("Realizadas automáticamente", distinct_completed(day_operations, mode="Automática"), "OID únicos"),
        ("Cancelaciones manuales", distinct_completed(day_operations, mode="Manual", operation_type="Cancelación"), "OID únicos completados"),
        ("Cancelaciones automáticas", distinct_completed(day_operations, mode="Automática", operation_type="Cancelación"), "OID únicos completados"),
        ("Pagadas con tiempo medible", len(day_paid_candidates), "Con primera detección en A Transferir y marca Pagada"),
        ("Eventos técnicos del día", len(day_events), "Eventos de trazabilidad recibidos"),
    ]
    summary.append([])
    summary.append(["Indicador", "Valor", "Unidad / definición"])
    for row in metrics:
        summary.append(row)
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    for row in range(5, 5 + len(metrics)):
        summary.cell(row, 1).font = Font(color=NAVY, bold=True)
        summary.cell(row, 2).alignment = Alignment(horizontal="right")
    summary.column_dimensions["A"].width, summary.column_dimensions["B"].width, summary.column_dimensions["C"].width = 40, 18, 44
    summary.column_dimensions["D"].width, summary.column_dimensions["E"].width = 18, 48
    mode_row = 16
    for column, value in enumerate(("Modalidad", "Realizadas", "Tiempo promedio detección → Pagada (min)", "Cancelaciones", "Tiempo prom. cancelación detección → Pagada (min)"), 1):
        summary.cell(mode_row, column, value)
    for cell in summary[mode_row][:5]:
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.font = Font(color=WHITE, bold=True)
    for row_index, mode in enumerate(("Manual", "Automática"), mode_row + 1):
        avg_all = average_time_to_paid(day_paid_candidates, mode=mode)
        avg_cancellation = average_time_to_paid(day_paid_candidates, mode=mode, operation_type="Cancelación")
        average_value = avg_all / 60 if avg_all is not None else ("Sin cobertura" if not candidate_coverage else "Sin datos")
        cancellation_average_value = avg_cancellation / 60 if avg_cancellation is not None else ("Sin cobertura" if not candidate_coverage else "Sin datos")
        values = (mode, distinct_completed(day_operations, mode=mode), average_value, distinct_completed(day_operations, mode=mode, operation_type="Cancelación"), cancellation_average_value)
        for column, value in enumerate(values, 1):
            summary.cell(row_index, column, value)
        summary.cell(row_index, 3).number_format = "0.00"
        summary.cell(row_index, 5).number_format = "0.00"
    chart = BarChart()
    chart.type, chart.title, chart.height, chart.width = "col", "Operaciones realizadas por modalidad", 7, 12
    chart.y_axis.title = "Solicitudes"
    chart.add_data(Reference(summary, min_col=2, min_row=mode_row, max_row=mode_row + 2), titles_from_data=True)
    chart.set_categories(Reference(summary, min_col=1, min_row=mode_row + 1, max_row=mode_row + 2))
    summary.add_chart(chart, "G4")
    has_duration_data = any(isinstance(summary.cell(row, 3).value, (int, float)) for row in range(mode_row + 1, mode_row + 3))
    if has_duration_data:
        duration_chart = BarChart()
        duration_chart.type, duration_chart.title, duration_chart.height, duration_chart.width = "bar", "Tiempo promedio: detección a Pagada (min)", 7, 12
        duration_chart.x_axis.title = "Minutos"
        duration_chart.add_data(Reference(summary, min_col=3, min_row=mode_row, max_row=mode_row + 2), titles_from_data=True)
        duration_chart.set_categories(Reference(summary, min_col=1, min_row=mode_row + 1, max_row=mode_row + 2))
        summary.add_chart(duration_chart, "G19")
    configure_print(summary, print_area="A1:R34" if has_duration_data else "A1:R18")

    operation_sheet = workbook.create_sheet("Operaciones app")
    operation_sheet.append(["Primera detección A Transferir", "Marcada Pagada", "Solicitud OID", "Modalidad", "Tipo", "Resultado", "Tiempo hasta Pagada segundos", "Tiempo hasta Pagada minutos", "Línea", "Línea ID", "Importe", "Operador", "Versión app", "ID externos", "Patas cancelación", "Eventos", "Inicio técnico del intento", "Fin técnico del intento", "Duración técnica segundos", "Sesión", "Instancia"])
    candidates_by_oid = {item.request_oid: item for item in candidates}
    for item in day_operations:
        candidate = candidates_by_oid.get(item.request_oid)
        detected_at = candidate.first_observed_at if candidate else None
        paid_at = item.paid_at
        time_to_paid = (paid_at - detected_at).total_seconds() if detected_at and paid_at else None
        operation_sheet.append([detected_at.replace(tzinfo=None) if detected_at else None, paid_at.replace(tzinfo=None) if paid_at else None, item.request_oid, item.mode, item.operation_type, item.result, max(time_to_paid, 0.0) if time_to_paid is not None else None, max(time_to_paid, 0.0) / 60 if time_to_paid is not None else None, item.credit_line, item.credit_line_id, item.amount, item.operator, item.application_version, ", ".join(item.external_ids), item.cancellation_legs, item.event_count, item.started_at.replace(tzinfo=None), item.finished_at.replace(tzinfo=None) if item.finished_at else None, item.flow_duration_seconds, item.session_id, item.client_instance_id])
    compact_sheet(operation_sheet, widths={1: 29, 2: 24, 7: 28, 8: 27, 9: 42, 12: 28, 14: 36, 17: 24, 18: 24, 20: 38, 21: 38})
    for row in range(2, operation_sheet.max_row + 1):
        operation_sheet.cell(row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
        operation_sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        operation_sheet.cell(row, 7).number_format = "0.00"
        operation_sheet.cell(row, 8).number_format = "0.00"
        operation_sheet.cell(row, 11).number_format = "#,##0.00"
        operation_sheet.cell(row, 17).number_format = "yyyy-mm-dd hh:mm:ss"
        operation_sheet.cell(row, 18).number_format = "yyyy-mm-dd hh:mm:ss"
        operation_sheet.cell(row, 19).number_format = "0.00"
    # Keep the printable view focused on the business timeline and amounts;
    # operator/version/technical identifiers remain available in the workbook.
    configure_print(operation_sheet, print_area=f"A1:K{operation_sheet.max_row}")

    candidates_sheet = workbook.create_sheet("Solicitudes observadas")
    candidates_sheet.append(["Primera observación", "Última observación", "Solicitud OID", "Estado observado", "Tipo", "Línea", "Línea ID", "Importe", "Observaciones", "Realizada vía app", "Marcada Pagada", "Modalidad Pagada", "Tiempo hasta Pagada segundos", "Tiempo hasta Pagada minutos"])
    for item in candidates:
        candidates_sheet.append([item.first_observed_at.replace(tzinfo=None), item.last_observed_at.replace(tzinfo=None), item.request_oid, item.request_status, item.operation_type, item.credit_line, item.credit_line_id, item.amount, item.observations, "Sí" if item.completed_via_app else "No", item.paid_at.replace(tzinfo=None) if item.paid_at else None, item.paid_mode, item.time_to_paid_seconds, item.time_to_paid_seconds / 60 if item.time_to_paid_seconds is not None else None])
    compact_sheet(candidates_sheet, widths={6: 42})
    for row in range(2, candidates_sheet.max_row + 1):
        candidates_sheet.cell(row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
        candidates_sheet.cell(row, 2).number_format = "yyyy-mm-dd hh:mm:ss"
        candidates_sheet.cell(row, 8).number_format = "#,##0.00"
        candidates_sheet.cell(row, 11).number_format = "yyyy-mm-dd hh:mm:ss"
        candidates_sheet.cell(row, 13).number_format = "0.00"
        candidates_sheet.cell(row, 14).number_format = "0.00"
        status = candidates_sheet.cell(row, 10)
        status.fill = PatternFill("solid", fgColor=GREEN if status.value == "Sí" else ORANGE)
    configure_print(candidates_sheet)

    types_sheet = workbook.create_sheet("Eventos técnicos")
    types_sheet.append(["Tipo de evento", "Cantidad", "Severidad"])
    counts = Counter((str(item.get("event_type") or ""), str(item.get("severity") or "")) for item in day_events)
    for (event_type, severity), count in sorted(counts.items(), key=lambda pair: (-pair[1], pair[0])):
        types_sheet.append([event_type, count, severity])
    compact_sheet(types_sheet, widths={1: 50})
    configure_print(types_sheet)

    methodology = workbook.create_sheet("Metodología")
    methodology.append(["Concepto", "Definición aplicada"])
    definitions = [
        ("Período", f"Día {run_date.isoformat()} en America/Argentina/Buenos_Aires; backlog reconstruido desde {coverage_from.isoformat()}."),
        ("Solicitud observada", "OID que la app informó mediante transfer_candidate_observed al verlo en la lista A Transferir."),
        ("No realizada vía app", "Solicitud observada hasta el cierre del día sin una transferencia confirmada por la app."),
        ("Realizada vía app", "Intento iniciado por la app con confirmación bancaria; puede quedar pendiente el registro final del comprobante."),
        ("Manual / automática", "Valor mode del evento transfer_started."),
        ("Cancelación", "Operación que contiene eventos cancellation_leg_* dentro del mismo intento."),
        ("Tiempo hasta Pagada", "Segundos entre la primera observación transfer_candidate_observed de la solicitud en A Transferir y mark_paid_request_succeeded. No se calcula si falta alguno de los dos eventos."),
        ("Tiempo técnico", "Inicio y fin del intento dentro de la app se conservan sólo como detalle técnico; no alimentan el tiempo operativo del resumen."),
        ("Cobertura", "Las solicitudes no realizadas vía app sólo son medibles desde clientes que emiten transfer_candidate_observed. Si no hay esa señal, el resumen muestra Sin cobertura."),
        ("Fuente única", "MetaMap Platform Server /api/v1/transfer-trace-events."),
        ("Privacidad", "Se omiten CBU, CUIL, documentos, payloads y respuestas HTTP crudas."),
    ]
    for row in definitions:
        methodology.append(row)
    compact_sheet(methodology, widths={1: 30, 2: 110})
    methodology.auto_filter.ref = None
    for row in range(2, methodology.max_row + 1):
        methodology.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        methodology.row_dimensions[row].height = 34
    configure_print(methodology)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    return workbook


def publish(workbook: Workbook, metadata: dict[str, Any], root: Path, run_date: date) -> tuple[Path, Path, Path]:
    report_dir = root / "contabilidad" / "transferencias-app"
    history_dir, metadata_dir = report_dir / "historico", report_dir / "metadata"
    history_dir.mkdir(parents=True, exist_ok=True)
    metadata_dir.mkdir(parents=True, exist_ok=True)
    dated, latest = history_dir / f"{run_date.isoformat()}.xlsx", report_dir / "ultimo.xlsx"
    metadata_path = metadata_dir / f"{run_date.isoformat()}.json"
    with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, dated)
        dated.chmod(0o644)
        with tempfile.NamedTemporaryFile(dir=report_dir, suffix=".xlsx", delete=False) as handle:
            latest_temporary = Path(handle.name)
        shutil.copy2(dated, latest_temporary)
        os.replace(latest_temporary, latest)
        latest.chmod(0o644)
    finally:
        temporary.unlink(missing_ok=True)
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    metadata_path.chmod(0o644)
    latest_metadata = report_dir / "ultimo.json"
    shutil.copy2(metadata_path, latest_metadata)
    latest_metadata.chmod(0o644)
    return latest, dated, metadata_path


def set_kestra_outputs(values: dict[str, Any]) -> None:
    if Kestra is not None:
        Kestra.outputs(values)
    else:
        print(json.dumps(values, ensure_ascii=False, default=str))


def main() -> int:
    run_date = parse_run_date()
    coverage_from = date.fromisoformat(env("TRANSFER_TRACE_COVERAGE_FROM", DEFAULT_COVERAGE_FROM))
    start = datetime.combine(coverage_from, datetime_time.min, ARGENTINA_TIMEZONE)
    end = datetime.combine(run_date + timedelta(days=1), datetime_time.min, ARGENTINA_TIMEZONE) - timedelta(microseconds=1)
    session = requests.Session()
    session.headers.update({"X-Client-Id": require_env("TRANSFER_TRACE_CLIENT_ID"), "X-Client-Secret": require_env("TRANSFER_TRACE_CLIENT_SECRET")})
    base_url = require_env("TRANSFER_TRACE_BASE_URL")
    historical_events: list[dict[str, Any]] = []
    for event_type in sorted(TRANSFER_EVENT_TYPES | {"transfer_candidate_observed"}):
        historical_events.extend(fetch_trace_events(session, base_url, start, end, event_type=event_type))
    day_start = datetime.combine(run_date, datetime_time.min, ARGENTINA_TIMEZONE)
    day_events = fetch_trace_events(session, base_url, day_start, end)
    events_by_id = {str(item.get("event_id") or ""): item for item in historical_events}
    events_by_id.update({str(item.get("event_id") or ""): item for item in day_events})
    events = list(events_by_id.values())
    operations = build_operations(events)
    candidates = build_candidates(events, operations)
    day_operations = [item for item in operations if item.started_at.date() == run_date]
    day_paid_candidates = [item for item in candidates if item.paid_at is not None and item.paid_at.date() == run_date]
    pending = [item for item in candidates if item.first_observed_at.date() <= run_date and not item.completed_via_app]
    new_candidates = [item for item in candidates if item.first_observed_at.date() == run_date]
    candidate_coverage = has_candidate_coverage(events)
    metadata = {
        "ok": True, "run_date": run_date.isoformat(), "coverage_from": coverage_from.isoformat(),
        "generated_at": datetime.now(ARGENTINA_TIMEZONE).isoformat(), "candidate_coverage_available": candidate_coverage,
        "trace_event_count": len(day_events), "new_candidate_count": len(new_candidates) if candidate_coverage else None,
        "not_transferred_via_app_count": len(pending) if candidate_coverage else None,
        "operation_attempt_count": len(day_operations), "completed_request_count": distinct_completed(day_operations),
        "completed_manual_count": distinct_completed(day_operations, mode="Manual"),
        "completed_automatic_count": distinct_completed(day_operations, mode="Automática"),
        "manual_cancellation_count": distinct_completed(day_operations, mode="Manual", operation_type="Cancelación"),
        "automatic_cancellation_count": distinct_completed(day_operations, mode="Automática", operation_type="Cancelación"),
        "paid_with_measurable_time_count": len(day_paid_candidates),
        "average_manual_detection_to_paid_seconds": average_time_to_paid(day_paid_candidates, mode="Manual"),
        "average_automatic_detection_to_paid_seconds": average_time_to_paid(day_paid_candidates, mode="Automática"),
        "average_manual_cancellation_detection_to_paid_seconds": average_time_to_paid(day_paid_candidates, mode="Manual", operation_type="Cancelación"),
        "average_automatic_cancellation_detection_to_paid_seconds": average_time_to_paid(day_paid_candidates, mode="Automática", operation_type="Cancelación"),
    }
    workbook = build_workbook(run_date, events, operations, candidates, coverage_from=coverage_from)
    latest, dated, metadata_path = publish(workbook, metadata, Path(env("TRANSFER_TRACE_REPORT_ROOT", DEFAULT_REPORT_ROOT)), run_date)
    set_kestra_outputs({**metadata, "latest_path": str(latest), "history_path": str(dated), "metadata_path": str(metadata_path)})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
