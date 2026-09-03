from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


MODULE_PATH = Path(__file__).parents[1] / "files" / "transfer_trace_report" / "generate.py"
SPEC = importlib.util.spec_from_file_location("transfer_trace_report", MODULE_PATH)
REPORT = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = REPORT
SPEC.loader.exec_module(REPORT)


def event(event_type, occurred_at, request_oid="100", mode="manual", data=None, session="s1"):
    return {"event_id": f"{event_type}-{occurred_at}-{request_oid}", "session_id": session, "client_instance_id": "desktop-1", "event_type": event_type, "occurred_at": occurred_at, "operator": "Operador", "application_version": "2.0.1", "request_oid": request_oid, "mode": mode, "severity": "info", "data": data or {}}


class TransferTraceReportTest(unittest.TestCase):
    def test_builds_manual_and_automatic_cancellation_metrics(self):
        events = [
            event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", mode=None, data={"request_status": "A Transferir"}),
            event("transfer_started", "2026-09-03T10:00:00-03:00", data={"credit_line": "Línea", "transfer_amount": "1000"}),
            event("confirmation_finished", "2026-09-03T10:00:30-03:00", data={"id_coelsa": "TX-1", "status": {"classification": "confirmed"}}),
            event("mark_paid_request_succeeded", "2026-09-03T10:00:40-03:00"),
            event("transfer_candidate_observed", "2026-09-03T10:30:00-03:00", request_oid="200", mode=None, data={"request_status": "A Transferir", "is_cancellation": True}),
            event("transfer_started", "2026-09-03T11:00:00-03:00", request_oid="200", mode="automatica"),
            event("cancellation_leg_payload_built", "2026-09-03T11:00:10-03:00", request_oid="200", mode="automatica", data={"leg": "member"}),
            event("cancellation_leg_payload_built", "2026-09-03T11:00:20-03:00", request_oid="200", mode="automatica", data={"leg": "creditor-1"}),
            event("mark_paid_request_succeeded", "2026-09-03T11:01:00-03:00", request_oid="200", mode="automatica"),
        ]
        operations = REPORT.build_operations(events)
        candidates = REPORT.build_candidates(events, operations)
        self.assertEqual(len(operations), 2)
        self.assertEqual(operations[0].flow_duration_seconds, 40)
        self.assertEqual(operations[1].operation_type, "Cancelación")
        self.assertEqual(operations[1].cancellation_legs, 2)
        self.assertEqual(REPORT.distinct_completed(operations, mode="Manual"), 1)
        self.assertEqual(REPORT.distinct_completed(operations, mode="Automática", operation_type="Cancelación"), 1)
        self.assertEqual(candidates[0].time_to_paid_seconds, 3640)
        self.assertEqual(REPORT.average_time_to_paid(candidates, mode="Automática"), 1860)
        self.assertEqual(REPORT.average_time_to_paid(candidates, mode="Automática", operation_type="Cancelación"), 1860)

    def test_candidates_are_the_universe_for_not_transferred_via_app(self):
        events = [
            event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", request_oid="100", mode=None, data={"request_status": "A Transferir"}),
            event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", request_oid="200", mode=None, data={"request_status": "A Transferir", "is_cancellation": True}),
            event("transfer_started", "2026-09-03T10:00:00-03:00", request_oid="100"),
            event("mark_paid_request_succeeded", "2026-09-03T10:00:40-03:00", request_oid="100"),
        ]
        candidates = REPORT.build_candidates(events, REPORT.build_operations(events))
        pending = [item for item in candidates if not item.completed_via_app]
        self.assertEqual(len(candidates), 2)
        self.assertEqual([item.request_oid for item in pending], ["200"])
        self.assertEqual(pending[0].operation_type, "Cancelación")

    def test_confirmed_transfer_with_pending_core_registration_is_still_via_app(self):
        events = [
            event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", mode=None),
            event("transfer_started", "2026-09-03T10:00:00-03:00"),
            event("confirmation_finished", "2026-09-03T10:00:30-03:00", data={"status": {"classification": "confirmed"}}),
            event("mark_paid_request_failed", "2026-09-03T10:00:40-03:00"),
        ]
        operations = REPORT.build_operations(events)
        candidates = REPORT.build_candidates(events, operations)
        self.assertEqual(operations[0].result, "Transferida; registro pendiente")
        self.assertTrue(candidates[0].completed_via_app)
        self.assertIsNone(candidates[0].paid_at)
        self.assertIsNone(candidates[0].time_to_paid_seconds)

    def test_candidate_first_observation_is_deduplicated_across_sessions(self):
        events = [event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", mode=None, session="s1"), event("transfer_candidate_observed", "2026-09-04T09:00:00-03:00", mode=None, session="s2")]
        candidates = REPORT.build_candidates(events, [])
        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0].first_observed_at.date(), date(2026, 9, 3))
        self.assertEqual(candidates[0].last_observed_at.date(), date(2026, 9, 4))
        self.assertEqual(candidates[0].observations, 2)

    def test_version_201_establishes_zero_candidate_coverage(self):
        self.assertTrue(REPORT.has_candidate_coverage([event("app_started", "2026-09-03T09:00:00-03:00", request_oid=None, mode=None)]))
        legacy = event("app_started", "2026-09-03T09:00:00-03:00", request_oid=None, mode=None)
        legacy["application_version"] = "2.0.0"
        self.assertFalse(REPORT.has_candidate_coverage([legacy]))

    def test_workbook_does_not_report_technical_duration_as_operational_time(self):
        events = [
            event("transfer_started", "2026-09-03T10:00:00-03:00"),
            event("mark_paid_request_succeeded", "2026-09-03T10:00:40-03:00"),
        ]
        operations = REPORT.build_operations(events)
        workbook = REPORT.build_workbook(
            date(2026, 9, 3), events, operations, [], coverage_from=date(2026, 9, 3)
        )

        self.assertEqual(workbook["Resumen"]["B12"].value, 0)
        self.assertEqual(workbook["Resumen"]["C17"].value, "Sin datos")
        self.assertIsNone(workbook["Operaciones app"]["G2"].value)
        self.assertEqual(workbook["Operaciones app"]["S2"].value, 40)
        self.assertEqual(len(workbook["Resumen"]._charts), 1)

    def test_workbook_and_publish_include_auditable_sheets(self):
        events = [event("transfer_candidate_observed", "2026-09-03T09:00:00-03:00", mode=None), event("transfer_started", "2026-09-03T10:00:00-03:00"), event("mark_paid_request_succeeded", "2026-09-03T10:00:40-03:00")]
        operations = REPORT.build_operations(events)
        candidates = REPORT.build_candidates(events, operations)
        workbook = REPORT.build_workbook(date(2026, 9, 3), events, operations, candidates, coverage_from=date(2026, 9, 3))
        self.assertEqual(workbook.sheetnames, ["Resumen", "Operaciones app", "Solicitudes observadas", "Eventos técnicos", "Metodología"])
        self.assertEqual(workbook["Resumen"]["B5"].value, 1)
        self.assertEqual(workbook["Resumen"]["B6"].value, 0)
        self.assertEqual(workbook["Resumen"]["B12"].value, 1)
        self.assertAlmostEqual(workbook["Resumen"]["C17"].value, 3640 / 60)
        self.assertEqual(workbook["Operaciones app"]["A2"].value.isoformat(), "2026-09-03T09:00:00")
        self.assertEqual(workbook["Operaciones app"]["B2"].value.isoformat(), "2026-09-03T10:00:40")
        self.assertEqual(len(workbook["Resumen"]._charts), 2)
        with tempfile.TemporaryDirectory() as directory:
            latest, dated, metadata = REPORT.publish(workbook, {"ok": True}, Path(directory), date(2026, 9, 3))
            reopened = load_workbook(latest, read_only=True, data_only=False)
            try:
                self.assertEqual(reopened.sheetnames, workbook.sheetnames)
            finally:
                reopened.close()
            self.assertTrue(dated.exists())
            self.assertTrue(metadata.exists())


if __name__ == "__main__":
    unittest.main()
