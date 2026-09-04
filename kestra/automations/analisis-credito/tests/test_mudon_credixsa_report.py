from __future__ import annotations

from contextlib import closing
from datetime import datetime, timezone
import json
from pathlib import Path
import sqlite3
import sys
import tempfile
import unittest
from unittest.mock import Mock, patch

from openpyxl import load_workbook


FILES_ROOT = Path(__file__).resolve().parents[1] / "files"
if str(FILES_ROOT) not in sys.path:
    sys.path.insert(0, str(FILES_ROOT))

from mudon_credixsa_report.core import (  # noqa: E402
    CoreMember,
    build_active_loan_filter,
    fetch_active_mudon_members,
)
from mudon_credixsa_report.credix import (  # noqa: E402
    is_issn_employer,
    load_shared_cache,
    process_output,
)
from mudon_credixsa_report.excel import atomic_publish, write_workbook  # noqa: E402
from mudon_credixsa_report import entrypoint  # noqa: E402
from mudon_credixsa_report.state import StateStore  # noqa: E402


class FlowConfigTests(unittest.TestCase):
    def test_defaulted_inputs_are_required_by_kestra(self) -> None:
        flow_path = Path(__file__).resolve().parents[1] / "flows/mudon_credixsa_report.yaml"
        flow = flow_path.read_text(encoding="utf-8")
        self.assertEqual(flow.count("    required: true\n    defaults: false"), 2)

    def test_schedules_pass_an_explicit_run_mode(self) -> None:
        flow_path = Path(__file__).resolve().parents[1] / "flows/mudon_credixsa_report.yaml"
        flow = flow_path.read_text(encoding="utf-8")
        self.assertIn('RUN_MODE: "{{ inputs.run_mode }}"', flow)
        self.assertIn("      run_mode: monthly", flow)
        self.assertIn("      run_mode: resume", flow)
        self.assertNotIn("trigger.id", flow)


class CoreTests(unittest.TestCase):
    def test_filter_contains_both_lines_and_active_balance(self) -> None:
        value = build_active_loan_filter(
            ("MUDON HABERES", "MUDON HABERES SOCIOS NUEVOS")
        )
        self.assertIn("[SaldoPrestamo] > 0.0m", value)
        self.assertIn("MUDON HABERES'", value)
        self.assertIn("MUDON HABERES SOCIOS NUEVOS", value)

    @patch("mudon_credixsa_report.core.requests.Session")
    def test_core_rows_are_deduplicated_and_accounts_are_aggregated(self, session_cls: Mock) -> None:
        response = Mock()
        response.json.return_value = [
            ["20-12345678-3", "12345678", "Persona Uno", "10", "100", "20", "MUDON HABERES"],
            ["20123456783", "12345678", "Persona Uno", "10", "101", "30", "MUDON HABERES SOCIOS NUEVOS"],
        ]
        response.raise_for_status.return_value = None
        session_cls.return_value.post.return_value = response

        members = fetch_active_mudon_members(base_url="https://core.test:5002")

        self.assertEqual(len(members), 1)
        self.assertEqual(members[0].cuit, "20123456783")
        self.assertEqual(members[0].loan_accounts, {"100", "101"})
        payload = session_cls.return_value.post.call_args.kwargs["json"]
        self.assertEqual(payload["tipo"], "F.Module.Cuentas.Prestamos.Prestamo")

    @patch("mudon_credixsa_report.core.requests.Session")
    def test_core_aborts_when_maximum_is_reached(self, session_cls: Mock) -> None:
        response = Mock()
        response.json.return_value = [["20123456783"]] * 2
        response.raise_for_status.return_value = None
        session_cls.return_value.post.return_value = response
        with self.assertRaisesRegex(RuntimeError, "alcanzo el limite"):
            fetch_active_mudon_members(base_url="https://core.test:5002", max_rows=2)


class CredixTests(unittest.TestCase):
    def test_issn_name_accepts_real_variant(self) -> None:
        self.assertTrue(is_issn_employer("INSTITUTO DE SEGURIDAD SOCIAL DEL NEUQUEN O. P."))
        self.assertTrue(is_issn_employer("Instituto de Seguridad Social del Neuquén"))
        self.assertFalse(is_issn_employer("PROVINCIA DEL NEUQUEN ADMINISTRACION CENTRAL"))

    def test_process_output_extracts_employer_and_periods(self) -> None:
        normalized = {
            "previsional": {
                "empleadores": [],
                "situaciones_por_empleador": [
                    {
                        "empleador": {
                            "nombre": "INSTITUTO DE SEGURIDAD SOCIAL DEL NEUQUEN O. P.",
                            "cuit": "30999999991",
                        },
                        "periodos": [{"periodo": "07/2026", "aportes_seguridad_social": "INFORMATIVO"}],
                    }
                ],
            }
        }
        result = process_output(
            {
                "ok": True,
                "status": "single",
                "normalized_json": json.dumps(normalized),
            },
            source="online",
            checked_at="2026-08-24T12:00:00+00:00",
        )
        self.assertTrue(result.qualifies_issn)
        self.assertEqual(result.employers[0]["periodos"][0]["periodo"], "07/2026")

    def test_shared_cache_preserves_original_timestamp(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "credix.sqlite"
            checked_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
            payload = {
                "version": 4,
                "cached_at": checked_at,
                "result": {
                    "ok": True,
                    "status": "single",
                    "cuit": "20123456783",
                    "nombre": "Persona Uno",
                    "data": [{"title": "Datos Filiatorios"}],
                    "normalized": {"previsional": {"empleadores": []}},
                },
            }
            with closing(sqlite3.connect(path)) as connection:
                connection.execute(
                    "CREATE TABLE credixsa_cache (lookup_key TEXT PRIMARY KEY, payload_json TEXT NOT NULL)"
                )
                connection.execute(
                    "INSERT INTO credixsa_cache VALUES (?, ?)",
                    ("credixsa.cuil.20123456783", json.dumps(payload)),
                )
                connection.commit()
            output = load_shared_cache(str(path), "20123456783", 7)
            self.assertIsNotNone(output)
            self.assertTrue(output["cache_hit"])
            self.assertEqual(output["cached_at"], checked_at)
            self.assertFalse(output["cache_should_persist"])


class StateTests(unittest.TestCase):
    def test_state_claims_only_valid_cuil_and_persists_recent_result(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            store = StateStore(str(Path(directory) / "state.sqlite"))
            members = [
                CoreMember("20123456783", "20123456783", "12345678", "Persona Uno", "10"),
                CoreMember("dni:87654321", "", "87654321", "Persona Dos", "11"),
            ]
            store.create_run("run-1", "manual", members, force_refresh=False)
            claimed = store.claim_members("run-1", 5)
            self.assertEqual([row["member_key"] for row in claimed], ["20123456783"])
            self.assertEqual(store.stats("run-1")["errors"], 1)

            checked_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
            output = {"ok": True, "status": "none", "normalized_json": "{}"}
            store.save_result("20123456783", "none", checked_at, output)
            self.assertEqual(store.load_recent_result("20123456783", 7), output)

    def test_failed_member_is_requeued_until_max_attempts(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            store = StateStore(str(Path(directory) / "state.sqlite"))
            member = CoreMember("20123456783", "20123456783", "12345678", "Persona Uno", "10")
            store.create_run("run-1", "manual", [member], force_refresh=False)
            store.claim_members("run-1", 1)
            store.fail_member("run-1", member.member_key, "timeout", max_attempts=2)
            self.assertEqual(store.stats("run-1")["pending"], 1)
            store.claim_members("run-1", 1)
            store.fail_member("run-1", member.member_key, "timeout", max_attempts=2)
            self.assertEqual(store.stats("run-1")["errors"], 1)


class ExcelTests(unittest.TestCase):
    def test_workbook_and_atomic_publication(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "source.xlsx"
            run = {"run_id": "mudon-2026-08"}
            employers = [
                {
                    "nombre": "INSTITUTO DE SEGURIDAD SOCIAL DEL NEUQUEN O. P.",
                    "cuit": "30999999991",
                    "periodos": [{"periodo": "07/2026"}],
                }
            ]
            members = [
                {
                    "member_number": "10",
                    "full_name": "Persona Uno",
                    "dni": "12345678",
                    "cuit": "20123456783",
                    "loan_accounts_json": '["100"]',
                    "loan_lines_json": '["MUDON HABERES"]',
                    "employers_json": json.dumps(employers),
                    "qualifies_issn": 1,
                    "credix_status": "single",
                    "result_source": "online",
                    "checked_at": "2026-08-24T12:00:00+00:00",
                    "error": "",
                }
            ]
            stats = {
                "total": 1,
                "completed": 1,
                "cache": 0,
                "online": 1,
                "qualifying": 1,
                "errors": 0,
            }
            write_workbook(source, run, members, stats)
            workbook = load_workbook(source)
            self.assertEqual(workbook.sheetnames, ["Resumen", "Socios"])
            self.assertEqual(workbook["Socios"]["A2"].value, "10")
            self.assertEqual(workbook["Socios"]["J2"].value, "Si")

            latest, history = atomic_publish(
                source,
                root / "reports",
                datetime(2026, 8, 24, 12, 0, 0),
            )
            self.assertTrue(latest.exists())
            self.assertEqual(history.name, "2026-08-24.xlsx")


class EntrypointTests(unittest.TestCase):
    def test_resume_mode_never_creates_a_run_when_none_is_active(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            store = StateStore(Path(directory) / "state.sqlite")
            with patch.object(
                entrypoint,
                "fetch_active_mudon_members",
                side_effect=AssertionError("resume must not query Core"),
            ):
                run = entrypoint.ensure_run(
                    store,
                    "resume",
                    {},
                    datetime(2026, 8, 24, 12, 0, 0, tzinfo=timezone.utc),
                )

        self.assertIsNone(run)

    @patch("mudon_credixsa_report.entrypoint.load_shared_cache")
    @patch("mudon_credixsa_report.entrypoint.fetch_active_mudon_members")
    def test_manual_run_completes_from_fresh_cache_and_publishes(
        self,
        fetch_members: Mock,
        load_cache: Mock,
    ) -> None:
        fetch_members.return_value = [
            CoreMember("20123456783", "20123456783", "12345678", "Persona Uno", "10")
        ]
        load_cache.return_value = {
            "ok": True,
            "status": "single",
            "cached_at": "2026-08-24T12:00:00+00:00",
            "normalized_json": json.dumps(
                {
                    "previsional": {
                        "empleadores": [
                            {"nombre": "INSTITUTO DE SEGURIDAD SOCIAL DEL NEUQUEN O. P."}
                        ],
                        "situaciones_por_empleador": [],
                    }
                }
            ),
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            variables = {
                "RUN_MODE": "manual",
                "TRIGGER_BODY_JSON": "{}",
                "MUDON_STATE_DB_PATH": str(root / "state.sqlite"),
                "CREDIX_CACHE_SQLITE_PATH": str(root / "credix.sqlite"),
                "REPORTS_ROOT": str(root / "reports"),
                "MUDON_CORE_BASE_URL": "https://core.test:5002",
                "MUDON_CREDIXSA_BATCH_SIZE": "5",
                "MUDON_CREDIXSA_DELAY_SECONDS": "0",
            }
            with patch.dict("os.environ", variables, clear=False):
                result = entrypoint.run_batch(
                    datetime(2026, 8, 24, 12, 0, 0, tzinfo=timezone.utc)
                )

            self.assertEqual(result["status"], "completed")
            self.assertEqual(result["qualifying"], 1)
            self.assertTrue((root / "reports/cobranzas/mudon-jubilados/ultimo.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
