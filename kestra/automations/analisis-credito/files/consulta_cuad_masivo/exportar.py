"""Exporta los resultados de una corrida de CUAD a Excel."""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

PATRONES_NDJSON = ("resultados*.ndjson", "resultados_cuad_main2*.ndjson")

ENCABEZADOS_TOTALES = [
    "Empleador", "CUIL consultado", "Apellido y nombre", "DNI", "Organizacion", "Jurisdiccion", "Entidad", "Nro afiliado", "Bruto", "Neto", "Cupo", "Afectado", "% Afectado", "PreCancelado", "% PreCancelado", "Disponible", "% Disponible", "Deuda",
]
ENCABEZADOS_MOVIMIENTOS = [
    "Empleador", "CUIL consultado", "Apellido y nombre", "DNI", "Organizacion", "Jurisdiccion", "Entidad empleado", "Nro afiliado", "Organismo", "Sector", "Entidad movimiento", "Cupo", "Afectado", "% Afectado", "PreCancelado", "% PreCancelado", "Deuda",
]


def texto(valor):
    return "" if valor is None else str(valor).strip()


def leer_ndjson(path):
    with path.open("r", encoding="utf-8-sig") as archivo:
        for linea in archivo:
            linea = linea.strip()
            if linea:
                yield json.loads(linea)


def cargar_ultimos_registros(path):
    ultimos = {}
    for registro in leer_ndjson(path):
        cuil = texto(registro.get("cuil"))
        if cuil:
            ultimos.pop(cuil, None)
            ultimos[cuil] = registro
    return list(ultimos.values())


def buscar_ndjson_mas_reciente(directorio=Path("corridas")):
    candidatos = set()
    for patron in PATRONES_NDJSON:
        candidatos.update(path for path in Path(directorio).glob(f"*/{patron}") if path.is_file())
    return max(candidatos, key=lambda path: path.stat().st_mtime) if candidatos else None


def datos_comunes(registro):
    parsed = registro.get("parsed") or {}
    empleado = parsed.get("tabla_empleado") or {}
    empleador = texto(registro.get("emr_nombre")) or texto((registro.get("payload") or {}).get("Emr_Nombre"))
    return empleador, empleado, parsed.get("tabla_totales") or {}, parsed.get("tabla_movimientos") or {}


def fila_totales(registro):
    empleador, empleado, totales, _ = datos_comunes(registro)
    return [
        empleador, texto(registro.get("cuil")), texto(empleado.get("apellido_nombre")), texto(empleado.get("documento")), texto(empleado.get("organizacion")), texto(empleado.get("jurisdiccion_codigo")), texto(empleado.get("entidad")), texto(empleado.get("nro_afiliado")), texto(totales.get("bruto")), texto(totales.get("neto")), texto(totales.get("cupo")), texto(totales.get("afectado")), texto(totales.get("afectado_porcentaje")), texto(totales.get("precancelado")), texto(totales.get("precancelado_porcentaje")), texto(totales.get("disponible")), texto(totales.get("disponible_porcentaje")), texto(totales.get("deuda")),
    ]


def filas_movimientos(registro):
    empleador, empleado, _, movimientos = datos_comunes(registro)
    filas = []
    for movimiento in movimientos.get("registros") or []:
        filas.append([
            empleador, texto(registro.get("cuil")), texto(empleado.get("apellido_nombre")), texto(empleado.get("documento")), texto(empleado.get("organizacion")), texto(empleado.get("jurisdiccion_codigo")), texto(empleado.get("entidad")), texto(empleado.get("nro_afiliado")), texto(movimiento.get("Organismo")), texto(movimiento.get("Sector")), texto(movimiento.get("Entidad")), texto(movimiento.get("Cupo")), texto(movimiento.get("Afectado")), texto(movimiento.get("%_1")), texto(movimiento.get("PreCancelado")), texto(movimiento.get("%_2")), texto(movimiento.get("Deuda")),
        ])
    return filas


def ajustar_anchos(hoja):
    for columna in hoja.columns:
        ancho = max(len(str(celda.value or "")) for celda in columna)
        hoja.column_dimensions[get_column_letter(columna[0].column)].width = max(ancho + 2, 12)


def exportar(archivo_ndjson, tipo):
    registros = cargar_ultimos_registros(archivo_ndjson)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Totales" if tipo == "totales" else "Movimientos"
    hoja.append(ENCABEZADOS_TOTALES if tipo == "totales" else ENCABEZADOS_MOVIMIENTOS)
    filas = 0
    for registro in registros:
        if texto(registro.get("status")).lower() != "ok":
            continue
        if tipo == "totales":
            hoja.append(fila_totales(registro))
            filas += 1
        else:
            for fila in filas_movimientos(registro):
                hoja.append(fila)
                filas += 1
    hoja.freeze_panes = "A2"
    ajustar_anchos(hoja)
    salida = archivo_ndjson.with_name(f"{archivo_ndjson.stem}_{tipo}.xlsx")
    libro.save(salida)
    return salida, len(registros), filas


def main(argv=None):
    parser = argparse.ArgumentParser(description="Exporta resultados de CUAD a Excel.")
    parser.add_argument("archivo_ndjson", nargs="?", type=Path, help="Archivo .ndjson; por defecto, el mas reciente en corridas/.")
    parser.add_argument("--tipo", required=True, choices=("totales", "movimientos"))
    args = parser.parse_args(argv)
    archivo = args.archivo_ndjson or buscar_ndjson_mas_reciente()
    if archivo is None:
        parser.error("no se encontro ningun NDJSON en corridas/.")
    if not archivo.is_file():
        parser.error(f"no existe el archivo: {archivo}")
    salida, leidos, exportados = exportar(archivo, args.tipo)
    print(f"Archivo generado: {salida}")
    print(f"Registros leidos: {leidos}")
    print(f"Filas exportadas: {exportados}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
