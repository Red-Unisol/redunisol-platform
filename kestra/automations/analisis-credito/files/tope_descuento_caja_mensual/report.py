from __future__ import annotations

import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .checkpoint import ResultRow
from .contact_data import ContactKind, PreferredContact, choose_preferred, display_values
from .sources import Candidate, SourceStats

HEADER_FILL = "17365D"
HEADER_FONT = "FFFFFF"
ERROR_FILL = "FFC7CE"
NOT_FOUND_FILL = "FFEB9C"
PRIORITY_FILL = "1F6E78"
AUXILIARY_CONTACT_HEADERS = {
    "email_core",
    "telefono_core",
    "emails_bitrix_lead",
    "telefonos_bitrix_lead",
    "emails_bitrix_contacto",
    "telefonos_bitrix_contacto",
    "criterio_mail",
    "criterio_telefono",
}


def build_workbook(
    candidates: list[Candidate],
    results: dict[str, ResultRow],
    stats: SourceStats,
    generated_at: datetime,
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"
    headers = [
        "cuil",
        "nombre",
        "apellido",
        "mail",
        "telefono",
        "disponible",
        "tope_descuento",
        "estado",
        "consultado",
        "error",
        "origen_core",
        "origen_bitrix_jubilado",
        "origen_bitrix_pensionado_bancor",
        "cuil_desde_lead",
        "cuil_desde_contacto",
        "prestamos_core",
        "lineas_core",
        "leads_bitrix",
        "email_core",
        "telefono_core",
        "emails_bitrix_lead",
        "telefonos_bitrix_lead",
        "emails_bitrix_contacto",
        "telefonos_bitrix_contacto",
        "criterio_mail",
        "criterio_telefono",
    ]
    sheet.append(headers)
    for candidate in candidates:
        result = results.get(candidate.cuil)
        preferred_mail = choose_preferred(
            cuil=candidate.cuil,
            kind="email",
            core=candidate.core_emails,
            leads=candidate.bitrix_lead_emails,
            contacts=candidate.bitrix_contact_emails,
        )
        preferred_phone = choose_preferred(
            cuil=candidate.cuil,
            kind="phone",
            core=candidate.core_phones,
            leads=candidate.bitrix_lead_phones,
            contacts=candidate.bitrix_contact_phones,
        )
        sheet.append(
            [
                candidate.cuil,
                result.nombre if result else "",
                result.apellido if result else "",
                preferred_mail.value,
                preferred_phone.value,
                result.disponible if result else None,
                result.tope_descuento if result else None,
                result.status if result else "pending",
                result.checked_at if result else "",
                result.error if result else "",
                candidate.from_core,
                candidate.from_bitrix_jubilado,
                candidate.from_bitrix_pensionado_bancor,
                candidate.cuil_from_lead,
                candidate.cuil_from_contact,
                ", ".join(sorted(candidate.core_loan_ids)),
                ", ".join(sorted(candidate.core_line_names)),
                ", ".join(sorted(candidate.bitrix_lead_ids, key=_numeric_key)),
                display_values(candidate.core_emails),
                display_values(candidate.core_phones),
                display_values(candidate.bitrix_lead_emails),
                display_values(candidate.bitrix_lead_phones),
                display_values(candidate.bitrix_contact_emails),
                display_values(candidate.bitrix_contact_phones),
                preferred_mail.criterion,
                preferred_phone.criterion,
            ]
        )
    _format_results(sheet)
    _add_summary(workbook, candidates, results, stats, generated_at)
    return workbook


def save_workbook(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def atomic_publish(
    source: Path, reports_root: Path, run_month: str
) -> tuple[Path, Path]:
    report_dir = reports_root / "analisis-credito" / "tope-descuento-caja"
    history_dir = report_dir / "historico"
    history_dir.mkdir(parents=True, exist_ok=True)
    latest = report_dir / "ultimo.xlsx"
    history = history_dir / f"{run_month}.xlsx"
    for destination in (latest, history):
        with tempfile.NamedTemporaryFile(
            dir=destination.parent, suffix=".xlsx", delete=False
        ) as handle:
            temporary = Path(handle.name)
        try:
            shutil.copyfile(source, temporary)
            os.replace(temporary, destination)
            destination.chmod(0o644)
        finally:
            temporary.unlink(missing_ok=True)
    return latest, history


def _format_results(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in (4, 5):
        sheet.cell(1, column).fill = PatternFill("solid", fgColor=PRIORITY_FILL)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row, 1).number_format = "@"
        sheet.cell(row, 4).number_format = "@"
        sheet.cell(row, 5).number_format = "@"
        sheet.cell(row, 6).number_format = "#,##0.00"
        sheet.cell(row, 7).number_format = '0"%"'
        status = sheet.cell(row, 8).value
        if status == "not_found":
            sheet.cell(row, 8).fill = PatternFill("solid", fgColor=NOT_FOUND_FILL)
        elif status != "completed":
            sheet.cell(row, 8).fill = PatternFill("solid", fgColor=ERROR_FILL)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    for column in range(1, sheet.max_column + 1):
        width = max(
            len(str(sheet.cell(row, column).value or ""))
            for row in range(1, min(sheet.max_row, 200) + 1)
        )
        sheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 42)
        header = sheet.cell(1, column).value
        if header in AUXILIARY_CONTACT_HEADERS:
            dimension = sheet.column_dimensions[get_column_letter(column)]
            dimension.hidden = True
            dimension.outlineLevel = 1


def _add_summary(
    workbook: Workbook,
    candidates: list[Candidate],
    results: dict[str, ResultRow],
    stats: SourceStats,
    generated_at: datetime,
) -> None:
    sheet = workbook.create_sheet("Resumen", 0)
    completed = sum(row.status == "completed" for row in results.values())
    not_found = sum(row.status == "not_found" for row in results.values())
    invalid_cuils = sum(row.status == "invalid_cuil" for row in results.values())
    preferred_mails = sum(
        bool(_preferred(candidate, "email").value) for candidate in candidates
    )
    preferred_phones = sum(
        bool(_preferred(candidate, "phone").value) for candidate in candidates
    )
    rows = [
        ("Generado", generated_at.isoformat(timespec="seconds")),
        ("CUILs unicos", len(candidates)),
        ("Completados", completed),
        ("No encontrados", not_found),
        ("CUILs invalidos", invalid_cuils),
        ("CUILs con mail probable", preferred_mails),
        ("CUILs con telefono probable", preferred_phones),
        ("Prestamos Core leidos", stats.core_rows),
        ("Prestamos Core sin CUIL", stats.core_without_cuil),
        ("Leads jubilados provinciales Cordoba", stats.bitrix_jubilado_rows),
        ("Leads pensionados Cordoba Bancor", stats.bitrix_pensionado_rows),
        ("Leads sin CUIL directo", stats.bitrix_without_direct_cuil),
        ("Contactos Bitrix revisados", stats.bitrix_contact_ids_checked),
        ("CUILs recuperados desde contacto", stats.bitrix_contacts_recovered),
    ]
    sheet.append(["Metrica", "Valor"])
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 28


def _numeric_key(value: str) -> tuple[int, str]:
    return (0, f"{int(value):020d}") if value.isdigit() else (1, value)


def _preferred(candidate: Candidate, kind: ContactKind) -> PreferredContact:
    return choose_preferred(
        cuil=candidate.cuil,
        kind=kind,
        core=candidate.core_emails if kind == "email" else candidate.core_phones,
        leads=(
            candidate.bitrix_lead_emails
            if kind == "email"
            else candidate.bitrix_lead_phones
        ),
        contacts=(
            candidate.bitrix_contact_emails
            if kind == "email"
            else candidate.bitrix_contact_phones
        ),
    )
