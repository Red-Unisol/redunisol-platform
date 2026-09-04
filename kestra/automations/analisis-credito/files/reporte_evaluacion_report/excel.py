from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional, Sequence

from .analysis import business_seconds_between, summarize_metric_rows_by_status, summarize_minutes
from .core import MonthlyReport, format_period_label, normalize_text

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    Workbook = None
    Comment = None
    ColorScaleRule = None
    Alignment = None
    Border = None
    Font = None
    PatternFill = None
    Side = None
    Table = None
    TableStyleInfo = None
    get_column_letter = None


STATE_AREA_MAP = {
    "liquidada": "Vendedores",
    "preaprobado": "Vendedores",
    "revisar": "Vendedores",
    "verificacion de documentos y firma": "Vendedores",
    "revisionriesgo": "Analistas",
    "confirmada": "Analistas",
    "verificaciondocumentacion": "Analistas",
    "verificardocumentacion": "Analistas",
    "a transferir": "Tesoreria",
}


def build_unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def summarize_report_collection(month_reports: Sequence[MonthlyReport]) -> Dict[str, Any]:
    all_first_values = [row["minutos"] for report in month_reports for row in report.first_response]
    all_transfer_values = [row["minutos"] for report in month_reports for row in report.transfer]
    all_end_to_end_rows = [row for report in month_reports for row in report.end_to_end]
    all_end_to_end_values = [row["minutos"] for row in all_end_to_end_rows]

    return {
        "overall_first_summary": summarize_minutes(all_first_values),
        "overall_transfer_summary": summarize_minutes(all_transfer_values),
        "overall_end_to_end_summary": summarize_minutes(all_end_to_end_values),
        "overall_end_to_end_by_status": summarize_metric_rows_by_status(all_end_to_end_rows),
        "total_solicitudes": sum(report.summary["solicitudes_count"] for report in month_reports),
        "total_cerradas": sum(report.summary["closed_solicitudes_count"] for report in month_reports),
        "total_novedades": sum(report.summary["month_novedades_count"] for report in month_reports),
        "total_historical": sum(report.summary["historical_novedades_count"] for report in month_reports),
        "analysis_excluded_solicitudes_count": sum(
            report.summary.get("analysis_excluded_solicitudes_count", 0)
            for report in month_reports
        ),
    }


def _sanitize_sheet_name(name: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join(ch for ch in name if ch not in invalid).strip()
    if not cleaned:
        cleaned = "Hoja"
    return cleaned[:31]


def _sanitize_table_name(name: str) -> str:
    cleaned = re.sub(r"[^0-9A-Za-z_]", "_", name)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if not cleaned:
        cleaned = "Tabla"
    if cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:255]


def _format_cell_value(value: Any) -> Any:
    if isinstance(value, float):
        return round(value, 2)
    return value


def _autofit_columns(ws: Any, max_width: int = 42) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            rendered = value.strftime("%d/%m/%Y %H:%M") if isinstance(value, datetime) else str(value)
            max_length = max(max_length, len(rendered))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), max_width)


def _style_range(
    ws: Any,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    *,
    fill: Any = None,
    font: Any = None,
    alignment: Any = None,
    border: Any = None,
) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border


def _add_excel_table(
    ws: Any,
    table_name: str,
    start_row: int,
    start_col: int,
    headers: Sequence[str],
    row_count: int,
) -> None:
    if row_count <= 0:
        return
    end_col = start_col + len(headers) - 1
    end_row = start_row + row_count
    table = Table(
        displayName=_sanitize_table_name(table_name),
        ref=f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _write_kpi_card(
    ws: Any,
    *,
    start_row: int,
    start_col: int,
    width: int,
    title: str,
    value: str,
    accent_color: str,
    border: Any,
) -> None:
    end_col = start_col + width - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=end_col)
    title_cell = ws.cell(start_row, start_col, title)
    value_cell = ws.cell(start_row + 1, start_col, value)
    title_fill = PatternFill("solid", fgColor=accent_color)
    value_fill = PatternFill("solid", fgColor="F7FAFC")
    _style_range(
        ws,
        start_row,
        start_row,
        start_col,
        end_col,
        fill=title_fill,
        font=Font(color="FFFFFF", bold=True, size=11),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=border,
    )
    _style_range(
        ws,
        start_row + 1,
        start_row + 2,
        start_col,
        end_col,
        fill=value_fill,
        font=Font(color="1F2937", bold=True, size=18),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=border,
    )
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_key_value_block(
    ws: Any,
    *,
    start_row: int,
    title: str,
    rows: Sequence[tuple[str, Any]],
    section_fill: Any,
    label_fill: Any,
    value_fill: Any,
    border: Any,
) -> int:
    end_col = 6
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    ws.cell(start_row, 1, title)
    _style_range(
        ws,
        start_row,
        start_row,
        1,
        end_col,
        fill=section_fill,
        font=Font(color="FFFFFF", bold=True, size=11),
        alignment=Alignment(horizontal="left", vertical="center"),
        border=border,
    )

    current_row = start_row + 1
    for label, value in rows:
        ws.cell(current_row, 1, label)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=end_col)
        ws.cell(current_row, 2, value)
        _style_range(
            ws,
            current_row,
            current_row,
            1,
            1,
            fill=label_fill,
            font=Font(bold=True, color="1F2937"),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=border,
        )
        _style_range(
            ws,
            current_row,
            current_row,
            2,
            end_col,
            fill=value_fill,
            font=Font(color="111827"),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=border,
        )
        current_row += 1

    return current_row


def _write_summary_sheet(
    ws: Any,
    *,
    period_label: str,
    run_started_at: datetime,
    overview_rows: Sequence[tuple[str, Any]],
    kpi_cards: Sequence[Dict[str, str]],
) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"
    card_width = 3
    total_columns = max(12, len(kpi_cards) * card_width)

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="DCE6F1")
    section_fill = PatternFill("solid", fgColor="305496")
    label_fill = PatternFill("solid", fgColor="EDF2F7")
    value_fill = PatternFill("solid", fgColor="FFFFFF")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Reporte evaluatorio comercial"
    _style_range(
        ws,
        1,
        1,
        1,
        total_columns,
        fill=title_fill,
        font=Font(color="FFFFFF", bold=True, size=16),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border,
    )

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    ws["A2"] = (
        f"Periodo: {period_label} | Generado: {run_started_at.strftime('%d/%m/%Y %H:%M')} | "
        "Horario laboral considerado: lunes a viernes de 08:00 a 17:00"
    )
    _style_range(
        ws,
        2,
        2,
        1,
        total_columns,
        fill=subtitle_fill,
        font=Font(color="1F2937", italic=True),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border,
    )

    card_colors = ["4472C4", "70AD47", "ED7D31", "A5A5A5", "5B9BD5", "C55A11"]
    for index, card in enumerate(kpi_cards):
        _write_kpi_card(
            ws,
            start_row=4,
            start_col=1 + index * card_width,
            width=card_width,
            title=card["title"],
            value=card["value"],
            accent_color=card_colors[index % len(card_colors)],
            border=thin_border,
        )

    _write_key_value_block(
        ws,
        start_row=8,
        title="Contexto y criterios",
        rows=overview_rows,
        section_fill=section_fill,
        label_fill=label_fill,
        value_fill=value_fill,
        border=thin_border,
    )

    ws.column_dimensions["A"].width = 32
    for column_index in range(2, total_columns + 1):
        ws.column_dimensions[get_column_letter(column_index)].width = 18


def _write_table_sheet(
    ws: Any,
    *,
    title: str,
    subtitle: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    table_name: str,
    header_comments: Optional[Sequence[Optional[str]]] = None,
    hidden_header_descriptions: Optional[Sequence[Optional[str]]] = None,
    datetime_cols: Optional[Sequence[int]] = None,
    float_cols: Optional[Sequence[int]] = None,
    int_cols: Optional[Sequence[int]] = None,
    max_width: int = 42,
) -> None:
    datetime_cols = set(datetime_cols or [])
    float_cols = set(float_cols or [])
    int_cols = set(int_cols or [])

    ws.sheet_view.showGridLines = False
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws.cell(1, 1, title)
    _style_range(
        ws,
        1,
        1,
        1,
        max(1, len(headers)),
        fill=PatternFill("solid", fgColor="1F4E78"),
        font=Font(color="FFFFFF", bold=True, size=15),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border,
    )

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(headers)))
    ws.cell(2, 1, subtitle)
    _style_range(
        ws,
        2,
        2,
        1,
        max(1, len(headers)),
        fill=PatternFill("solid", fgColor="EAF0F7"),
        font=Font(color="374151", italic=True),
        alignment=Alignment(horizontal="center", vertical="center"),
        border=thin_border,
    )

    if hidden_header_descriptions:
        for col_index, description in enumerate(hidden_header_descriptions, start=1):
            ws.cell(3, col_index, description)
        _style_range(
            ws,
            3,
            3,
            1,
            max(1, len(headers)),
            fill=PatternFill("solid", fgColor="F7FAFC"),
            font=Font(color="6B7280", italic=True, size=9),
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
            border=thin_border,
        )
        ws.row_dimensions[3].hidden = True

    header_row = 4
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_index, header)
        if header_comments and len(header_comments) >= col_index:
            comment_text = header_comments[col_index - 1]
            if comment_text:
                cell.comment = Comment(comment_text, "Codex")
    _style_range(
        ws,
        header_row,
        header_row,
        1,
        len(headers),
        fill=PatternFill("solid", fgColor="305496"),
        font=Font(color="FFFFFF", bold=True),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=thin_border,
    )

    if rows:
        for row_offset, row_values in enumerate(rows, start=1):
            current_row = header_row + row_offset
            for col_index, value in enumerate(row_values, start=1):
                cell = ws.cell(current_row, col_index, _format_cell_value(value))
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_index in datetime_cols and value is not None:
                    cell.number_format = "DD/MM/YYYY HH:MM"
                elif col_index in float_cols and value is not None:
                    cell.number_format = "0.00"
                elif col_index in int_cols and value is not None:
                    cell.number_format = "0"
        _add_excel_table(
            ws,
            table_name=table_name,
            start_row=header_row,
            start_col=1,
            headers=headers,
            row_count=len(rows),
        )
    else:
        ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max(1, len(headers)))
        ws.cell(5, 1, "Sin registros para este periodo.")
        _style_range(
            ws,
            5,
            5,
            1,
            max(1, len(headers)),
            fill=PatternFill("solid", fgColor="F9FAFB"),
            font=Font(color="6B7280", italic=True),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=thin_border,
        )

    ws.freeze_panes = f"A{header_row + 1}"
    _autofit_columns(ws, max_width=max_width)


def _format_card_metric(value: Optional[float]) -> str:
    if value is None:
        return "Sin datos"
    return f"{value:.2f} min"


def _with_period_context(report: MonthlyReport, rows: Sequence[Dict[str, Any]]) -> list[Dict[str, Any]]:
    return [{"periodo": report.month_value, "mes": report.month_label, **row} for row in rows]


def _classify_state_area(state_value: Optional[str]) -> str:
    return STATE_AREA_MAP.get(normalize_text(state_value), "Misc")


def _build_state_duration_rows(report: MonthlyReport) -> list[Dict[str, Any]]:
    grouped_rows: dict[int, list[Dict[str, Any]]] = defaultdict(list)
    for row in report.base_rows:
        if row.get("estado_detectado") and row.get("event_ts") is not None:
            grouped_rows[int(row["solicitud_oid"])].append(row)

    duration_rows: list[Dict[str, Any]] = []
    for solicitud_oid, rows in grouped_rows.items():
        ordered = sorted(
            rows,
            key=lambda item: (
                item.get("event_ts") is None,
                item.get("event_ts"),
                item.get("event_id"),
                item.get("seq_estado_en_solicitud"),
            ),
        )
        previous_state: Optional[str] = None
        segment_starts: list[Dict[str, Any]] = []
        for row in ordered:
            current_state = row.get("estado_detectado")
            if current_state != previous_state:
                segment_starts.append(row)
                previous_state = current_state

        for index, row in enumerate(segment_starts):
            start_ts = row.get("event_ts")
            next_row = segment_starts[index + 1] if index + 1 < len(segment_starts) else None
            end_ts = next_row.get("event_ts") if next_row else None
            calendar_minutes = 0.0
            business_minutes = 0.0
            if start_ts is not None and end_ts is not None:
                calendar_minutes = max((end_ts - start_ts).total_seconds() / 60.0, 0.0)
                business_minutes = business_seconds_between(start_ts, end_ts) / 60.0

            duration_rows.append(
                {
                    "periodo": report.month_value,
                    "mes": report.month_label,
                    "solicitud_oid": solicitud_oid,
                    "nro_socio": row.get("nro_socio"),
                    "linea": row.get("linea"),
                    "segmento_idx": index + 1,
                    "estado": row.get("estado_detectado"),
                    "area": _classify_state_area(row.get("estado_detectado")),
                    "inicio_estado": start_ts,
                    "fin_estado": end_ts,
                    "duracion_calendario_min": calendar_minutes,
                    "duracion_laboral_min": business_minutes,
                }
            )
    return duration_rows


def _state_time_summary_rows(
    rows: Sequence[Dict[str, Any]],
    *,
    periodo: str,
    mes: str,
    nivel: str,
) -> list[Dict[str, Any]]:
    grouped: dict[tuple[str, Optional[str]], Dict[str, Any]] = {}
    for row in rows:
        area = str(row.get("area") or "Misc")
        estado = str(row.get("estado") or "") if nivel == "Estado" else None
        key = (area, estado)
        if key not in grouped:
            grouped[key] = {
                "solicitudes": set(),
                "calendario": [],
                "laboral": [],
            }
        grouped[key]["solicitudes"].add(int(row["solicitud_oid"]))
        grouped[key]["calendario"].append(float(row["duracion_calendario_min"]))
        grouped[key]["laboral"].append(float(row["duracion_laboral_min"]))

    summary_rows: list[Dict[str, Any]] = []
    for (area, estado), values in sorted(grouped.items(), key=lambda item: (item[0][0], item[0][1] or "")):
        calendario = values["calendario"]
        laboral = values["laboral"]
        calendario_stats = summarize_minutes(calendario)
        laboral_stats = summarize_minutes(laboral)
        summary_rows.append(
            {
                "nivel": nivel,
                "periodo": periodo,
                "mes": mes,
                "area": area,
                "estado": estado,
                "permanencias": len(calendario),
                "solicitudes": len(values["solicitudes"]),
                "total_calendario_min": sum(calendario),
                "total_calendario_horas": sum(calendario) / 60.0,
                "promedio_calendario_min": calendario_stats["promedio_minutos"],
                "p10_calendario_min": calendario_stats["p10_minutos"],
                "q1_calendario_min": calendario_stats["q1_minutos"],
                "mediana_calendario_min": calendario_stats["mediana_minutos"],
                "q3_calendario_min": calendario_stats["q3_minutos"],
                "p90_calendario_min": calendario_stats["p90_minutos"],
                "p95_calendario_min": calendario_stats["p95_minutos"],
                "maximo_calendario_min": calendario_stats["maximo_minutos"],
                "total_laboral_min": sum(laboral),
                "total_laboral_horas": sum(laboral) / 60.0,
                "promedio_laboral_min": laboral_stats["promedio_minutos"],
                "p10_laboral_min": laboral_stats["p10_minutos"],
                "q1_laboral_min": laboral_stats["q1_minutos"],
                "mediana_laboral_min": laboral_stats["mediana_minutos"],
                "q3_laboral_min": laboral_stats["q3_minutos"],
                "p90_laboral_min": laboral_stats["p90_minutos"],
                "p95_laboral_min": laboral_stats["p95_minutos"],
                "maximo_laboral_min": laboral_stats["maximo_minutos"],
            }
        )
    return summary_rows


def _best_month_text(month_reports: Sequence[MonthlyReport], metric_key: str) -> str:
    comparable = [
        (report.summary[metric_key]["promedio_minutos"], report.month_label)
        for report in month_reports
        if report.summary[metric_key]["promedio_minutos"] is not None
    ]
    if not comparable:
        return "Sin datos"
    best_value, best_label = min(comparable, key=lambda item: item[0])
    return f"{best_label} ({best_value:.2f} min)"


def write_report_workbook(
    path: Path,
    *,
    month_reports: Sequence[MonthlyReport],
    effective_seed: int,
    run_started_at: datetime,
    dataset_created_at: Optional[datetime] = None,
) -> None:
    if Workbook is None:
        raise RuntimeError(
            "Para generar el Excel con formato se necesita openpyxl. "
            "Instalar con: python -m pip install openpyxl"
        )

    month_values = [report.month_value for report in month_reports]
    period_label = format_period_label(month_values)
    summary = summarize_report_collection(month_reports)

    quantile_stat_fields = [
        ("casos", "Casos"),
        ("promedio_minutos", "Promedio (min)"),
        ("promedio_sin_extremos", "Promedio sin extremos (min)"),
        ("minimo_minutos", "Minimo (min)"),
        ("p10_minutos", "P10 (min)"),
        ("q1_minutos", "Q1 (min)"),
        ("mediana_minutos", "Mediana (min)"),
        ("q3_minutos", "Q3 (min)"),
        ("p90_minutos", "P90 (min)"),
        ("p95_minutos", "P95 (min)"),
        ("maximo_minutos", "Maximo (min)"),
    ]

    base_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("solicitud_oid", "Solicitud"),
        ("nro_socio", "Numero de socio"),
        ("linea", "Linea"),
        ("event_id", "Id de evento"),
        ("fecha", "Fecha"),
        ("texto_original", "Texto original"),
        ("estado_detectado", "Estado detectado"),
        ("estado_detectado_norm", "Estado detectado normalizado"),
        ("created_raw", "Registro original"),
        ("event_ts", "Fecha y hora del evento"),
        ("usuario_evento", "Usuario del evento"),
        ("seq_estado_en_solicitud", "Secuencia de estado en la solicitud"),
        ("estado_anterior", "Estado anterior"),
        ("event_ts_anterior", "Fecha y hora del estado anterior"),
        ("delta_min_desde_estado_anterior", "Minutos desde estado anterior"),
    ]
    first_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("solicitud_oid", "Solicitud"),
        ("nro_socio", "Numero de socio"),
        ("linea", "Linea"),
        ("revision_riesgo_inicio", "Inicio de revision de riesgo"),
        ("primera_respuesta", "Primera respuesta"),
        ("estado_primera_respuesta", "Estado de primera respuesta"),
        ("minutos", "Minutos"),
    ]
    transfer_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("solicitud_oid", "Solicitud"),
        ("nro_socio", "Numero de socio"),
        ("linea", "Linea"),
        ("a_transferir", "Fecha y hora de A Transferir"),
        ("pagada", "Fecha y hora de Pagada"),
        ("minutos", "Minutos"),
    ]
    end_to_end_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("solicitud_oid", "Solicitud"),
        ("nro_socio", "Numero de socio"),
        ("linea", "Linea"),
        ("inicio_proceso", "Inicio del proceso"),
        ("fin_proceso", "Fin del proceso"),
        ("estado_final", "Estado final"),
        ("minutos", "Minutos punta a punta"),
    ]
    end_to_end_status_fields = [
        ("estado_final", "Estado final"),
        *quantile_stat_fields,
    ]
    sample_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("solicitud_oid", "Solicitud"),
        ("nro_socio", "Numero de socio"),
        ("linea", "Linea"),
        ("estado_actual", "Estado actual"),
    ]

    comparison_fields = [
        ("mes", "Mes", "Mes calendario del reporte mensual."),
        ("solicitudes_analizadas", "Solicitudes analizadas", "Cantidad de solicitudes unicas incluidas en el analisis del mes."),
        (
            "primera_respuesta_mediana_min",
            "Mediana tiempo respuesta",
            "Mediana en minutos de primera respuesta: desde la primera RevisionRiesgo hasta el primer cambio posterior, usando solo horario laboral.",
        ),
        (
            "primera_respuesta_promedio_min",
            "Promedio tiempo respuesta",
            "Promedio en minutos de primera respuesta: desde la primera RevisionRiesgo hasta el primer cambio posterior, usando solo horario laboral.",
        ),
        (
            "primera_respuesta_p90_min",
            "P90 tiempo respuesta",
            "Percentil 90 del tiempo de primera respuesta en minutos, usando solo horario laboral.",
        ),
        (
            "transferencia_mediana_min",
            "Mediana tiempo transferencia",
            "Mediana en minutos entre A Transferir y la ultima Pagada asociada, usando solo horario laboral.",
        ),
        (
            "transferencia_promedio_min",
            "Promedio tiempo transferencia",
            "Promedio en minutos entre A Transferir y la ultima Pagada asociada, usando solo horario laboral.",
        ),
        (
            "transferencia_p90_min",
            "P90 tiempo transferencia",
            "Percentil 90 del tiempo de transferencia en minutos, usando solo horario laboral.",
        ),
        ("periodo", "Periodo", "Clave tecnica del mes en formato YYYY-MM."),
        ("novedades_mes", "Novedades del mes", "Cantidad de novedades descargadas para el mes calendario consultado."),
        ("solicitudes_cerradas", "Solicitudes cerradas", "Cantidad de solicitudes cerradas del mes, considerando los estados de cierre definidos por el reporte."),
        ("primera_respuesta_casos", "Primera respuesta casos", "Cantidad de solicitudes que tienen medicion valida de primera respuesta."),
        ("primera_respuesta_p10_min", "Primera respuesta P10 (min)", "Percentil 10 del tiempo de primera respuesta en minutos."),
        ("primera_respuesta_q1_min", "Primera respuesta Q1 (min)", "Primer cuartil del tiempo de primera respuesta en minutos."),
        ("primera_respuesta_mediana_min", "Primera respuesta mediana (min)", "Mediana del tiempo de primera respuesta en minutos."),
        ("primera_respuesta_q3_min", "Primera respuesta Q3 (min)", "Tercer cuartil del tiempo de primera respuesta en minutos."),
        ("primera_respuesta_p95_min", "Primera respuesta P95 (min)", "Percentil 95 del tiempo de primera respuesta en minutos."),
        ("transferencia_casos", "Transferencia casos", "Cantidad de solicitudes que tienen medicion valida de transferencia."),
        ("transferencia_p10_min", "Transferencia P10 (min)", "Percentil 10 del tiempo de transferencia en minutos."),
        ("transferencia_q1_min", "Transferencia Q1 (min)", "Primer cuartil del tiempo de transferencia en minutos."),
        ("transferencia_mediana_min", "Transferencia mediana (min)", "Mediana del tiempo de transferencia en minutos."),
        ("transferencia_q3_min", "Transferencia Q3 (min)", "Tercer cuartil del tiempo de transferencia en minutos."),
        ("transferencia_p95_min", "Transferencia P95 (min)", "Percentil 95 del tiempo de transferencia en minutos."),
        ("punta_a_punta_casos", "Punta a punta casos", "Cantidad de solicitudes que tienen medicion valida de punta a punta."),
        ("punta_a_punta_promedio_min", "Punta a punta promedio (min)", "Promedio en minutos desde el primer estado registrado hasta el cierre final, usando solo horario laboral."),
        ("punta_a_punta_p10_min", "Punta a punta P10 (min)", "Percentil 10 del tiempo punta a punta en minutos."),
        ("punta_a_punta_q1_min", "Punta a punta Q1 (min)", "Primer cuartil del tiempo punta a punta en minutos."),
        ("punta_a_punta_mediana_min", "Punta a punta mediana (min)", "Mediana del tiempo punta a punta en minutos."),
        ("punta_a_punta_q3_min", "Punta a punta Q3 (min)", "Tercer cuartil del tiempo punta a punta en minutos."),
        ("punta_a_punta_p90_min", "Punta a punta P90 (min)", "Percentil 90 del tiempo punta a punta en minutos."),
        ("punta_a_punta_p95_min", "Punta a punta P95 (min)", "Percentil 95 del tiempo punta a punta en minutos."),
        ("punta_a_punta_promedio_sin_extremos_min", "Punta a punta promedio sin extremos (min)", "Promedio recortado del tiempo punta a punta, excluyendo 5% inferior y 5% superior cuando aplica."),
        ("legajos_muestreados", "Legajos muestreados", "Cantidad de legajos incluidos en la muestra reproducible del mes."),
    ]
    comparison_headers = [label for _, label, _ in comparison_fields]
    comparison_comments = [description for _, _, description in comparison_fields]
    comparison_rows = [
        [
            report.month_label,
            report.summary["solicitudes_count"],
            report.summary["first_response"]["mediana_minutos"],
            report.summary["first_response"]["promedio_minutos"],
            report.summary["first_response"]["p90_minutos"],
            report.summary["transfer"]["mediana_minutos"],
            report.summary["transfer"]["promedio_minutos"],
            report.summary["transfer"]["p90_minutos"],
            report.month_value,
            report.summary["month_novedades_count"],
            report.summary["closed_solicitudes_count"],
            report.summary["first_response"]["cases"],
            report.summary["first_response"]["p10_minutos"],
            report.summary["first_response"]["q1_minutos"],
            report.summary["first_response"]["mediana_minutos"],
            report.summary["first_response"]["q3_minutos"],
            report.summary["first_response"]["p95_minutos"],
            report.summary["transfer"]["cases"],
            report.summary["transfer"]["p10_minutos"],
            report.summary["transfer"]["q1_minutos"],
            report.summary["transfer"]["mediana_minutos"],
            report.summary["transfer"]["q3_minutos"],
            report.summary["transfer"]["p95_minutos"],
            report.summary["end_to_end"]["cases"],
            report.summary["end_to_end"]["promedio_minutos"],
            report.summary["end_to_end"]["p10_minutos"],
            report.summary["end_to_end"]["q1_minutos"],
            report.summary["end_to_end"]["mediana_minutos"],
            report.summary["end_to_end"]["q3_minutos"],
            report.summary["end_to_end"]["p90_minutos"],
            report.summary["end_to_end"]["p95_minutos"],
            report.summary["end_to_end"]["promedio_sin_extremos"],
            report.summary["legajos_sample_count"],
        ]
        for report in month_reports
    ]

    overview_rows = [
        ("Periodo analizado", period_label),
        ("Fecha de ejecucion", run_started_at.strftime("%d/%m/%Y %H:%M:%S")),
        ("Meses incluidos", len(month_reports)),
        ("Novedades del rango", summary["total_novedades"]),
        ("Solicitudes cerradas consideradas", summary["total_cerradas"]),
        ("Novedades historicas relevadas", summary["total_historical"]),
        ("Solicitudes unicas analizadas", summary["total_solicitudes"]),
        ("Primera respuesta promedio", _format_card_metric(summary["overall_first_summary"]["promedio_minutos"])),
        ("Primera respuesta mediana", _format_card_metric(summary["overall_first_summary"]["mediana_minutos"])),
        ("Primera respuesta P90", _format_card_metric(summary["overall_first_summary"]["p90_minutos"])),
        ("Transferencia promedio", _format_card_metric(summary["overall_transfer_summary"]["promedio_minutos"])),
        ("Transferencia mediana", _format_card_metric(summary["overall_transfer_summary"]["mediana_minutos"])),
        ("Transferencia P90", _format_card_metric(summary["overall_transfer_summary"]["p90_minutos"])),
        ("Punta a punta promedio", _format_card_metric(summary["overall_end_to_end_summary"]["promedio_minutos"])),
        (
            "Punta a punta promedio sin extremos",
            _format_card_metric(summary["overall_end_to_end_summary"]["promedio_sin_extremos"]),
        ),
        ("Punta a punta mediana", _format_card_metric(summary["overall_end_to_end_summary"]["mediana_minutos"])),
        ("Punta a punta P90", _format_card_metric(summary["overall_end_to_end_summary"]["p90_minutos"])),
        (
            "Regla de primera respuesta",
            month_reports[0].summary["first_response_rule"],
        ),
        (
            "Regla de transferencia",
            month_reports[0].summary["transfer_rule"],
        ),
        (
            "Definicion de punta a punta",
            "desde el primer estado registrado de la solicitud hasta su estado de cierre final, excluyendo lineas Carlos Paz, usando solo horario laboral",
        ),
        (
            "Lineas excluidas del analisis",
            ", ".join(month_reports[0].summary["analysis_excluded_line_keywords"]),
        ),
        (
            "Solicitudes excluidas del analisis por linea",
            summary.get("analysis_excluded_solicitudes_count", 0),
        ),
        (
            "Lineas excluidas para transferencia",
            ", ".join(month_reports[0].summary["excluded_line_keywords_transfer"]),
        ),
        ("Semilla base de muestreo", effective_seed),
        ("Mejor mes en primera respuesta", _best_month_text(month_reports, "first_response")),
        ("Mejor mes en transferencia", _best_month_text(month_reports, "transfer")),
        ("Mejor mes punta a punta", _best_month_text(month_reports, "end_to_end")),
    ]
    if dataset_created_at is not None:
        overview_rows.insert(2, ("Base SQLite generada", dataset_created_at.strftime("%d/%m/%Y %H:%M:%S")))

    kpi_cards = [
        {"title": "Meses analizados", "value": str(len(month_reports))},
        {"title": "Solicitudes analizadas", "value": str(summary["total_solicitudes"])},
        {"title": "Prom. 1ra respuesta", "value": _format_card_metric(summary["overall_first_summary"]["promedio_minutos"])},
        {"title": "Med. 1ra respuesta", "value": _format_card_metric(summary["overall_first_summary"]["mediana_minutos"])},
        {"title": "Prom. transferencia", "value": _format_card_metric(summary["overall_transfer_summary"]["promedio_minutos"])},
        {"title": "Med. transferencia", "value": _format_card_metric(summary["overall_transfer_summary"]["mediana_minutos"])},
        {"title": "Prom. punta a punta", "value": _format_card_metric(summary["overall_end_to_end_summary"]["promedio_minutos"])},
    ]

    base_rows = [row for report in month_reports for row in _with_period_context(report, report.base_rows)]
    first_rows = [row for report in month_reports for row in _with_period_context(report, report.first_response)]
    transfer_rows = [row for report in month_reports for row in _with_period_context(report, report.transfer)]
    end_to_end_rows = [row for report in month_reports for row in _with_period_context(report, report.end_to_end)]
    sample_rows = [row for report in month_reports for row in _with_period_context(report, report.legajos_sample)]
    state_duration_rows = [row for report in month_reports for row in _build_state_duration_rows(report)]
    metric_quantile_rows = []
    metric_catalog = [
        ("Primera respuesta", "first_response"),
        ("Transferencia", "transfer"),
        ("Punta a punta", "end_to_end"),
    ]
    for report in month_reports:
        for metric_label, metric_key in metric_catalog:
            stats = report.summary[metric_key]
            metric_quantile_rows.append(
                {
                    "periodo": report.month_value,
                    "mes": report.month_label,
                    "metrica": metric_label,
                    "casos": stats["cases"],
                    **stats,
                }
            )
    overall_metric_rows = [
        {
            "periodo": "TOTAL",
            "mes": period_label,
            "metrica": "Primera respuesta",
            "casos": summary["overall_first_summary"].get("cases", len([row for row in first_rows if row.get("minutos") is not None])),
            **summary["overall_first_summary"],
        },
        {
            "periodo": "TOTAL",
            "mes": period_label,
            "metrica": "Transferencia",
            "casos": summary["overall_transfer_summary"].get("cases", len([row for row in transfer_rows if row.get("minutos") is not None])),
            **summary["overall_transfer_summary"],
        },
        {
            "periodo": "TOTAL",
            "mes": period_label,
            "metrica": "Punta a punta",
            "casos": summary["overall_end_to_end_summary"].get("cases", len([row for row in end_to_end_rows if row.get("minutos") is not None])),
            **summary["overall_end_to_end_summary"],
        },
    ]
    metric_quantile_fields = [
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("metrica", "Metrica"),
        *quantile_stat_fields,
    ]
    state_time_fields = [
        ("nivel", "Nivel"),
        ("periodo", "Periodo"),
        ("mes", "Mes"),
        ("area", "Area"),
        ("estado", "Estado"),
        ("permanencias", "Permanencias"),
        ("solicitudes", "Solicitudes"),
        ("total_calendario_min", "Total calendario (min)"),
        ("total_calendario_horas", "Total calendario (horas)"),
        ("promedio_calendario_min", "Promedio calendario (min)"),
        ("p10_calendario_min", "P10 calendario (min)"),
        ("q1_calendario_min", "Q1 calendario (min)"),
        ("mediana_calendario_min", "Mediana calendario (min)"),
        ("q3_calendario_min", "Q3 calendario (min)"),
        ("p90_calendario_min", "P90 calendario (min)"),
        ("p95_calendario_min", "P95 calendario (min)"),
        ("maximo_calendario_min", "Maximo calendario (min)"),
        ("total_laboral_min", "Total laboral (min)"),
        ("total_laboral_horas", "Total laboral (horas)"),
        ("promedio_laboral_min", "Promedio laboral (min)"),
        ("p10_laboral_min", "P10 laboral (min)"),
        ("q1_laboral_min", "Q1 laboral (min)"),
        ("mediana_laboral_min", "Mediana laboral (min)"),
        ("q3_laboral_min", "Q3 laboral (min)"),
        ("p90_laboral_min", "P90 laboral (min)"),
        ("p95_laboral_min", "P95 laboral (min)"),
        ("maximo_laboral_min", "Maximo laboral (min)"),
    ]
    state_time_summary_rows = []
    for report in month_reports:
        month_state_rows = _build_state_duration_rows(report)
        state_time_summary_rows.extend(
            _state_time_summary_rows(
                month_state_rows,
                periodo=report.month_value,
                mes=report.month_label,
                nivel="Area",
            )
        )
        state_time_summary_rows.extend(
            _state_time_summary_rows(
                month_state_rows,
                periodo=report.month_value,
                mes=report.month_label,
                nivel="Estado",
            )
        )
    state_time_summary_rows.extend(
        _state_time_summary_rows(
            state_duration_rows,
            periodo="TOTAL",
            mes=period_label,
            nivel="Area",
        )
    )
    state_time_summary_rows.extend(
        _state_time_summary_rows(
            state_duration_rows,
            periodo="TOTAL",
            mes=period_label,
            nivel="Estado",
        )
    )
    misc_states = sorted(
        {
            str(row.get("estado"))
            for row in state_duration_rows
            if _classify_state_area(row.get("estado")) == "Misc" and row.get("estado")
        }
    )

    detail_sheets = [
        {
            "name": "Base historica",
            "title": "Base historica de estados",
            "subtitle": f"Periodo {period_label}. Cada fila representa un cambio de estado detectado para una solicitud cerrada en el rango.",
            "headers": [label for _, label in base_fields],
            "rows": [[row.get(key) for key, _ in base_fields] for row in base_rows],
            "table_name": "BaseHistorica",
            "datetime_cols": [12, 16],
            "float_cols": [17],
            "int_cols": [3, 4, 6, 14],
            "max_width": 46,
        },
        {
            "name": "Primera respuesta",
            "title": "Detalle de primera respuesta",
            "subtitle": f"Periodo {period_label}. Tiempo entre la primera RevisionRiesgo y el primer cambio posterior.",
            "headers": [label for _, label in first_fields],
            "rows": [[row.get(key) for key, _ in first_fields] for row in first_rows],
            "table_name": "PrimeraRespuesta",
            "datetime_cols": [6, 7],
            "float_cols": [9],
            "int_cols": [3, 4],
            "max_width": 28,
        },
        {
            "name": "Transferencia",
            "title": "Detalle de transferencia",
            "subtitle": f"Periodo {period_label}. Tiempo entre A Transferir y la ultima Pagada previa.",
            "headers": [label for _, label in transfer_fields],
            "rows": [[row.get(key) for key, _ in transfer_fields] for row in transfer_rows],
            "table_name": "Transferencia",
            "datetime_cols": [6, 7],
            "float_cols": [8],
            "int_cols": [3, 4],
            "max_width": 28,
        },
        {
            "name": "Punta a punta",
            "title": "Detalle de punta a punta",
            "subtitle": f"Periodo {period_label}. Tiempo desde el primer estado registrado hasta el cierre final de la solicitud.",
            "headers": [label for _, label in end_to_end_fields],
            "rows": [[row.get(key) for key, _ in end_to_end_fields] for row in end_to_end_rows],
            "table_name": "PuntaAPunta",
            "datetime_cols": [6, 7],
            "float_cols": [9],
            "int_cols": [3, 4],
            "max_width": 28,
        },
        {
            "name": "Cuantiles metricas",
            "title": "Cuantiles por metrica",
            "subtitle": f"Periodo {period_label}. Incluye estadisticos mensuales y consolidado para primera respuesta, transferencia y punta a punta.",
            "headers": [label for _, label in metric_quantile_fields],
            "rows": [
                [row.get(key) for key, _ in metric_quantile_fields]
                for row in [*metric_quantile_rows, *overall_metric_rows]
            ],
            "table_name": "CuantilesMetricas",
            "float_cols": [index + 1 for index, (key, _) in enumerate(metric_quantile_fields) if key.endswith("_minutos")],
            "int_cols": [4],
            "max_width": 24,
        },
        {
            "name": "Punta a punta final",
            "title": "Punta a punta por estado final",
            "subtitle": f"Periodo {period_label}. Resumen consolidado del tiempo punta a punta segmentado por estado final.",
            "headers": [label for _, label in end_to_end_status_fields],
            "rows": [[row.get(key) for key, _ in end_to_end_status_fields] for row in summary["overall_end_to_end_by_status"]],
            "table_name": "PuntaAPuntaPorEstadoFinal",
            "float_cols": [index + 1 for index, (key, _) in enumerate(end_to_end_status_fields) if key.endswith("_minutos")],
            "int_cols": [2],
            "max_width": 30,
        },
        {
            "name": "Tiempos por estado",
            "title": "Tiempos por estado y area",
            "subtitle": (
                f"Periodo {period_label}. Permanencia reconstruida entre cambios consecutivos de estado. "
                f"Estados en Misc: {', '.join(misc_states) if misc_states else 'ninguno'}."
            ),
            "headers": [label for _, label in state_time_fields],
            "rows": [[row.get(key) for key, _ in state_time_fields] for row in state_time_summary_rows],
            "table_name": "TiemposPorEstado",
            "float_cols": [index + 1 for index, (key, _) in enumerate(state_time_fields) if key.endswith("_min") or key.endswith("_horas")],
            "int_cols": [6, 7],
            "max_width": 24,
        },
        {
            "name": "Muestreo legajos",
            "title": "Muestreo de legajos pagados",
            "subtitle": f"Periodo {period_label}. Muestra reproducible de hasta 30 solicitudes pagadas por mes.",
            "headers": [label for _, label in sample_fields],
            "rows": [[row.get(key) for key, _ in sample_fields] for row in sample_rows],
            "table_name": "MuestreoLegajos",
            "int_cols": [3, 4],
            "max_width": 24,
        },
    ]

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = _sanitize_sheet_name("Resumen ejecutivo")
    _write_summary_sheet(
        summary_ws,
        period_label=period_label,
        run_started_at=run_started_at,
        overview_rows=overview_rows,
        kpi_cards=kpi_cards,
    )

    comparison_ws = workbook.create_sheet(_sanitize_sheet_name("Comparativo mensual"))
    _write_table_sheet(
        comparison_ws,
        title="Comparativo mensual",
        subtitle="Indicadores consolidados por mes para facilitar seguimiento y comparación.",
        headers=comparison_headers,
        rows=comparison_rows,
        table_name="ComparativoMensual",
        header_comments=comparison_comments,
        hidden_header_descriptions=comparison_comments,
        float_cols=[3, 4, 5, 6, 7, 8, 13, 14, 15, 16, 17, 19, 20, 21, 22, 23, 25, 26, 27, 28, 29, 30, 31, 32],
        int_cols=[2, 10, 11, 12, 18, 24, 33],
        max_width=24,
    )
    if comparison_rows:
        for col_index in (3, 4, 5, 6, 7, 8, 25, 28, 30, 32):
            column_letter = get_column_letter(col_index)
            comparison_ws.conditional_formatting.add(
                f"{column_letter}5:{column_letter}{4 + len(comparison_rows)}",
                ColorScaleRule(
                    start_type="min",
                    start_color="7BC67E",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="F4E7B2",
                    end_type="max",
                    end_color="E1B77A",
                ),
            )

    for sheet in detail_sheets:
        ws = workbook.create_sheet(_sanitize_sheet_name(str(sheet["name"])))
        _write_table_sheet(
            ws,
            title=str(sheet["title"]),
            subtitle=str(sheet["subtitle"]),
            headers=list(sheet["headers"]),
            rows=list(sheet["rows"]),
            table_name=str(sheet["table_name"]),
            datetime_cols=sheet.get("datetime_cols"),
            float_cols=sheet.get("float_cols"),
            int_cols=sheet.get("int_cols"),
            max_width=int(sheet.get("max_width", 42)),
        )

    workbook.save(path)
