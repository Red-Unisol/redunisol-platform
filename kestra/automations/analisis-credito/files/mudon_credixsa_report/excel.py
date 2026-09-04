from __future__ import annotations

from datetime import datetime
import json
import os
from pathlib import Path
import shutil
import tempfile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "Numero de socio",
    "Nombre completo",
    "DNI",
    "CUIL",
    "Cuenta(s) activa(s)",
    "Linea(s) MUDON",
    "Empleador(es) CredixSA",
    "CUIT empleador(es)",
    "Ultimo periodo informado",
    "Jubilado ISSN",
    "Estado consulta",
    "Fuente",
    "Fecha consulta",
    "Observacion/error",
]


def decode_list(raw: Any) -> list[Any]:
    if isinstance(raw, list):
        return raw
    try:
        value = json.loads(str(raw or "[]"))
    except json.JSONDecodeError:
        return []
    return value if isinstance(value, list) else []


def period_sort_key(value: str) -> tuple[int, int]:
    try:
        month, year = value.split("/", 1)
        return int(year), int(month)
    except (TypeError, ValueError):
        return 0, 0


def employer_values(member: dict[str, Any]) -> tuple[str, str, str]:
    employers = [item for item in decode_list(member.get("employers_json")) if isinstance(item, dict)]
    names = sorted({str(item.get("nombre") or "").strip() for item in employers if str(item.get("nombre") or "").strip()})
    cuits = sorted({str(item.get("cuit") or "").strip() for item in employers if str(item.get("cuit") or "").strip()})
    periods = []
    for employer in employers:
        for period in employer.get("periodos") or []:
            if isinstance(period, dict) and period.get("periodo"):
                periods.append(str(period["periodo"]))
    latest_period = max(periods, key=period_sort_key) if periods else ""
    return " | ".join(names), " | ".join(cuits), latest_period


def write_workbook(path: Path, run: dict[str, Any], members: list[dict[str, Any]], stats: dict[str, int]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary_rows = [
        ("Reporte", "Socios MUDON enriquecidos con CredixSA"),
        ("Corrida", run["run_id"]),
        ("Generado", datetime.now().astimezone().isoformat(timespec="seconds")),
        ("Estado", "Completado con errores" if stats["errors"] else "Completado"),
        ("Socios unicos", stats["total"]),
        ("Procesados", stats["completed"]),
        ("Reutilizados desde cache", stats["cache"]),
        ("Consultados online", stats["online"]),
        ("Jubilados ISSN", stats["qualifying"]),
        ("Errores", stats["errors"]),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 48
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    sheet = workbook.create_sheet("Socios")
    sheet.append(HEADERS)
    for member in members:
        employer_names, employer_cuits, latest_period = employer_values(member)
        sheet.append(
            [
                member.get("member_number", ""),
                member.get("full_name", ""),
                member.get("dni", ""),
                member.get("cuit", ""),
                " | ".join(str(item) for item in decode_list(member.get("loan_accounts_json"))),
                " | ".join(str(item) for item in decode_list(member.get("loan_lines_json"))),
                employer_names,
                employer_cuits,
                latest_period,
                "Si" if int(member.get("qualifies_issn") or 0) else "No",
                member.get("credix_status") or member.get("status", ""),
                member.get("result_source", ""),
                member.get("checked_at", ""),
                member.get("error", ""),
            ]
        )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [18, 34, 14, 18, 24, 30, 52, 28, 24, 16, 20, 14, 28, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    if len(members) > 0:
        table = Table(displayName="SociosMudon", ref=f"A1:N{len(members) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    workbook.save(path)


def atomic_publish(source: Path, reports_root: Path, now: datetime) -> tuple[Path, Path]:
    report_dir = reports_root / "cobranzas" / "mudon-jubilados"
    history_dir = report_dir / "historico"
    history_dir.mkdir(parents=True, exist_ok=True)
    latest = report_dir / "ultimo.xlsx"
    history = history_dir / f"{now:%Y-%m-%d}.xlsx"
    for destination in (latest, history):
        with tempfile.NamedTemporaryFile(dir=destination.parent, suffix=".xlsx", delete=False) as handle:
            temporary = Path(handle.name)
        try:
            shutil.copyfile(source, temporary)
            os.replace(temporary, destination)
            destination.chmod(0o644)
        finally:
            temporary.unlink(missing_ok=True)
    return latest, history
