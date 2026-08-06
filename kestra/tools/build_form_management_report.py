#!/usr/bin/env python3
"""Build the management workbook from the Kestra execution and lead cross-check exports."""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NAVY = "17365D"
WHITE = "FFFFFF"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
ORANGE = "FCE4D6"
RED = "FFC7CE"
GRAY = "E7E6E6"


def read_sheet(path: Path, sheet: str) -> list[dict[str, Any]]:
    ws = load_workbook(path, read_only=True, data_only=True)[sheet]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    return [dict(zip(headers, row)) for row in rows]


def compact_table(ws, *, filter_: bool = True) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 18
        for cell in ws[row]:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
    if filter_ and ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def widths(ws, caps: dict[int, int] | None = None) -> None:
    caps = caps or {}
    for idx, column in enumerate(ws.columns, 1):
        longest = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(longest, caps.get(idx, 36))


def normalize_id(value: Any) -> str:
    if value in (None, ""):
        return ""
    return str(value)


def classify_not_sent(row: dict[str, Any], cross: dict[str, Any] | None) -> tuple[str, str, str]:
    if cross:
        result = cross.get("resultado_cruce")
        message = str(cross.get("message_hijo") or row.get("razon_error") or row.get("message") or "")
        persist_state = str(cross.get("estado_tarea_persistir") or "")
        if result == "SIN SUBEJECUCIÓN" and persist_state not in {"SKIPPED", "WARNING"}:
            return "Error técnico", message or "La persistencia no pudo ejecutarse", "Reintentar la persistencia"
        if "más de un contacto" in message or "mas de un contacto" in message:
            return "Duplicidad en Bitrix", message, "Unificar contactos duplicados por CUIL"
        if result != "SIN SUBEJECUCIÓN":
            return "Pendiente de verificación", message or str(result), "Revisar manualmente en Bitrix"

    action = str(row.get("action") or "").lower()
    message = str(row.get("razon_error") or row.get("message") or "")
    reason = str(row.get("reason") or "")
    if row.get("estado") in {"FAILED", "KILLED"}:
        return "Error técnico", message or reason, "Reintentar y revisar logs"
    if action == "rejected":
        return "Rechazo comercial", message or reason, "Sin acción técnica"
    if action == "error":
        lower = message.lower()
        if "falta el campo" in lower:
            return "Dato requerido faltante", message, "Validar el campo en el frontend"
        if "valor no soportado" in lower:
            return "Valor no soportado", message, "Normalizar catálogo y opciones"
        if "formato" in lower or "digitos" in lower:
            return "Formato inválido", message, "Agregar validación en el frontend"
        if "mas de un contacto" in lower or "más de un contacto" in lower:
            return "Duplicidad en Bitrix", message, "Unificar contactos duplicados por CUIL"
        return "Error lógico o de datos", message or reason, "Revisar validaciones"
    return "Pendiente de verificación", message or reason or action, "Revisar ejecución"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--executions", required=True, type=Path)
    parser.add_argument("--cross", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    executions = read_sheet(args.executions, "Ejecuciones")
    cross_rows = read_sheet(args.cross, "Cruce")
    cross_by_parent = {str(x["parent_execution_id"]): x for x in cross_rows}

    leads: dict[str, dict[str, Any]] = {}
    sent_parent_ids: set[str] = set()
    for row in executions:
        lead_id = normalize_id(row.get("lead_id"))
        if lead_id:
            leads[lead_id] = {
                **row, "lead_id": lead_id, "metodo_confirmacion": "Flow principal",
                "child_execution_id": "", "estado_bitrix": "Lead confirmado",
            }
            sent_parent_ids.add(str(row["execution_id"]))
    for cross in cross_rows:
        lead_id = normalize_id(cross.get("lead_id"))
        if not lead_id:
            continue
        parent_id = str(cross["parent_execution_id"])
        parent = next((x for x in executions if str(x["execution_id"]) == parent_id), {})
        leads[lead_id] = {
            **parent, "lead_id": lead_id, "contact_id": normalize_id(cross.get("contact_id")),
            "metodo_confirmacion": "Subejecución persistir_bitrix",
            "child_execution_id": cross.get("child_execution_id"), "estado_bitrix": "Lead confirmado",
            "action": cross.get("action_hijo") or parent.get("action"),
            "reason": cross.get("reason_hijo") or parent.get("reason"),
            "message": cross.get("message_hijo") or parent.get("message"),
        }
        sent_parent_ids.add(parent_id)

    not_sent = []
    for row in executions:
        parent_id = str(row["execution_id"])
        if parent_id in sent_parent_ids:
            continue
        category, reason, recommendation = classify_not_sent(row, cross_by_parent.get(parent_id))
        not_sent.append({**row, "categoria": category, "motivo_final": reason, "accion_recomendada": recommendation})

    wb = Workbook()
    summary = wb.active
    summary.title = "Resumen Ejecutivo"
    summary.append(["INFORME EJECUTIVO FORMULARIO → BITRIX", "Valor"])
    total = len(executions)
    lead_count = len(leads)
    rejected = sum(x["categoria"] == "Rechazo comercial" for x in not_sent)
    leads_approved = sum(str(x.get("action") or "").lower() != "rejected" for x in leads.values())
    leads_rejected = sum(str(x.get("action") or "").lower() == "rejected" for x in leads.values())
    data_errors = sum(x["categoria"] not in {"Rechazo comercial", "Error técnico", "Pendiente de verificación"} for x in not_sent)
    technical = sum(x["categoria"] == "Error técnico" for x in not_sent)
    pending = sum(x["categoria"] == "Pendiente de verificación" for x in not_sent)
    dates = [x.get("inicio") for x in executions if isinstance(x.get("inicio"), datetime)]
    metrics = [
        ("Período", f"{min(dates):%d/%m/%Y} al {max(dates):%d/%m/%Y}" if dates else ""),
        ("Formularios recibidos", total),
        ("Leads confirmados en Bitrix", lead_count),
        ("Leads aprobados/ingresados", leads_approved),
        ("Leads rechazados en Bitrix", leads_rejected),
        ("Conversión formulario → lead", lead_count / total if total else 0),
        ("Rechazados antes de llegar a Bitrix", rejected),
        ("Errores de datos", data_errors),
        ("Errores técnicos", technical),
        ("Pendientes de verificación", pending),
        ("Formularios sin lead", len(not_sent)),
    ]
    for metric in metrics:
        summary.append(metric)
    for row in range(2, summary.max_row + 1):
        summary.cell(row, 1).font = Font(bold=True, color=NAVY)
        label = summary.cell(row, 1).value
        color = GREEN if label in {"Leads confirmados en Bitrix", "Leads aprobados/ingresados"} else RED if label in {"Errores técnicos", "Pendientes de verificación"} else YELLOW if label in {"Leads rechazados en Bitrix", "Rechazados antes de llegar a Bitrix", "Errores de datos"} else WHITE
        for cell in summary[row]: cell.fill = PatternFill("solid", fgColor=color)
        if label == "Conversión formulario → lead": summary.cell(row, 2).number_format = "0.00%"
    compact_table(summary, filter_=False)
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 28

    # Helper data for management charts, placed away from the visible KPI block.
    summary["D1"], summary["E1"] = "Resultado", "Cantidad"
    funnel = [("Leads aprobados", leads_approved), ("Leads rechazados en Bitrix", leads_rejected),
              ("Rechazados antes de Bitrix", rejected),
              ("Errores de datos", data_errors), ("Errores técnicos", technical),
              ("Pendientes", pending)]
    for idx, item in enumerate(funnel, 2):
        summary.cell(idx, 4, item[0]); summary.cell(idx, 5, item[1])
    pie = PieChart(); pie.title = "Resultado de formularios"
    pie.add_data(Reference(summary, min_col=5, min_row=1, max_row=1 + len(funnel)), titles_from_data=True)
    pie.set_categories(Reference(summary, min_col=4, min_row=2, max_row=1 + len(funnel)))
    pie.height = 8; pie.width = 12; summary.add_chart(pie, "G2")

    leads_ws = wb.create_sheet("Leads en Bitrix")
    lead_headers = ["fecha", "lead_id", "contact_id", "nombre", "cuil", "email", "whatsapp", "provincia",
                    "situacion_laboral", "banco_cobro", "origen", "utm_source", "utm_medium", "utm_campaign",
                    "landing", "resultado_kestra", "motivo", "estado_bitrix", "metodo_confirmacion",
                    "ejecucion_kestra", "subejecucion", "revision"]
    leads_ws.append(lead_headers)
    for row in sorted(leads.values(), key=lambda x: str(x.get("inicio") or "")):
        leads_ws.append([row.get("inicio"), row.get("lead_id"), row.get("contact_id"), row.get("nombre"),
                         row.get("cuil"), row.get("email"), row.get("whatsapp"), row.get("provincia"),
                         row.get("situacion_laboral"), row.get("banco_cobro"), row.get("origen_lead"),
                         row.get("utm_source"), row.get("utm_medium"), row.get("utm_campaign"), row.get("landing"),
                         row.get("action"), row.get("reason"), row.get("estado_bitrix"), row.get("metodo_confirmacion"),
                         row.get("execution_id"), row.get("child_execution_id"), row.get("revision")])
    compact_table(leads_ws)
    for row in range(2, leads_ws.max_row + 1): leads_ws.cell(row, 18).fill = PatternFill("solid", fgColor=GREEN)
    widths(leads_ws, {1: 20, 4: 28, 6: 32, 10: 38, 14: 42, 19: 32, 20: 28, 21: 28})

    no_ws = wb.create_sheet("No enviados")
    no_headers = ["fecha", "nombre", "cuil", "email", "whatsapp", "provincia", "origen", "resultado",
                  "categoria", "motivo", "accion_recomendada", "etapa_alcanzada", "ejecucion_kestra", "revision"]
    no_ws.append(no_headers)
    for row in sorted(not_sent, key=lambda x: str(x.get("inicio") or "")):
        cross = cross_by_parent.get(str(row["execution_id"]))
        stage = "Subejecución de persistencia" if cross and cross.get("child_execution_id") else "Flow principal"
        no_ws.append([row.get("inicio"), row.get("nombre"), row.get("cuil"), row.get("email"), row.get("whatsapp"),
                      row.get("provincia"), row.get("origen_lead"), row.get("resultado"), row["categoria"],
                      row["motivo_final"], row["accion_recomendada"], stage, row.get("execution_id"), row.get("revision")])
    compact_table(no_ws)
    for row in range(2, no_ws.max_row + 1):
        category = no_ws.cell(row, 9).value
        color = YELLOW if category == "Rechazo comercial" else RED if category in {"Error técnico", "Pendiente de verificación"} else ORANGE
        no_ws.cell(row, 9).fill = PatternFill("solid", fgColor=color)
    widths(no_ws, {2: 28, 4: 32, 9: 28, 10: 62, 11: 44, 13: 28})

    quality = wb.create_sheet("Motivos y Calidad")
    quality.append(["Categoría", "Motivo", "Cantidad", "% total", "% no enviados", "Acción recomendada"])
    reason_counts = Counter((x["categoria"], x["motivo_final"], x["accion_recomendada"]) for x in not_sent)
    for (category, reason, recommendation), count in reason_counts.most_common():
        quality.append([category, reason, count, count / total if total else 0,
                        count / len(not_sent) if not_sent else 0, recommendation])
    compact_table(quality)
    for row in range(2, quality.max_row + 1):
        quality.cell(row, 4).number_format = quality.cell(row, 5).number_format = "0.00%"
    widths(quality, {1: 28, 2: 72, 6: 48})
    if quality.max_row > 1:
        chart = BarChart(); chart.type = "bar"; chart.style = 10; chart.title = "Principales motivos de no creación"
        max_chart_row = min(quality.max_row, 11)
        chart.add_data(Reference(quality, min_col=3, min_row=1, max_row=max_chart_row), titles_from_data=True)
        chart.set_categories(Reference(quality, min_col=2, min_row=2, max_row=max_chart_row))
        chart.height = 9; chart.width = 17; quality.add_chart(chart, "H2")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"{args.output}|formularios={total}|leads={lead_count}|no_enviados={len(not_sent)}")


if __name__ == "__main__":
    main()
