#!/usr/bin/env python3
"""Export Kestra form executions and failed-execution logs to a styled XLSX."""

from __future__ import annotations

import argparse
import json
import os
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


COLORS = {
    "navy": "17365D",
    "blue": "5B9BD5",
    "green": "C6EFCE",
    "green_text": "006100",
    "red": "FFC7CE",
    "red_text": "9C0006",
    "yellow": "FFEB9C",
    "yellow_text": "9C6500",
    "gray": "E7E6E6",
    "white": "FFFFFF",
}

SENSITIVE_HEADERS = {
    "authorization", "cookie", "set-cookie", "x-api-key", "proxy-authorization"
}


def api_get(session: requests.Session, url: str, *, params: dict[str, Any] | None = None) -> Any:
    last_error: Exception | None = None
    for attempt in range(5):
        try:
            response = session.get(url, params=params, timeout=(10, 90))
            response.raise_for_status()
            return response.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            if attempt == 4:
                break
            time.sleep(2 ** attempt)
    raise RuntimeError(f"No se pudo consultar {url}: {last_error}")


def fetch_executions(session: requests.Session, base_url: str, tenant: str,
                     namespace: str, flow_id: str) -> list[dict[str, Any]]:
    endpoint = f"{base_url}/api/v1/{tenant}/executions/search"
    page, size, rows = 1, 100, []
    while True:
        payload = api_get(session, endpoint, params={
            "namespace": namespace, "flowId": flow_id, "page": page, "size": size,
        })
        batch = payload.get("results", [])
        rows.extend(batch)
        if not batch or len(rows) >= int(payload.get("total", len(rows))):
            return rows
        page += 1


def fetch_logs(session: requests.Session, base_url: str, tenant: str,
               execution_id: str) -> list[dict[str, Any]]:
    payload = api_get(session, f"{base_url}/api/v1/{tenant}/logs/{execution_id}")
    return payload if isinstance(payload, list) else payload.get("results", [])


def state_of(item: dict[str, Any]) -> str:
    return str((item.get("state") or {}).get("current") or "UNKNOWN")


def iso_value(value: Any) -> Any:
    if not value or not isinstance(value, str):
        return value
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return parsed.replace(tzinfo=None)
    except ValueError:
        return value


def json_text(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)


def safe_headers(headers: Any) -> dict[str, Any]:
    if not isinstance(headers, dict):
        return {}
    return {k: v for k, v in headers.items() if k.lower() not in SENSITIVE_HEADERS}


def error_summary(logs: list[dict[str, Any]], execution: dict[str, Any]) -> str:
    error_messages = [str(x.get("message", "")).strip() for x in logs
                      if str(x.get("level", "")).upper() in {"ERROR", "WARN"} and x.get("message")]
    if error_messages:
        return error_messages[-1]
    failed_tasks = [x.get("taskId", "") for x in execution.get("taskRunList", [])
                    if state_of(x) in {"FAILED", "KILLED"}]
    return "Tarea fallida: " + ", ".join(failed_tasks) if failed_tasks else state_of(execution)


def outcome_of(execution: dict[str, Any]) -> str:
    status = state_of(execution)
    action = str((execution.get("outputs") or {}).get("action") or "").lower()
    if status in {"FAILED", "KILLED"}:
        return "ERROR TÉCNICO"
    if action == "error":
        return "ERROR LÓGICO"
    if action == "rejected":
        return "RECHAZADO"
    if status == "SUCCESS":
        return "EXITOSO"
    return "OTRO"


def style_table(ws, freeze: str = "A2", filter_range: str | None = None,
                wrap_data: bool = False) -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    if filter_range:
        ws.auto_filter.ref = filter_range
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=wrap_data)
    for row_number in range(2, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 18


def autosize(ws, caps: dict[int, int] | None = None) -> None:
    caps = caps or {}
    for idx, column in enumerate(ws.columns, 1):
        width = max((len(str(c.value)) if c.value is not None else 0) for c in column) + 2
        ws.column_dimensions[get_column_letter(idx)].width = min(width, caps.get(idx, 45))


def build_workbook(executions: list[dict[str, Any]], logs_by_execution: dict[str, list[dict[str, Any]]],
                   namespace: str, flow_id: str, generated_at: datetime) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    states = Counter(state_of(x) for x in executions)
    outcomes = Counter(outcome_of(x) for x in executions)
    starts = [iso_value((x.get("state") or {}).get("startDate")) for x in executions]
    starts = [x for x in starts if isinstance(x, datetime)]
    success = states.get("SUCCESS", 0)
    failed = sum(states.get(x, 0) for x in ("FAILED", "KILLED"))
    ws.append(["AUDITORÍA DE FORMULARIO KESTRA", "Valor"])
    summary = [
        ("Namespace", namespace), ("Flow", flow_id), ("Generado", generated_at),
        ("Primera ejecución", min(starts) if starts else ""),
        ("Última ejecución", max(starts) if starts else ""),
        ("Total", len(executions)), ("Exitosas", success), ("Fallidas/Killed", failed),
        ("Tasa de éxito", success / len(executions) if executions else 0),
        ("Resultados exitosos/ingresados", outcomes.get("EXITOSO", 0)),
        ("Rechazados por reglas", outcomes.get("RECHAZADO", 0)),
        ("Errores lógicos/de datos", outcomes.get("ERROR LÓGICO", 0)),
        ("Errores técnicos", outcomes.get("ERROR TÉCNICO", 0)),
        ("Revisiones observadas", ", ".join(map(str, sorted({x.get('flowRevision') for x in executions})))),
    ]
    for row in summary:
        ws.append(row)
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == "Tasa de éxito":
            ws.cell(row, 2).number_format = "0.00%"
    style_table(ws, freeze="A2")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 48
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(bold=True, color=COLORS["navy"])
    for row in range(2, ws.max_row + 1):
        label = ws.cell(row, 1).value
        if label in {"Exitosas", "Resultados exitosos/ingresados"}:
            for cell in ws[row]: cell.fill = PatternFill("solid", fgColor=COLORS["green"])
        elif label in {"Fallidas/Killed", "Errores lógicos/de datos", "Errores técnicos"}:
            for cell in ws[row]: cell.fill = PatternFill("solid", fgColor=COLORS["red"])
        elif label == "Rechazados por reglas":
            for cell in ws[row]: cell.fill = PatternFill("solid", fgColor=COLORS["yellow"])

    detail = wb.create_sheet("Ejecuciones")
    fields = [
        "execution_id", "estado", "resultado", "inicio", "fin", "duracion", "revision", "ok", "action", "reason",
        "message", "contact_id", "lead_id", "cuil", "nombre", "email", "whatsapp", "provincia",
        "situacion_laboral", "banco_cobro", "origen_lead", "canal_envio", "landing", "landing_url",
        "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "tarea_fallida",
        "razon_error", "parametros_json", "headers_sanitizados_json", "body_json",
    ]
    detail.append(fields)
    error_rows = []
    daily: dict[str, Counter] = defaultdict(Counter)
    for item in sorted(executions, key=lambda x: str((x.get("state") or {}).get("startDate", ""))):
        trigger_vars = (item.get("trigger") or {}).get("variables") or {}
        body = trigger_vars.get("body") if isinstance(trigger_vars.get("body"), dict) else {}
        outputs = item.get("outputs") or {}
        state = item.get("state") or {}
        status = state_of(item)
        failed_tasks = [x.get("taskId", "") for x in item.get("taskRunList", []) if state_of(x) in {"FAILED", "KILLED"}]
        logs = logs_by_execution.get(item["id"], [])
        if status in {"FAILED", "KILLED"}:
            reason_error = error_summary(logs, item)
        elif str(outputs.get("action") or "").lower() == "error":
            reason_error = str(outputs.get("message") or outputs.get("reason") or "Error lógico sin detalle")
        else:
            reason_error = ""
        start = iso_value(state.get("startDate"))
        day = start.date().isoformat() if isinstance(start, datetime) else "Sin fecha"
        daily[day][status] += 1
        row = [
            item.get("id"), status, outcome_of(item), start, iso_value(state.get("endDate")), state.get("duration"),
            item.get("flowRevision"), outputs.get("ok"), outputs.get("action"), outputs.get("reason"),
            outputs.get("message"), outputs.get("contact_id"), outputs.get("lead_id"), body.get("cuil"),
            body.get("full_name"), body.get("email"), body.get("whatsapp"), body.get("province"),
            body.get("employment_status"), body.get("payment_bank"), body.get("lead_source"),
            body.get("submission_channel"), body.get("landing_slug"), body.get("landing_url"),
            body.get("utm_source"), body.get("utm_medium"), body.get("utm_campaign"), body.get("utm_term"),
            body.get("utm_content"), ", ".join(failed_tasks), reason_error,
            json_text(trigger_vars.get("parameters") or {}), json_text(safe_headers(trigger_vars.get("headers"))),
            json_text(body),
        ]
        detail.append(row)
        if outcome_of(item) in {"ERROR TÉCNICO", "ERROR LÓGICO"}:
            error_rows.append(row)
    style_table(detail, filter_range=f"A1:{get_column_letter(detail.max_column)}{detail.max_row}")
    detail.auto_filter.ref = f"A1:{get_column_letter(detail.max_column)}{detail.max_row}"
    detail.conditional_formatting.add(f"B2:B{detail.max_row}", FormulaRule(formula=["B2=\"SUCCESS\""], fill=PatternFill("solid", fgColor=COLORS["green"])))
    detail.conditional_formatting.add(f"B2:B{detail.max_row}", FormulaRule(formula=["OR(B2=\"FAILED\",B2=\"KILLED\")"], fill=PatternFill("solid", fgColor=COLORS["red"])))
    detail.conditional_formatting.add(f"C2:C{detail.max_row}", FormulaRule(formula=["C2=\"EXITOSO\""], fill=PatternFill("solid", fgColor=COLORS["green"])))
    detail.conditional_formatting.add(f"C2:C{detail.max_row}", FormulaRule(formula=["OR(C2=\"ERROR TÉCNICO\",C2=\"ERROR LÓGICO\")"], fill=PatternFill("solid", fgColor=COLORS["red"])))
    detail.conditional_formatting.add(f"C2:C{detail.max_row}", FormulaRule(formula=["C2=\"RECHAZADO\""], fill=PatternFill("solid", fgColor=COLORS["yellow"])))
    for r in range(2, detail.max_row + 1):
        detail.cell(r, 4).number_format = detail.cell(r, 5).number_format = "yyyy-mm-dd hh:mm:ss"
    autosize(detail, {1: 28, 11: 55, 24: 55, 31: 80, 32: 45, 33: 70, 34: 100})
    detail.column_dimensions["G"].width = 10
    detail.column_dimensions["H"].width = 9
    detail.column_dimensions["I"].width = 14
    # Raw request data remains available but hidden to keep the operational view compact.
    for column in range(32, 35):
        detail.column_dimensions[get_column_letter(column)].hidden = True

    errors = wb.create_sheet("Errores")
    errors.append(fields)
    for row in error_rows:
        errors.append(row)
    style_table(errors, filter_range=f"A1:{get_column_letter(errors.max_column)}{errors.max_row}")
    for row in errors.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="FFF2F2")
    autosize(errors, {1: 28, 11: 55, 31: 100, 33: 70, 34: 100})
    errors.column_dimensions["G"].width = 10
    errors.column_dimensions["H"].width = 9
    errors.column_dimensions["I"].width = 14
    for column in range(32, 35):
        errors.column_dimensions[get_column_letter(column)].hidden = True

    logs_ws = wb.create_sheet("Logs errores")
    log_fields = ["execution_id", "timestamp", "nivel", "task_id", "task_run_id", "intento", "mensaje"]
    logs_ws.append(log_fields)
    for execution_id, logs in logs_by_execution.items():
        for log in logs:
            logs_ws.append([execution_id, iso_value(log.get("timestamp")), log.get("level"), log.get("taskId"),
                            log.get("taskRunId"), log.get("attemptNumber"), log.get("message")])
    style_table(logs_ws, filter_range=f"A1:G{logs_ws.max_row}")
    logs_ws.conditional_formatting.add(f"C2:C{logs_ws.max_row}", FormulaRule(formula=["C2=\"ERROR\""], fill=PatternFill("solid", fgColor=COLORS["red"])))
    autosize(logs_ws, {1: 28, 7: 120})

    daily_ws = wb.create_sheet("Análisis diario")
    daily_ws.append(["Fecha", "Total", "Exitosas técnicas", "Errores técnicos", "Errores lógicos", "Rechazados", "Tasa éxito técnica"])
    for day, counts in sorted(daily.items()):
        total = sum(counts.values()); ok = counts.get("SUCCESS", 0); bad = counts.get("FAILED", 0) + counts.get("KILLED", 0)
        day_items = [x for x in executions if (iso_value((x.get("state") or {}).get("startDate")).date().isoformat() if isinstance(iso_value((x.get("state") or {}).get("startDate")), datetime) else "Sin fecha") == day]
        day_outcomes = Counter(outcome_of(x) for x in day_items)
        daily_ws.append([day, total, ok, bad, day_outcomes.get("ERROR LÓGICO", 0), day_outcomes.get("RECHAZADO", 0), ok / total if total else 0])
    style_table(daily_ws, filter_range=f"A1:G{daily_ws.max_row}")
    for r in range(2, daily_ws.max_row + 1): daily_ws.cell(r, 7).number_format = "0.00%"
    autosize(daily_ws)
    if daily_ws.max_row > 1:
        chart = BarChart(); chart.title = "Ejecuciones por día"; chart.y_axis.title = "Cantidad"
        chart.add_data(Reference(daily_ws, min_col=3, max_col=6, min_row=1, max_row=daily_ws.max_row), titles_from_data=True)
        chart.set_categories(Reference(daily_ws, min_col=1, min_row=2, max_row=daily_ws.max_row)); chart.height = 8; chart.width = 18
        daily_ws.add_chart(chart, "H2")

    states_ws = wb.create_sheet("Estados")
    states_ws.append(["Estado", "Cantidad"])
    for state, count in states.most_common(): states_ws.append([state, count])
    style_table(states_ws)
    pie = PieChart(); pie.title = "Distribución de estados"
    pie.add_data(Reference(states_ws, min_col=2, min_row=1, max_row=states_ws.max_row), titles_from_data=True)
    pie.set_categories(Reference(states_ws, min_col=1, min_row=2, max_row=states_ws.max_row)); states_ws.add_chart(pie, "D2")
    autosize(states_ws)

    notes = wb.create_sheet("Notas")
    notes.append(["Campo", "Descripción"])
    notes_rows = [
        ("Alcance", "Todas las ejecuciones devueltas por la API para namespace y flow indicados."),
        ("Inputs", "Campos del trigger.body aplanados y body_json completo."),
        ("Headers", "Se excluyen Authorization, Cookie, Set-Cookie, X-API-Key y Proxy-Authorization."),
        ("Errores", "Razón tomada del último log ERROR/WARN; en su defecto, de la tarea fallida."),
        ("Logs", "Se descargan íntegramente únicamente para ejecuciones FAILED/KILLED."),
        ("Privacidad", "El archivo contiene datos personales y debe manejarse con acceso restringido."),
    ]
    for row in notes_rows: notes.append(row)
    style_table(notes, wrap_data=True)
    notes.column_dimensions["A"].width = 22; notes.column_dimensions["B"].width = 110
    return wb


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--namespace", default="redunisol.prod.marketing-crm")
    parser.add_argument("--flow-id", default="bitrix24_form_webhook")
    parser.add_argument("--tenant", default=os.getenv("KESTRA_TENANT", "main"))
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    base_url = os.environ["KESTRA_URL"].rstrip("/")
    session = requests.Session()
    session.auth = (os.environ["KESTRA_USERNAME"], os.environ["KESTRA_PASSWORD"])
    executions = fetch_executions(session, base_url, args.tenant, args.namespace, args.flow_id)
    failed = [x for x in executions if state_of(x) in {"FAILED", "KILLED"}]
    logs = {x["id"]: fetch_logs(session, base_url, args.tenant, x["id"]) for x in failed}
    workbook = build_workbook(executions, logs, args.namespace, args.flow_id, datetime.now())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(json.dumps({"output": str(args.output), "total": len(executions), "failed": len(failed)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
